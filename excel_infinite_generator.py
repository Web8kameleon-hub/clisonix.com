#!/usr/bin/env python3
"""
🔷 EXCEL ∞ INFINITE GENERATOR
==============================
Clisonix Deterministic SaaS Protocol

Gjeneron Excel dashboard nga API lista me:
- Ngjyrosje automatike sipas folders
- Status colors (READY/PENDING/ERROR)
- cURL dhe Python snippets
- VBA-ready për pop-up notifications

Usage:
    python excel_infinite_generator.py --input apis.json --output Clisonix_API_Dashboard.xlsx
"""

import json
import argparse
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️  openpyxl not installed. Run: pip install openpyxl")


# ============================================================================
# CONFIGURATION - NGJYRAT DHE STILET
# ============================================================================

FOLDER_COLORS = {
    "📋 Health & Status": "CCFFCC",
    "🎤 Ask & Neural Symphony": "FFE699",
    "📤 Uploads & Processing": "CCE5FF",
    "💳 Billing": "FFC7CE",
    "🤖 ASI Trinity": "D9EAD3",
    "🧠 Brain Engine": "FFF2CC",
    "📊 ALBA Data Collection": "D9D2E9",
    "🏭 Industrial Dashboard": "EAD1DC",
    "📡 Signal Generation": "E2EFDA",
    "🤖 ASI Real Metrics": "FDE9D9",
    "🤖 OpenAI Neural": "F4CCCC",
    "🌐 External APIs": "D9D9D9",
    "🖥️ Frontend Modules": "DDEBF7",
    "🔧 External Tools": "FFF2CC",
    "🔬 Protocol Kitchen": "E6E6FA",
    "📈 Excel Dashboard": "98FB98",
}

STATUS_COLORS = {
    "READY": "92D050",      # Green
    "APPROVED": "92D050",   # Green
    "BAKED": "92D050",      # Green
    "PENDING": "FFC000",    # Orange
    "TESTING": "FFFF00",    # Yellow
    "ERROR": "FF0000",      # Red
    "DEPRECATED": "808080", # Gray
}

METHOD_COLORS = {
    "GET": "61AFFE",
    "POST": "49CC90",
    "PUT": "FCA130",
    "DELETE": "F93E3E",
    "PATCH": "50E3C2",
}

HEADER_STYLE = {
    "font": Font(bold=True, color="FFFFFF", size=11),
    "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
    "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
    "border": Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
}

# ============================================================================
# HEADERS - STRUKTURA E TABELËS
# ============================================================================

HEADERS = [
    "Row_ID",
    "Folder",
    "Method", 
    "Endpoint",
    "Description",
    "Sample_Response",
    "cURL",
    "Python_Snippet",
    "Status",
    "Version",
    "Security",
    "ML_Confidence",
    "Last_Updated"
]

COLUMN_WIDTHS = {
    "Row_ID": 8,
    "Folder": 25,
    "Method": 10,
    "Endpoint": 35,
    "Description": 40,
    "Sample_Response": 50,
    "cURL": 60,
    "Python_Snippet": 60,
    "Status": 12,
    "Version": 10,
    "Security": 10,
    "ML_Confidence": 15,
    "Last_Updated": 20
}


# ============================================================================
# GENERATOR FUNCTIONS
# ============================================================================

def generate_curl(method: str, endpoint: str, base_url: str = "https://clisonix.com") -> str:
    """Gjeneron cURL command për endpoint"""
    if method.upper() == "GET":
        return f'curl -X GET "{base_url}{endpoint}" -H "Authorization: Bearer $API_KEY"'
    elif method.upper() == "POST":
        return f'curl -X POST "{base_url}{endpoint}" -H "Authorization: Bearer $API_KEY" -H "Content-Type: application/json" -d \'{{"data": "example"}}\''
    elif method.upper() == "PUT":
        return f'curl -X PUT "{base_url}{endpoint}" -H "Authorization: Bearer $API_KEY" -H "Content-Type: application/json" -d \'{{"data": "update"}}\''
    elif method.upper() == "DELETE":
        return f'curl -X DELETE "{base_url}{endpoint}" -H "Authorization: Bearer $API_KEY"'
    else:
        return f'curl -X {method.upper()} "{base_url}{endpoint}"'


def generate_python_snippet(method: str, endpoint: str, base_url: str = "https://clisonix.com") -> str:
    """Gjeneron Python snippet për endpoint"""
    return f'''import requests

response = requests.{method.lower()}(
    "{base_url}{endpoint}",
    headers={{"Authorization": "Bearer API_KEY"}}
)
print(response.json())'''


def create_excel_infinite(
    apis: List[Dict[str, Any]], 
    output_path: str = "Clisonix_API_Dashboard.xlsx",
    base_url: str = "https://clisonix.com"
) -> str:
    """
    Krijon Excel ∞ Dashboard nga lista e API-ve
    
    Args:
        apis: Lista e API-ve me keys: folder, method, endpoint, description, status, etc.
        output_path: Path për output file
        base_url: Base URL për cURL dhe snippets
    
    Returns:
        Path i file të krijuar
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "API Catalog"
    
    # === HEADER ROW ===
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_STYLE["font"]
        cell.fill = HEADER_STYLE["fill"]
        cell.alignment = HEADER_STYLE["alignment"]
        cell.border = HEADER_STYLE["border"]
        
        # Set column width
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS.get(header, 15)
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # === DATA ROWS ===
    for row_idx, api in enumerate(apis, 2):
        row_id = row_idx - 1
        folder = api.get("folder", "📋 Health & Status")
        method = api.get("method", "GET").upper()
        endpoint = api.get("endpoint", "/")
        description = api.get("description", "")
        sample_response = api.get("sample_response", '{"status": "ok"}')
        status = api.get("status", "PENDING").upper()
        version = api.get("version", "v1")
        security = api.get("security", "PASS")
        ml_confidence = api.get("ml_confidence", 95)
        last_updated = api.get("last_updated", datetime.now().strftime("%Y-%m-%d %H:%M"))
        
        # Generate snippets
        curl_cmd = generate_curl(method, endpoint, base_url)
        python_code = generate_python_snippet(method, endpoint, base_url)
        
        # Row data
        row_data = [
            row_id,
            folder,
            method,
            endpoint,
            description,
            sample_response if isinstance(sample_response, str) else json.dumps(sample_response),
            curl_cmd,
            python_code,
            status,
            version,
            security,
            f"{ml_confidence}%",
            last_updated
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(
                left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC")
            )
            
            # === NGJYROSJE ===
            
            # Folder color
            if col_idx == 2:  # Folder column
                color = FOLDER_COLORS.get(folder, "FFFFFF")
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            
            # Method color
            if col_idx == 3:  # Method column
                color = METHOD_COLORS.get(method, "CCCCCC")
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            # Status color
            if col_idx == 9:  # Status column
                color = STATUS_COLORS.get(status, "CCCCCC")
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                if status in ["READY", "APPROVED", "BAKED"]:
                    cell.font = Font(bold=True, color="FFFFFF")
                elif status == "ERROR":
                    cell.font = Font(bold=True, color="FFFFFF")
            
            # Security color
            if col_idx == 11:  # Security column
                if security == "PASS":
                    cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
                elif security == "FAIL":
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    cell.font = Font(color="FFFFFF")
    
    # === METADATA SHEET ===
    meta_ws = wb.create_sheet("Metadata")
    meta_data = [
        ["Property", "Value"],
        ["Generated", datetime.now().isoformat()],
        ["Total APIs", len(apis)],
        ["Ready Count", sum(1 for a in apis if a.get("status", "").upper() in ["READY", "APPROVED", "BAKED"])],
        ["Pending Count", sum(1 for a in apis if a.get("status", "").upper() == "PENDING")],
        ["Base URL", base_url],
        ["Generator", "Clisonix Excel ∞ Infinite Generator"],
        ["Version", "1.0.0"]
    ]
    for row_idx, row in enumerate(meta_data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = meta_ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True)
    
    # === SAVE ===
    wb.save(output_path)
    print(f"✅ Excel ∞ Dashboard created: {output_path}")
    print(f"   📊 Total APIs: {len(apis)}")
    print(f"   🟢 Ready: {sum(1 for a in apis if a.get('status', '').upper() in ['READY', 'APPROVED', 'BAKED'])}")
    
    return output_path


def load_apis_from_json(json_path: str) -> List[Dict[str, Any]]:
    """Ngarkon API lista nga JSON file"""
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Handle different JSON structures
    if isinstance(data, list):
        return data
    elif isinstance(data, dict):
        if "apis" in data:
            return data["apis"]
        elif "item" in data:  # Postman collection format
            return extract_from_postman(data)
    
    return []


def extract_from_postman(collection: Dict) -> List[Dict[str, Any]]:
    """Ekstrakton API-të nga Postman collection"""
    apis = []
    
    def process_items(items, folder_name=""):
        for item in items:
            if "item" in item:  # It's a folder
                process_items(item["item"], item.get("name", folder_name))
            elif "request" in item:  # It's a request
                request = item["request"]
                method = request.get("method", "GET") if isinstance(request, dict) else "GET"
                
                # Get URL
                url = request.get("url", {}) if isinstance(request, dict) else request
                if isinstance(url, dict):
                    endpoint = "/" + "/".join(url.get("path", []))
                elif isinstance(url, str):
                    endpoint = url.split("clisonix.com")[-1] if "clisonix.com" in url else url
                else:
                    endpoint = "/"
                
                apis.append({
                    "folder": f"📁 {folder_name}" if folder_name else "📋 General",
                    "method": method,
                    "endpoint": endpoint,
                    "description": item.get("name", ""),
                    "status": "READY",
                    "version": "v1",
                    "security": "PASS",
                    "ml_confidence": 95
                })
    
    if "item" in collection:
        process_items(collection["item"])
    
    return apis


# ============================================================================
# SAMPLE API DATA (për demo)
# ============================================================================

SAMPLE_APIS = [
    {"folder": "📋 Health & Status", "method": "GET", "endpoint": "/health", "description": "System health check", "status": "READY", "sample_response": '{"status": "healthy"}'},
    {"folder": "📋 Health & Status", "method": "GET", "endpoint": "/status", "description": "Full system status", "status": "READY", "sample_response": '{"status": "operational"}'},
    {"folder": "🤖 ASI Trinity", "method": "GET", "endpoint": "/asi/status", "description": "ASI Trinity architecture status", "status": "READY", "sample_response": '{"trinity": {"alba": {}, "albi": {}, "jona": {}}}'},
    {"folder": "🤖 ASI Trinity", "method": "GET", "endpoint": "/asi/alba/metrics", "description": "Alba network metrics", "status": "READY"},
    {"folder": "🤖 ASI Trinity", "method": "GET", "endpoint": "/asi/albi/metrics", "description": "Albi neural processor metrics", "status": "READY"},
    {"folder": "🤖 ASI Trinity", "method": "GET", "endpoint": "/asi/jona/metrics", "description": "Jona coordinator metrics", "status": "READY"},
    {"folder": "🔬 Protocol Kitchen", "method": "GET", "endpoint": "/api/kitchen/status", "description": "Protocol Kitchen status", "status": "READY"},
    {"folder": "🔬 Protocol Kitchen", "method": "GET", "endpoint": "/api/kitchen/layers", "description": "All pipeline layers", "status": "READY"},
    {"folder": "🔬 Protocol Kitchen", "method": "POST", "endpoint": "/api/kitchen/intake", "description": "Process intake through pipeline", "status": "READY"},
    {"folder": "📈 Excel Dashboard", "method": "GET", "endpoint": "/api/excel/status", "description": "Excel module status", "status": "READY"},
    {"folder": "📈 Excel Dashboard", "method": "GET", "endpoint": "/api/excel/dashboards", "description": "List all Excel dashboards", "status": "READY"},
    {"folder": "📈 Excel Dashboard", "method": "GET", "endpoint": "/api/excel/summary", "description": "Summary of all dashboards", "status": "READY"},
    {"folder": "🎤 Ask & Neural Symphony", "method": "POST", "endpoint": "/api/ask", "description": "AI Q&A endpoint", "status": "READY"},
    {"folder": "📤 Uploads & Processing", "method": "POST", "endpoint": "/api/uploads/eeg/process", "description": "EEG file processing", "status": "READY"},
    {"folder": "📤 Uploads & Processing", "method": "POST", "endpoint": "/api/uploads/audio/process", "description": "Audio file processing", "status": "READY"},
    {"folder": "💳 Billing", "method": "POST", "endpoint": "/billing/paypal/order", "description": "Create PayPal order", "status": "READY"},
    {"folder": "💳 Billing", "method": "POST", "endpoint": "/billing/stripe/payment-intent", "description": "Create Stripe payment intent", "status": "READY"},
    {"folder": "🧠 Brain Engine", "method": "GET", "endpoint": "/brain/youtube/insight", "description": "YouTube insight generator", "status": "READY"},
    {"folder": "🧠 Brain Engine", "method": "POST", "endpoint": "/brain/energy/check", "description": "Daily energy check", "status": "READY"},
    {"folder": "🏭 Industrial Dashboard", "method": "GET", "endpoint": "/industrial/metrics", "description": "Industrial metrics", "status": "READY"},
    {"folder": "🌐 External APIs", "method": "GET", "endpoint": "/api/crypto/market", "description": "Crypto market data", "status": "READY"},
    {"folder": "🌐 External APIs", "method": "GET", "endpoint": "/api/weather", "description": "Weather data", "status": "READY"},
]


# ============================================================================
# MAIN
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description="Clisonix Excel ∞ Infinite Generator")
    parser.add_argument("--input", "-i", help="Input JSON file with APIs")
    parser.add_argument("--output", "-o", default="Clisonix_API_Infinite_Dashboard.xlsx", help="Output Excel file")
    parser.add_argument("--base-url", default="https://clisonix.com", help="Base URL for API endpoints")
    parser.add_argument("--demo", action="store_true", help="Generate demo dashboard with sample APIs")
    
    args = parser.parse_args()
    
    if args.demo or not args.input:
        print("🔷 Generating demo Excel ∞ Dashboard...")
        apis = SAMPLE_APIS
    else:
        print(f"📂 Loading APIs from: {args.input}")
        apis = load_apis_from_json(args.input)
    
    create_excel_infinite(apis, args.output, args.base_url)


if __name__ == "__main__":
    main()
