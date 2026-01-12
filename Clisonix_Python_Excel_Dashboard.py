"""
🎯 CLISONIX PYTHON + EXCEL INTEGRATION
Krijon Excel workbook me Python formulas që ekzekutohen brenda Excel

KËRKON: Microsoft 365 me Python in Excel (Preview/Beta feature)

Ky script krijon një Excel file me:
1. Data tables
2. Python formulas në qeliza (=PY(...))
3. Charts të gjeneruara nga Python
4. Statistical analysis
"""

import json
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False
    print("pip install openpyxl")

# ============================================================
# PYTHON IN EXCEL FORMULAS
# These are examples of =PY() formulas for Excel 365
# ============================================================

PYTHON_EXCEL_FORMULAS = {
    "stats_summary": '''=PY("
import pandas as pd
# Merr data nga tabela Excel
df = xl('A1:E100', headers=True)
# Statistika
stats = df.describe()
stats
")''',

    "histogram": '''=PY("
import matplotlib.pyplot as plt
import pandas as pd
df = xl('A1:E100', headers=True)
fig, ax = plt.subplots(figsize=(8,5))
df['Status'].value_counts().plot(kind='bar', ax=ax, color=['green','red','orange'])
ax.set_title('Status Distribution')
fig
")''',

    "pie_chart": '''=PY("
import matplotlib.pyplot as plt
import pandas as pd
df = xl('A1:E100', headers=True)
fig, ax = plt.subplots(figsize=(6,6))
df['Maturity'].value_counts().plot(kind='pie', ax=ax, autopct='%1.1f%%')
ax.set_title('Maturity Levels')
fig
")''',

    "correlation": '''=PY("
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
df = xl('A1:E100', headers=True)
numeric_cols = df.select_dtypes(include='number')
fig, ax = plt.subplots(figsize=(8,6))
sns.heatmap(numeric_cols.corr(), annot=True, cmap='coolwarm', ax=ax)
fig
")''',

    "ml_prediction": '''=PY("
import pandas as pd
from sklearn.ensemble import RandomForestClassifier
df = xl('A1:E100', headers=True)
# Prepare data
X = df[['ML_Confidence', 'Score']].fillna(0)
y = df['Status'].map({'PASS': 1, 'FAIL': 0}).fillna(0)
# Train model
model = RandomForestClassifier(n_estimators=100)
model.fit(X, y)
# Predict
predictions = model.predict(X)
pd.DataFrame({'Prediction': predictions})
")''',

    "time_series": '''=PY("
import pandas as pd
import matplotlib.pyplot as plt
df = xl('A1:E100', headers=True)
df['Timestamp'] = pd.to_datetime(df['Timestamp'])
fig, ax = plt.subplots(figsize=(10,5))
df.set_index('Timestamp')['ML_Confidence'].plot(ax=ax)
ax.set_title('ML Confidence Over Time')
fig
")''',
}


def create_python_excel_workbook(output_path: str = None):
    """
    Krijon Excel workbook me shembuj të Python formulas.
    """
    if not OPENPYXL_OK:
        return None
    
    base_path = Path(__file__).parent
    output_path = output_path or base_path / "Clisonix_Python_Dashboard.xlsx"
    
    wb = Workbook()
    
    # ========== SHEET 1: Sample Data ==========
    ws_data = wb.active
    ws_data.title = "Data"
    
    # Headers
    headers = ['ID', 'Endpoint', 'Status', 'Maturity', 'ML_Confidence', 'Score', 'Timestamp']
    for col, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    # Sample data
    sample_data = [
        [1, '/health', 'PASS', 'Production', 0.95, 98, '2026-01-11 10:00'],
        [2, '/status', 'PASS', 'Production', 0.92, 95, '2026-01-11 10:01'],
        [3, '/api/alba/health', 'PASS', 'Stable', 0.88, 90, '2026-01-11 10:02'],
        [4, '/api/albi/neural', 'FAIL', 'Beta', 0.45, 60, '2026-01-11 10:03'],
        [5, '/asi/status', 'PASS', 'Production', 0.91, 94, '2026-01-11 10:04'],
        [6, '/brain/analyze', 'PASS', 'Stable', 0.87, 89, '2026-01-11 10:05'],
        [7, '/billing/plans', 'PASS', 'Production', 0.93, 96, '2026-01-11 10:06'],
        [8, '/jona/coordinate', 'FAIL', 'Beta', 0.52, 65, '2026-01-11 10:07'],
        [9, '/api/metrics', 'PASS', 'Stable', 0.89, 91, '2026-01-11 10:08'],
        [10, '/stream/audio', 'PASS', 'Production', 0.94, 97, '2026-01-11 10:09'],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_data.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True)
            
            # Color coding for Status
            if col_idx == 3:  # Status column
                if value == 'PASS':
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                elif value == 'FAIL':
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    # Column widths
    ws_data.column_dimensions['A'].width = 8
    ws_data.column_dimensions['B'].width = 25
    ws_data.column_dimensions['C'].width = 12
    ws_data.column_dimensions['D'].width = 15
    ws_data.column_dimensions['E'].width = 15
    ws_data.column_dimensions['F'].width = 10
    ws_data.column_dimensions['G'].width = 20
    
    # ========== SHEET 2: Python Formulas Guide ==========
    ws_py = wb.create_sheet("Python_Formulas")
    
    ws_py['A1'] = "🐍 PYTHON IN EXCEL - Formula Guide"
    ws_py['A1'].font = Font(bold=True, size=16, color='1F4E79')
    ws_py.merge_cells('A1:D1')
    
    ws_py['A3'] = "Këto formula funksionojnë në Excel 365 me Python in Excel të aktivizuar."
    ws_py['A3'].font = Font(italic=True)
    ws_py.merge_cells('A3:D3')
    
    # Formula examples
    formulas = [
        ("Statistics Summary", PYTHON_EXCEL_FORMULAS['stats_summary'], "Llogarit statistika për të dhënat"),
        ("Histogram Chart", PYTHON_EXCEL_FORMULAS['histogram'], "Krijon histogram për Status"),
        ("Pie Chart", PYTHON_EXCEL_FORMULAS['pie_chart'], "Krijon pie chart për Maturity"),
        ("Correlation Heatmap", PYTHON_EXCEL_FORMULAS['correlation'], "Tregon korrelacionin mes kolonave"),
        ("ML Prediction", PYTHON_EXCEL_FORMULAS['ml_prediction'], "Random Forest prediction"),
        ("Time Series", PYTHON_EXCEL_FORMULAS['time_series'], "Grafik i ML Confidence në kohë"),
    ]
    
    row = 5
    ws_py.cell(row=row, column=1, value="Emri").font = Font(bold=True)
    ws_py.cell(row=row, column=2, value="Formula").font = Font(bold=True)
    ws_py.cell(row=row, column=3, value="Përshkrimi").font = Font(bold=True)
    
    for name, formula, desc in formulas:
        row += 1
        ws_py.cell(row=row, column=1, value=name)
        ws_py.cell(row=row, column=2, value=formula)
        ws_py.cell(row=row, column=3, value=desc)
        ws_py.row_dimensions[row].height = 80
    
    ws_py.column_dimensions['A'].width = 25
    ws_py.column_dimensions['B'].width = 80
    ws_py.column_dimensions['C'].width = 40
    
    # Wrap text for formula column
    for row in ws_py.iter_rows(min_row=6, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # ========== SHEET 3: Quick Analysis ==========
    ws_analysis = wb.create_sheet("Quick_Analysis")
    
    ws_analysis['A1'] = "📊 QUICK ANALYSIS"
    ws_analysis['A1'].font = Font(bold=True, size=14)
    
    # Status counts (manually calculated for demo)
    ws_analysis['A3'] = "Status Distribution"
    ws_analysis['A3'].font = Font(bold=True)
    ws_analysis['A4'] = "PASS"
    ws_analysis['B4'] = 8
    ws_analysis['A5'] = "FAIL"
    ws_analysis['B5'] = 2
    
    # Create bar chart
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Status Distribution"
    data = Reference(ws_analysis, min_col=2, min_row=4, max_row=5)
    cats = Reference(ws_analysis, min_col=1, min_row=4, max_row=5)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    chart.shape = 4
    ws_analysis.add_chart(chart, "D3")
    
    # Maturity counts
    ws_analysis['A8'] = "Maturity Levels"
    ws_analysis['A8'].font = Font(bold=True)
    ws_analysis['A9'] = "Production"
    ws_analysis['B9'] = 4
    ws_analysis['A10'] = "Stable"
    ws_analysis['B10'] = 3
    ws_analysis['A11'] = "Beta"
    ws_analysis['B11'] = 2
    
    # Pie chart
    pie = PieChart()
    pie.title = "Maturity Distribution"
    data = Reference(ws_analysis, min_col=2, min_row=9, max_row=11)
    cats = Reference(ws_analysis, min_col=1, min_row=9, max_row=11)
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(cats)
    ws_analysis.add_chart(pie, "D12")
    
    # ========== SHEET 4: Instructions ==========
    ws_help = wb.create_sheet("How_To_Use")
    
    instructions = [
        "🎯 SI TË PËRDORËSH PYTHON NË EXCEL",
        "",
        "1️⃣ AKTIVIZO PYTHON IN EXCEL:",
        "   • Duhet Microsoft 365 (jo free version)",
        "   • File → Options → General → Enable Python (Preview)",
        "   • Ose: Insert → Python → Trust Python",
        "",
        "2️⃣ SHKRUAJ FORMULA PYTHON:",
        "   • Në një qelizë shkruaj: =PY(\"kod python\")",
        "   • Ose: Ctrl+Shift+P për Python cell",
        "",
        "3️⃣ PËRDOR XL() PËR TË MARRË DATA:",
        "   • xl('A1:E10') - merr range si DataFrame",
        "   • xl('Sheet1!A1:B5', headers=True)",
        "",
        "4️⃣ SHFAQ REZULTATE:",
        "   • Return DataFrame → shfaqet si tabelë",
        "   • Return Figure → shfaqet si imazh",
        "   • Return vlerë → shfaqet në qelizë",
        "",
        "5️⃣ BIBLIOTEKA TË DISPONUESHME:",
        "   • pandas, numpy, matplotlib, seaborn",
        "   • scikit-learn, scipy, statsmodels",
        "   • Dhe shumë të tjera!",
        "",
        "📌 SHEMBULL:",
        "   =PY(\"",
        "   import pandas as pd",
        "   df = xl('Data!A1:G11', headers=True)",
        "   df.describe()",
        "   \")",
    ]
    
    for idx, line in enumerate(instructions, 1):
        ws_help.cell(row=idx, column=1, value=line)
        if line.startswith("🎯") or line.startswith("1️⃣") or line.startswith("📌"):
            ws_help.cell(row=idx, column=1).font = Font(bold=True, size=12)
    
    ws_help.column_dimensions['A'].width = 60
    
    # Save
    wb.save(output_path)
    print(f"✅ Python+Excel Dashboard created: {output_path}")
    
    return str(output_path)


def create_api_analysis_workbook(api_data_path: str = None, output_path: str = None):
    """
    Krijon workbook me analiza API duke përdorur Python formulas.
    """
    if not OPENPYXL_OK:
        return None
    
    base_path = Path(__file__).parent
    output_path = output_path or base_path / "Clisonix_API_Analysis.xlsx"
    
    # Load API data if available
    api_data = []
    if api_data_path and Path(api_data_path).exists():
        try:
            with open(api_data_path, 'r', encoding='utf-8') as f:
                api_data = json.load(f)
        except:
            pass
    
    wb = Workbook()
    ws = wb.active
    ws.title = "API_Endpoints"
    
    # Headers
    headers = ['Endpoint', 'Method', 'Description', 'Sample_Response', 'cURL', 'Python_Code']
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Column widths - wrap text handles display
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 60
    ws.column_dimensions['F'].width = 60
    
    # Sample API endpoints
    endpoints = [
        {
            'endpoint': '/health',
            'method': 'GET',
            'description': 'Kontrollon gjendjen e shëndetit të sistemit Clisonix Cloud',
            'sample': '{"status": "healthy", "version": "2.1.0", "uptime": 86400}',
            'curl': 'curl -X GET "https://api.clisonix.com/health" -H "Authorization: Bearer $TOKEN"',
            'python': 'response = requests.get("https://api.clisonix.com/health", headers={"Authorization": f"Bearer {TOKEN}"})'
        },
        {
            'endpoint': '/api/alba/health',
            'method': 'GET',
            'description': 'Status i ALBA Network shërbimit për audio streaming',
            'sample': '{"service": "alba", "status": "active", "connections": 42}',
            'curl': 'curl -X GET "https://api.clisonix.com/api/alba/health"',
            'python': 'response = requests.get("https://api.clisonix.com/api/alba/health")'
        },
        {
            'endpoint': '/asi/trinity/status',
            'method': 'GET',
            'description': 'ASI Trinity Engine status dhe metrika',
            'sample': '{"trinity": "active", "goroutines": 156, "memory_mb": 512}',
            'curl': 'curl -X GET "https://api.clisonix.com/asi/trinity/status"',
            'python': 'response = requests.get("https://api.clisonix.com/asi/trinity/status")'
        },
    ]
    
    # Add data
    cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    code_font = Font(name='Consolas', size=10)
    
    for row_idx, ep in enumerate(endpoints, 2):
        ws.cell(row=row_idx, column=1, value=ep['endpoint']).alignment = cell_alignment
        ws.cell(row=row_idx, column=2, value=ep['method']).alignment = cell_alignment
        ws.cell(row=row_idx, column=3, value=ep['description']).alignment = cell_alignment
        
        sample_cell = ws.cell(row=row_idx, column=4, value=ep['sample'])
        sample_cell.alignment = cell_alignment
        sample_cell.font = code_font
        
        curl_cell = ws.cell(row=row_idx, column=5, value=ep['curl'])
        curl_cell.alignment = cell_alignment
        curl_cell.font = code_font
        
        python_cell = ws.cell(row=row_idx, column=6, value=ep['python'])
        python_cell.alignment = cell_alignment
        python_cell.font = code_font
    
    # Freeze header
    ws.freeze_panes = 'A2'
    
    # Save
    wb.save(output_path)
    print(f"✅ API Analysis workbook created: {output_path}")
    
    return str(output_path)


# ============================================================
# MAIN
# ============================================================

if __name__ == "__main__":
    print("🎯 Clisonix Python + Excel Dashboard Creator")
    print("=" * 50)
    
    # Create Python Excel dashboard
    path1 = create_python_excel_workbook()
    
    # Create API analysis workbook  
    path2 = create_api_analysis_workbook()
    
    print()
    print("📂 Files created:")
    print(f"   1. {path1}")
    print(f"   2. {path2}")
    print()
    print("💡 Hap këto file në Excel 365 dhe aktivizo Python in Excel")
    print("   për të përdorur =PY() formulas!")
