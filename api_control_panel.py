"""
🔴 LIVE API CONTROL PANEL - Me butona START/STOP
Dashboard me GUI - Nuk krijon Excel të reja, përditëson të njëjtën!
"""
import tkinter as tk
from tkinter import ttk, messagebox
import requests
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
import threading
import os

# Config
EXCEL_PATH = 'c:/Users/Admin/Desktop/LIVE_API_DASHBOARD.xlsx'
BASE_URL = 'https://clisonix.com'

class LiveDashboardApp:
    def __init__(self, root):
        self.root = root
        self.root.title("🔴 Clisonix LIVE API Monitor")
        self.root.geometry("900x600")
        self.root.configure(bg='#1a1a2e')
        
        self.is_running = False
        self.refresh_interval = 30
        self.monitor_thread = None
        
        # Styles
        self.online_color = '#00ff88'
        self.offline_color = '#ff4444'
        self.degraded_color = '#ffaa00'
        
        self.setup_ui()
        
    def setup_ui(self):
        # Title
        title_frame = tk.Frame(self.root, bg='#1a1a2e')
        title_frame.pack(fill='x', pady=10)
        
        self.status_indicator = tk.Label(title_frame, text="⚫", font=('Arial', 24), bg='#1a1a2e', fg='gray')
        self.status_indicator.pack(side='left', padx=20)
        
        tk.Label(title_frame, text="Clisonix API Monitor", font=('Arial', 20, 'bold'), 
                 bg='#1a1a2e', fg='white').pack(side='left')
        
        self.status_label = tk.Label(title_frame, text="STOPPED", font=('Arial', 14, 'bold'),
                                     bg='#1a1a2e', fg='gray')
        self.status_label.pack(side='right', padx=20)
        
        # Control Panel
        control_frame = tk.Frame(self.root, bg='#16213e', pady=15)
        control_frame.pack(fill='x', padx=20)
        
        # START Button
        self.start_btn = tk.Button(control_frame, text="▶ START", font=('Arial', 14, 'bold'),
                                   bg='#00aa55', fg='white', width=15, height=2,
                                   command=self.start_monitoring, cursor='hand2')
        self.start_btn.pack(side='left', padx=20)
        
        # STOP Button
        self.stop_btn = tk.Button(control_frame, text="⏹ STOP", font=('Arial', 14, 'bold'),
                                  bg='#cc3333', fg='white', width=15, height=2,
                                  command=self.stop_monitoring, cursor='hand2', state='disabled')
        self.stop_btn.pack(side='left', padx=20)
        
        # Refresh interval
        tk.Label(control_frame, text="Refresh (s):", font=('Arial', 11),
                 bg='#16213e', fg='white').pack(side='left', padx=10)
        
        self.interval_var = tk.StringVar(value="30")
        interval_spin = tk.Spinbox(control_frame, from_=5, to=300, width=5,
                                   textvariable=self.interval_var, font=('Arial', 12))
        interval_spin.pack(side='left')
        
        # Open Excel Button
        tk.Button(control_frame, text="📊 Open Excel", font=('Arial', 11),
                  bg='#2d6a4f', fg='white', command=self.open_excel,
                  cursor='hand2').pack(side='right', padx=20)
        
        # Last update
        self.last_update_label = tk.Label(control_frame, text="Last: --:--:--", 
                                          font=('Arial', 10), bg='#16213e', fg='#888')
        self.last_update_label.pack(side='right', padx=10)
        
        # API Status Table
        table_frame = tk.Frame(self.root, bg='#1a1a2e')
        table_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        # Treeview
        columns = ('status', 'name', 'endpoint', 'response', 'last_check')
        self.tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=10)
        
        # Style treeview
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('Treeview', background='#1a1a2e', foreground='white',
                        fieldbackground='#1a1a2e', font=('Arial', 10))
        style.configure('Treeview.Heading', background='#16213e', foreground='white',
                        font=('Arial', 11, 'bold'))
        
        self.tree.heading('status', text='Status')
        self.tree.heading('name', text='API Name')
        self.tree.heading('endpoint', text='Endpoint')
        self.tree.heading('response', text='Response (ms)')
        self.tree.heading('last_check', text='Last Check')
        
        self.tree.column('status', width=80, anchor='center')
        self.tree.column('name', width=200)
        self.tree.column('endpoint', width=300)
        self.tree.column('response', width=120, anchor='center')
        self.tree.column('last_check', width=100, anchor='center')
        
        self.tree.pack(fill='both', expand=True)
        
        # Summary
        summary_frame = tk.Frame(self.root, bg='#16213e', pady=10)
        summary_frame.pack(fill='x', padx=20, pady=10)
        
        self.online_label = tk.Label(summary_frame, text="🟢 Online: 0", font=('Arial', 12, 'bold'),
                                     bg='#16213e', fg=self.online_color)
        self.online_label.pack(side='left', padx=30)
        
        self.degraded_label = tk.Label(summary_frame, text="🟠 Degraded: 0", font=('Arial', 12, 'bold'),
                                       bg='#16213e', fg=self.degraded_color)
        self.degraded_label.pack(side='left', padx=30)
        
        self.offline_label = tk.Label(summary_frame, text="🔴 Offline: 0", font=('Arial', 12, 'bold'),
                                      bg='#16213e', fg=self.offline_color)
        self.offline_label.pack(side='left', padx=30)
        
        self.avg_response_label = tk.Label(summary_frame, text="⚡ Avg: 0ms", font=('Arial', 12, 'bold'),
                                           bg='#16213e', fg='white')
        self.avg_response_label.pack(side='right', padx=30)
        
        # Endpoints to monitor
        self.endpoints = [
            {'name': 'Reporting Dashboard', 'endpoint': '/api/reporting/dashboard'},
            {'name': 'System Status', 'endpoint': '/api/system-status'},
            {'name': 'Core Health', 'endpoint': '/health'},
            {'name': 'Docker Containers', 'endpoint': '/api/reporting/docker-containers'},
            {'name': 'Docker Stats', 'endpoint': '/api/reporting/docker-stats'},
            {'name': 'Excel Export', 'endpoint': '/api/reporting/export-excel'},
            {'name': 'PPTX Export', 'endpoint': '/api/reporting/export-pptx'},
        ]
        
    def start_monitoring(self):
        self.is_running = True
        self.start_btn.config(state='disabled', bg='#555')
        self.stop_btn.config(state='normal', bg='#cc3333')
        self.status_indicator.config(text="🔴", fg='red')
        self.status_label.config(text="LIVE", fg='#00ff88')
        
        try:
            self.refresh_interval = int(self.interval_var.get())
        except:
            self.refresh_interval = 30
        
        self.monitor_thread = threading.Thread(target=self.monitoring_loop, daemon=True)
        self.monitor_thread.start()
        
    def stop_monitoring(self):
        self.is_running = False
        self.start_btn.config(state='normal', bg='#00aa55')
        self.stop_btn.config(state='disabled', bg='#555')
        self.status_indicator.config(text="⚫", fg='gray')
        self.status_label.config(text="STOPPED", fg='gray')
        
    def monitoring_loop(self):
        while self.is_running:
            self.check_all_apis()
            for i in range(self.refresh_interval):
                if not self.is_running:
                    break
                time.sleep(1)
    
    def check_all_apis(self):
        results = []
        
        for ep in self.endpoints:
            url = BASE_URL + ep['endpoint']
            start_time = time.time()
            
            try:
                response = requests.get(url, timeout=10)
                response_time = int((time.time() - start_time) * 1000)
                
                if response.status_code == 200:
                    status = 'ONLINE' if response_time < 2000 else 'DEGRADED'
                else:
                    status = 'DEGRADED' if response.status_code < 500 else 'OFFLINE'
            except:
                status = 'OFFLINE'
                response_time = 0
            
            results.append({
                'name': ep['name'],
                'endpoint': ep['endpoint'],
                'status': status,
                'response_time': response_time,
                'last_check': datetime.now().strftime('%H:%M:%S')
            })
        
        # Update UI in main thread
        self.root.after(0, lambda: self.update_ui(results))
        
        # Update Excel
        self.update_excel(results)
        
    def update_ui(self, results):
        # Clear tree
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        online = degraded = offline = 0
        total_response = 0
        
        for r in results:
            if r['status'] == 'ONLINE':
                status_text = '🟢 ONLINE'
                online += 1
            elif r['status'] == 'DEGRADED':
                status_text = '🟠 DEGRADED'
                degraded += 1
            else:
                status_text = '🔴 OFFLINE'
                offline += 1
            
            total_response += r['response_time']
            
            self.tree.insert('', 'end', values=(
                status_text, r['name'], r['endpoint'], 
                f"{r['response_time']}ms", r['last_check']
            ))
        
        # Update summary
        self.online_label.config(text=f"🟢 Online: {online}")
        self.degraded_label.config(text=f"🟠 Degraded: {degraded}")
        self.offline_label.config(text=f"🔴 Offline: {offline}")
        
        avg = total_response // len(results) if results else 0
        self.avg_response_label.config(text=f"⚡ Avg: {avg}ms")
        
        self.last_update_label.config(text=f"Last: {datetime.now().strftime('%H:%M:%S')}")
        
    def update_excel(self, results):
        """Përditëson të njëjtin Excel - Postman style, pa ngjyra!"""
        try:
            # Krijo ose ngarko
            if os.path.exists(EXCEL_PATH):
                try:
                    wb = load_workbook(EXCEL_PATH)
                    ws = wb.active
                except:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = 'API_Status'
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = 'API_Status'
            
            # Clear dhe rishkruaj
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None
                    cell.fill = PatternFill()  # Remove any fill
            
            # Simple Postman-style headers (no colors)
            headers = ['Name', 'Method', 'URL', 'Status', 'Response_Time_ms', 'Timestamp']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
            
            # Data - plain text, no colors
            for row_idx, r in enumerate(results, 2):
                ws.cell(row=row_idx, column=1, value=r['name'])
                ws.cell(row=row_idx, column=2, value='GET')
                ws.cell(row=row_idx, column=3, value=BASE_URL + r['endpoint'])
                ws.cell(row=row_idx, column=4, value=r['status'])
                ws.cell(row=row_idx, column=5, value=r['response_time'])
                ws.cell(row=row_idx, column=6, value=r['last_check'])
            
            # Column widths
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 8
            ws.column_dimensions['C'].width = 50
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 12
            
            wb.save(EXCEL_PATH)
            
        except PermissionError:
            pass  # Excel is open, skip silently
        except Exception as e:
            print(f"Excel error: {e}")
            
    def open_excel(self):
        if os.path.exists(EXCEL_PATH):
            os.startfile(EXCEL_PATH)
        else:
            messagebox.showinfo("Info", "Excel nuk ekziston ende. Kliko START për ta krijuar.")

def main():
    root = tk.Tk()
    app = LiveDashboardApp(root)
    root.mainloop()

if __name__ == '__main__':
    main()
