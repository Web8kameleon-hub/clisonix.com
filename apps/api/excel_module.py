#!/usr/bin/env python3
"""
🔷 EXCEL MODULE - FULLY ISOLATED
================================
Clisonix Deterministic SaaS Protocol

NO PANDAS - Uses openpyxl only to avoid numpy binary conflicts
Isolated for container deployment

Author: Clisonix
Version: 1.0.0
"""

import os
import json
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional
from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse, JSONResponse

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ============================================================================
# CONFIGURATION
# ============================================================================

# Detect environment - container vs local
if os.path.exists("/app"):
    PROJECT_ROOT = Path("/app")
else:
    PROJECT_ROOT = Path(__file__).parent.parent.parent

EXCEL_DIRS = [
    PROJECT_ROOT / "dashboard",
    PROJECT_ROOT / "excel",
    PROJECT_ROOT / "excels",
    PROJECT_ROOT,
]

# ============================================================================
# ROUTER
# ============================================================================

router = APIRouter(prefix="/api/excel", tags=["Excel Dashboard"])


def find_excel_files() -> List[Dict[str, Any]]:
    """Finds all Excel files in configured directories"""
    files = []
    seen = set()
    
    for directory in EXCEL_DIRS:
        if directory.exists():
            for pattern in ["*.xlsx", "*.xls", "*.xlsm"]:
                for file in directory.glob(pattern):
                    if file.name not in seen and not file.name.startswith("~$"):
                        seen.add(file.name)
                        try:
                            stat = file.stat()
                            files.append({
                                "name": file.name,
                                "path": str(file),
                                "size_kb": round(stat.st_size / 1024, 2),
                                "modified": datetime.fromtimestamp(stat.st_mtime).isoformat(),
                                "directory": str(directory)
                            })
                        except Exception:
                            pass
    
    return sorted(files, key=lambda x: x["name"])


def read_excel_summary(filepath: Path) -> Dict[str, Any]:
    """Reads Excel file summary without pandas"""
    if not OPENPYXL_AVAILABLE:
        return {"error": "openpyxl not available", "sheets": [], "total_rows": 0}
    
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        sheets = []
        total_rows = 0
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            row_count = ws.max_row or 0
            col_count = ws.max_column or 0
            total_rows += row_count
            
            # Get headers (first row)
            headers = []
            if row_count > 0:
                for col in range(1, min(col_count + 1, 20)):  # Max 20 columns
                    cell = ws.cell(row=1, column=col)
                    if cell.value:
                        headers.append(str(cell.value))
            
            sheets.append({
                "name": sheet_name,
                "rows": row_count,
                "columns": col_count,
                "headers": headers[:10]  # First 10 headers
            })
        
        wb.close()
        
        return {
            "sheets": sheets,
            "total_rows": total_rows,
            "sheet_count": len(sheets)
        }
    except Exception as e:
        return {
            "error": str(e),
            "sheets": [],
            "total_rows": 0
        }


# ============================================================================
# ENDPOINTS
# ============================================================================

@router.get("/status")
async def excel_status():
    """Excel module status check"""
    files = find_excel_files()
    
    return {
        "status": "operational",
        "module": "excel_dashboard_isolated",
        "version": "1.0.0",
        "openpyxl_available": OPENPYXL_AVAILABLE,
        "files_found": len(files),
        "directories_scanned": [str(d) for d in EXCEL_DIRS if d.exists()],
        "timestamp": datetime.now().isoformat(),
        "isolation": "full",
        "dependencies": ["openpyxl"],
        "no_pandas": True
    }


@router.get("/dashboards")
async def list_dashboards():
    """List all Excel dashboard files"""
    files = find_excel_files()
    
    return {
        "dashboards": files,
        "total": len(files),
        "timestamp": datetime.now().isoformat()
    }


@router.get("/summary")
async def dashboards_summary():
    """Get summary of all Excel dashboards with sheet info"""
    files = find_excel_files()
    summaries = []
    
    for file in files:
        filepath = Path(file["path"])
        summary = read_excel_summary(filepath)
        summaries.append({
            "file": file["name"],
            "size_kb": file["size_kb"],
            "modified": file["modified"],
            **summary
        })
    
    return {
        "summaries": summaries,
        "total_files": len(files),
        "total_sheets": sum(s.get("sheet_count", 0) for s in summaries),
        "total_rows": sum(s.get("total_rows", 0) for s in summaries),
        "timestamp": datetime.now().isoformat()
    }


@router.get("/download/{filename}")
async def download_dashboard(filename: str):
    """Download a specific Excel file"""
    files = find_excel_files()
    
    for file in files:
        if file["name"] == filename:
            filepath = Path(file["path"])
            if filepath.exists():
                return FileResponse(
                    path=str(filepath),
                    filename=filename,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    
    raise HTTPException(status_code=404, detail=f"Dashboard '{filename}' not found")


@router.get("/preview/{filename}")
async def preview_dashboard(filename: str, sheet: Optional[str] = None, max_rows: int = 50):
    """Preview Excel file contents as JSON (first N rows)"""
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=503, detail="openpyxl not available")
    
    files = find_excel_files()
    filepath = None
    
    for file in files:
        if file["name"] == filename:
            filepath = Path(file["path"])
            break
    
    if not filepath or not filepath.exists():
        raise HTTPException(status_code=404, detail=f"Dashboard '{filename}' not found")
    
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        
        # Get sheet
        if sheet and sheet in wb.sheetnames:
            ws = wb[sheet]
        else:
            ws = wb.active
        
        # Read data
        data = []
        headers = []
        
        for row_idx, row in enumerate(ws.iter_rows(max_row=max_rows + 1), 1):
            row_data = [cell.value for cell in row]
            if row_idx == 1:
                headers = [str(h) if h else f"Col{i}" for i, h in enumerate(row_data)]
            else:
                data.append(dict(zip(headers, row_data)))
        
        wb.close()
        
        return {
            "filename": filename,
            "sheet": ws.title,
            "headers": headers,
            "rows": data,
            "row_count": len(data),
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error reading file: {str(e)}")


@router.get("/sheets/{filename}")
async def list_sheets(filename: str):
    """List all sheets in an Excel file"""
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=503, detail="openpyxl not available")
    
    files = find_excel_files()
    filepath = None
    
    for file in files:
        if file["name"] == filename:
            filepath = Path(file["path"])
            break
    
    if not filepath or not filepath.exists():
        raise HTTPException(status_code=404, detail=f"Dashboard '{filename}' not found")
    
    try:
        wb = load_workbook(filepath, read_only=True)
        sheets = []
        
        for name in wb.sheetnames:
            ws = wb[name]
            sheets.append({
                "name": name,
                "rows": ws.max_row,
                "columns": ws.max_column
            })
        
        wb.close()
        
        return {
            "filename": filename,
            "sheets": sheets,
            "sheet_count": len(sheets)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error reading file: {str(e)}")


# ============================================================================
# HEALTH CHECK
# ============================================================================

@router.get("/health")
async def excel_health():
    """Quick health check"""
    return {
        "status": "healthy",
        "module": "excel_isolated",
        "openpyxl": OPENPYXL_AVAILABLE,
        "timestamp": datetime.now().isoformat()
    }
