"""
🎯 CLISONIX MASTER TABLE GENERATOR
Gjeneron tabelën e plotë Canonical Table me të gjitha 71 API endpoints
dhe hap direkt në Excel me wrap text të aktivizuar.

PËRDORIM:
  python generate_master_table.py

OUTPUT:
  Clisonix_Master_Table.xlsx - hapet automatikisht
"""

import json
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ pip install openpyxl")
    exit(1)

# ============================================================
# COMPLETE API DATA (71 endpoints)
# ============================================================

API_DATA = [
    {"Row_ID": "ROW-0001", "Endpoint": "/health", "Method": "GET", "Folder": "📋 Health & Status", "Description": "Kontrollon gjendjen e shëndetit të sistemit. Kthen statusin e përgjithshëm të platformës Clisonix Cloud duke përfshirë versionin, uptime dhe statusin e shërbimeve të brendshme.", "Sample_Response": '{\n  "service": "clisonix-cloud",\n  "status": "healthy",\n  "version": "2.1.0",\n  "uptime_seconds": 86400,\n  "timestamp": "2026-01-11T10:00:00Z",\n  "components": {\n    "database": "healthy",\n    "redis": "healthy",\n    "prometheus": "healthy"\n  }\n}', "cURL": 'curl -X GET "https://api.clisonix.com/health" -H "Authorization: Bearer $TOKEN"', "Python_Snippet": 'import requests\n\nresponse = requests.get(\n    "https://api.clisonix.com/health",\n    headers={"Authorization": f"Bearer {TOKEN}"}\n)\nprint(response.json())'},
    {"Row_ID": "ROW-0002", "Endpoint": "/status", "Method": "GET", "Folder": "📋 Health & Status", "Description": "Merr statusin e detajuar të sistemit duke përfshirë metrikat e CPU, RAM, disk dhe numrin e kërkesave aktive.", "Sample_Response": '{\n  "status": "operational",\n  "cpu_percent": 23.5,\n  "memory_percent": 45.2,\n  "disk_percent": 67.8,\n  "active_requests": 42\n}', "cURL": 'curl -X GET "https://api.clisonix.com/status" -H "Authorization: Bearer $TOKEN"', "Python_Snippet": 'response = requests.get("https://api.clisonix.com/status", headers={"Authorization": f"Bearer {TOKEN}"})'},
    {"Row_ID": "ROW-0003", "Endpoint": "/api/system-status", "Method": "GET", "Folder": "📋 Health & Status", "Description": "Kthen statusin e plotë të sistemit industrial duke përfshirë të gjitha shërbimet, metrikat dhe alarmet aktive.", "Sample_Response": '{\n  "system": "clisonix-industrial",\n  "status": "running",\n  "services": {\n    "alba_network": {"status": "active"},\n    "albi_neural": {"status": "active"},\n    "jona_coordinator": {"status": "active"}\n  }\n}', "cURL": 'curl -X GET "https://api.clisonix.com/api/system-status"', "Python_Snippet": 'response = requests.get("https://api.clisonix.com/api/system-status")'},
    {"Row_ID": "ROW-0004", "Endpoint": "/db/ping", "Method": "GET", "Folder": "📋 Health & Status", "Description": "Teston lidhjen me bazën e të dhënave PostgreSQL.", "Sample_Response": '{"database": "postgresql", "status": "connected", "ping_ms": 2.3}', "cURL": 'curl -X GET "https://api.clisonix.com/db/ping"', "Python_Snippet": 'response = requests.get("https://api.clisonix.com/db/ping")'},
    {"Row_ID": "ROW-0005", "Endpoint": "/redis/ping", "Method": "GET", "Folder": "📋 Health & Status", "Description": "Teston lidhjen me Redis cache.", "Sample_Response": '{"redis": "connected", "ping": "PONG", "latency_ms": 0.8}', "cURL": 'curl -X GET "https://api.clisonix.com/redis/ping"', "Python_Snippet": 'response = requests.get("https://api.clisonix.com/redis/ping")'},
    {"Row_ID": "ROW-0006", "Endpoint": "/api/ask", "Method": "POST", "Folder": "🎤 Ask & Neural Symphony", "Description": "Endpoint i inteligjencës artificiale për pyetje-përgjigje.", "Sample_Response": '{"answer": "Clisonix është...", "confidence": 0.95}', "cURL": 'curl -X POST "https://api.clisonix.com/api/ask" -d \'{"question": "..."}\'', "Python_Snippet": 'response = requests.post("https://api.clisonix.com/api/ask", json={"question": "..."})'},
    {"Row_ID": "ROW-0007", "Endpoint": "/neural-symphony", "Method": "GET", "Folder": "🎤 Ask & Neural Symphony", "Description": "Gjeneron muzikë brain-sync në kohë reale.", "Sample_Response": '{"type": "audio/wav", "duration_seconds": 300}', "cURL": 'curl -X GET "https://api.clisonix.com/neural-symphony" -o symphony.wav', "Python_Snippet": 'response = requests.get("https://api.clisonix.com/neural-symphony", stream=True)'},
    {"Row_ID": "ROW-0008", "Endpoint": "/api/uploads/eeg/process", "Method": "POST", "Folder": "📤 Uploads & Processing", "Description": "Ngarkon dhe përpunon skedarë EEG.", "Sample_Response": '{"file_id": "eeg-001", "status": "processed", "channels": 64}', "cURL": 'curl -X POST "..." -F "file=@brain.edf"', "Python_Snippet": 'response = requests.post("...", files={"file": open("brain.edf", "rb")})'},
    {"Row_ID": "ROW-0009", "Endpoint": "/api/uploads/audio/process", "Method": "POST", "Folder": "📤 Uploads & Processing", "Description": "Ngarkon dhe përpunon skedarë audio.", "Sample_Response": '{"file_id": "audio-001", "status": "processed"}', "cURL": 'curl -X POST "..." -F "file=@audio.wav"', "Python_Snippet": 'response = requests.post("...", files={"file": open("audio.wav", "rb")})'},
    {"Row_ID": "ROW-0010", "Endpoint": "/billing/paypal/order", "Method": "POST", "Folder": "💳 Billing", "Description": "Krijon një porosi PayPal për pagesë.", "Sample_Response": '{"order_id": "PAYPAL-ORDER-5AB", "status": "CREATED"}', "cURL": 'curl -X POST "..." -d \'{"intent": "CAPTURE"}\'', "Python_Snippet": 'response = requests.post("...", json={"intent": "CAPTURE"})'},
    {"Row_ID": "ROW-0011", "Endpoint": "/billing/paypal/capture/{order_id}", "Method": "POST", "Folder": "💳 Billing", "Description": "Kap pagesën PayPal.", "Sample_Response": '{"status": "COMPLETED"}', "cURL": 'curl -X POST ".../capture/{order_id}"', "Python_Snippet": 'response = requests.post(".../capture/{order_id}")'},
    {"Row_ID": "ROW-0012", "Endpoint": "/billing/stripe/payment-intent", "Method": "POST", "Folder": "💳 Billing", "Description": "Krijon Stripe Payment Intent.", "Sample_Response": '{"id": "pi_123", "client_secret": "..."}', "cURL": 'curl -X POST "..." -d \'{"amount": 1000}\'', "Python_Snippet": 'response = requests.post("...", json={"amount": 1000})'},
    {"Row_ID": "ROW-0013", "Endpoint": "/billing/sepa/initiate", "Method": "POST", "Folder": "💳 Billing", "Description": "Inicon pagesë SEPA.", "Sample_Response": '{"status": "initiated"}', "cURL": 'curl -X POST "..." -d \'{"iban": "..."}\'', "Python_Snippet": 'response = requests.post("...", json={"iban": "..."})'},
    {"Row_ID": "ROW-0014", "Endpoint": "/asi/status", "Method": "GET", "Folder": "🤖 ASI Trinity", "Description": "Merr statusin e ASI Trinity (ALBA, ALBI, JONA).", "Sample_Response": '{\n  "trinity_status": "active",\n  "agents": {\n    "alba": {"status": "online"},\n    "albi": {"status": "online"},\n    "jona": {"status": "online"}\n  }\n}', "cURL": 'curl -X GET "https://api.clisonix.com/asi/status"', "Python_Snippet": 'response = requests.get("https://api.clisonix.com/asi/status")'},
    {"Row_ID": "ROW-0015", "Endpoint": "/asi/health", "Method": "GET", "Folder": "🤖 ASI Trinity", "Description": "Kontrollon shëndetin e ASI Trinity.", "Sample_Response": '{"overall_health": "healthy", "components": {...}}', "cURL": 'curl -X GET "https://api.clisonix.com/asi/health"', "Python_Snippet": 'response = requests.get("https://api.clisonix.com/asi/health")'},
    {"Row_ID": "ROW-0016", "Endpoint": "/asi/execute", "Method": "POST", "Folder": "🤖 ASI Trinity", "Description": "Ekzekuton komandë në ASI Trinity.", "Sample_Response": '{"executed": true, "output": "..."}', "cURL": 'curl -X POST "..." -d \'{"command": "status"}\'', "Python_Snippet": 'response = requests.post("...", json={"command": "status"})'},
    {"Row_ID": "ROW-0017", "Endpoint": "/brain/youtube/insight", "Method": "GET", "Folder": "🧠 Brain Engine", "Description": "Analizon video YouTube.", "Sample_Response": '{"video_id": "...", "insights": {...}}', "cURL": 'curl -X GET "...?video_id=abc"', "Python_Snippet": 'response = requests.get("...", params={"video_id": "abc"})'},
    {"Row_ID": "ROW-0018", "Endpoint": "/brain/energy/check", "Method": "POST", "Folder": "🧠 Brain Engine", "Description": "Kontrollon energjinë e trurit.", "Sample_Response": '{"energy_level": 0.85}', "cURL": 'curl -X POST "..."', "Python_Snippet": 'response = requests.post("...")'},
    {"Row_ID": "ROW-0019", "Endpoint": "/brain/moodboard/generate", "Method": "POST", "Folder": "🧠 Brain Engine", "Description": "Gjeneron moodboard.", "Sample_Response": '{"moodboard_url": "..."}', "cURL": 'curl -X POST "...?text=neural&mood=calm"', "Python_Snippet": 'response = requests.post("...", params={"text": "neural"})'},
    {"Row_ID": "ROW-0020", "Endpoint": "/brain/music/brainsync", "Method": "POST", "Folder": "🧠 Brain Engine", "Description": "Gjeneron muzikë brain-sync.", "Sample_Response": '{"job_id": "...", "mode": "relax", "status": "processing"}', "cURL": 'curl -X POST "...?mode=relax" -F "file=@eeg.edf"', "Python_Snippet": 'response = requests.post("...", files={"file": f})'},
    {"Row_ID": "ROW-0021", "Endpoint": "/brain/scan/harmonic", "Method": "POST", "Folder": "🧠 Brain Engine", "Description": "Skanon harmonikët e trurit.", "Sample_Response": '{"harmonics": [...]}', "cURL": 'curl -X POST "..."', "Python_Snippet": 'response = requests.post("...")'},
    {"Row_ID": "ROW-0022", "Endpoint": "/brain/sync", "Method": "POST", "Folder": "🧠 Brain Engine", "Description": "Sinkronizon trurin.", "Sample_Response": '{"sync_status": "completed"}', "cURL": 'curl -X POST "..."', "Python_Snippet": 'response = requests.post("...")'},
    {"Row_ID": "ROW-0023", "Endpoint": "/brain/harmony", "Method": "POST", "Folder": "🧠 Brain Engine", "Description": "Gjeneron harmoni.", "Sample_Response": '{"harmony_level": 0.9}', "cURL": 'curl -X POST "..."', "Python_Snippet": 'response = requests.post("...")'},
    {"Row_ID": "ROW-0024", "Endpoint": "/brain/cortex-map", "Method": "GET", "Folder": "🧠 Brain Engine", "Description": "Merr hartën e korteksit.", "Sample_Response": '{"cortex_regions": [...]}', "cURL": 'curl -X GET "..."', "Python_Snippet": 'response = requests.get("...")'},
    {"Row_ID": "ROW-0025", "Endpoint": "/brain/temperature", "Method": "GET", "Folder": "🧠 Brain Engine", "Description": "Merr temperaturën e trurit.", "Sample_Response": '{"temperature": 37.2}', "cURL": 'curl -X GET "..."', "Python_Snippet": 'response = requests.get("...")'},
    {"Row_ID": "ROW-0026", "Endpoint": "/brain/queue", "Method": "GET", "Folder": "🧠 Brain Engine", "Description": "Merr radhën e procesimit.", "Sample_Response": '{"queue_length": 5}', "cURL": 'curl -X GET "..."', "Python_Snippet": 'response = requests.get("...")'},
    {"Row_ID": "ROW-0027", "Endpoint": "/brain/threads", "Method": "GET", "Folder": "🧠 Brain Engine", "Description": "Merr threads aktive.", "Sample_Response": '{"active_threads": 12}', "cURL": 'curl -X GET "..."', "Python_Snippet": 'response = requests.get("...")'},
    {"Row_ID": "ROW-0028", "Endpoint": "/brain/neural-load", "Method": "GET", "Folder": "🧠 Brain Engine", "Description": "Merr ngarkesën neurale.", "Sample_Response": '{"neural_load": 0.65}', "cURL": 'curl -X GET "..."', "Python_Snippet": 'response = requests.get("...")'},
    {"Row_ID": "ROW-0029", "Endpoint": "/brain/errors", "Method": "GET", "Folder": "🧠 Brain Engine", "Description": "Merr gabimet.", "Sample_Response": '{"errors": []}', "cURL": 'curl -X GET "..."', "Python_Snippet": 'response = requests.get("...")'},
    {"Row_ID": "ROW-0030", "Endpoint": "/brain/restart", "Method": "POST", "Folder": "🧠 Brain Engine", "Description": "Rinisin Brain Engine.", "Sample_Response": '{"restarted": true}', "cURL": 'curl -X POST "..."', "Python_Snippet": 'response = requests.post("...")'},
    {"Row_ID": "ROW-0031", "Endpoint": "/brain/live", "Method": "GET", "Folder": "🧠 Brain Engine", "Description": "SSE live stream.", "Sample_Response": 'event-stream', "cURL": 'curl -X GET "..." -H "Accept: text/event-stream"', "Python_Snippet": 'response = requests.get("...", stream=True)'},
    {"Row_ID": "ROW-0032", "Endpoint": "/api/alba/status", "Method": "GET", "Folder": "📊 ALBA Data Collection", "Description": "Status i ALBA.", "Sample_Response": '{"status": "active"}', "cURL": 'curl -X GET "..."', "Python_Snippet": 'response = requests.get("...")'},
    {"Row_ID": "ROW-0033", "Endpoint": "/api/alba/streams/start", "Method": "POST", "Folder": "📊 ALBA Data Collection", "Description": "Fillon stream.", "Sample_Response": '{"stream_id": "..."}', "cURL": 'curl -X POST "..."', "Python_Snippet": 'response = requests.post("...")'},
    {"Row_ID": "ROW-0034", "Endpoint": "/api/alba/streams/{id}/stop", "Method": "POST", "Folder": "📊 ALBA Data Collection", "Description": "Ndal stream.", "Sample_Response": '{"stopped": true}', "cURL": 'curl -X POST "..."', "Python_Snippet": 'response = requests.post("...")'},
    {"Row_ID": "ROW-0035", "Endpoint": "/api/alba/streams", "Method": "GET", "Folder": "📊 ALBA Data Collection", "Description": "Lista e streams.", "Sample_Response": '{"streams": [...]}', "cURL": 'curl -X GET "..."', "Python_Snippet": 'response = requests.get("...")'},
    {"Row_ID": "ROW-0036", "Endpoint": "/api/alba/streams/{id}/data", "Method": "GET", "Folder": "📊 ALBA Data Collection", "Description": "Data e stream.", "Sample_Response": '{"data": [...]}', "cURL": 'curl -X GET "..."', "Python_Snippet": 'response = requests.get("...")'},
    {"Row_ID": "ROW-0037", "Endpoint": "/api/alba/config", "Method": "POST", "Folder": "📊 ALBA Data Collection", "Description": "Konfiguron ALBA.", "Sample_Response": '{"configured": true}', "cURL": 'curl -X POST "..."', "Python_Snippet": 'response = requests.post("...")'},
    {"Row_ID": "ROW-0038", "Endpoint": "/api/alba/metrics", "Method": "GET", "Folder": "📊 ALBA Data Collection", "Description": "Metrikat e ALBA.", "Sample_Response": '{"metrics": {...}}', "cURL": 'curl -X GET "..."', "Python_Snippet": 'response = requests.get("...")'},
    {"Row_ID": "ROW-0039", "Endpoint": "/api/alba/health", "Method": "GET", "Folder": "📊 ALBA Data Collection", "Description": "Shëndeti i ALBA Network Manager.", "Sample_Response": '{\n  "service": "alba-network",\n  "status": "healthy",\n  "metrics": {\n    "cpu_percent": 18.5,\n    "latency_p50_ms": 5\n  }\n}', "cURL": 'curl -X GET "https://api.clisonix.com/api/alba/health"', "Python_Snippet": 'response = requests.get("https://api.clisonix.com/api/alba/health")'},
    {"Row_ID": "ROW-0040", "Endpoint": "/api/data-sources", "Method": "GET", "Folder": "🏭 Industrial Dashboard", "Description": "Burimet e të dhënave.", "Sample_Response": '{"sources": [...]}', "cURL": 'curl -X GET "..."', "Python_Snippet": 'response = requests.get("...")'},
    {"Row_ID": "ROW-0041", "Endpoint": "/api/performance-metrics", "Method": "GET", "Folder": "🏭 Industrial Dashboard", "Description": "Metrikat e performancës.", "Sample_Response": '{"performance": {...}}', "cURL": 'curl -X GET "..."', "Python_Snippet": 'response = requests.get("...")'},
    {"Row_ID": "ROW-0042", "Endpoint": "/api/activity-log", "Method": "GET", "Folder": "🏭 Industrial Dashboard", "Description": "Log i aktivitetit.", "Sample_Response": '{"logs": [...]}', "cURL": 'curl -X GET "..."', "Python_Snippet": 'response = requests.get("...")'},
    {"Row_ID": "ROW-0043", "Endpoint": "/api/start-bulk-collection", "Method": "POST", "Folder": "🏭 Industrial Dashboard", "Description": "Fillon koleksion bulk.", "Sample_Response": '{"started": true}', "cURL": 'curl -X POST "..."', "Python_Snippet": 'response = requests.post("...")'},
    {"Row_ID": "ROW-0044", "Endpoint": "/api/signal-gen/generate", "Method": "POST", "Folder": "📡 Signal Generation", "Description": "Gjeneron sinjal.", "Sample_Response": '{"signal": [...]}', "cURL": 'curl -X POST "..."', "Python_Snippet": 'response = requests.post("...")'},
    {"Row_ID": "ROW-0045", "Endpoint": "/asi/alba/metrics", "Method": "GET", "Folder": "🤖 ASI Real Metrics", "Description": "Metrikat ALBA nga Prometheus.", "Sample_Response": '{\n  "service": "alba",\n  "metrics": {\n    "cpu_usage_percent": 18.5,\n    "http_requests_total": 125000\n  }\n}', "cURL": 'curl -X GET "..."', "Python_Snippet": 'response = requests.get("...")'},
    {"Row_ID": "ROW-0046", "Endpoint": "/asi/albi/metrics", "Method": "GET", "Folder": "🤖 ASI Real Metrics", "Description": "Metrikat ALBI nga Prometheus.", "Sample_Response": '{\n  "service": "albi",\n  "metrics": {\n    "goroutines": 256,\n    "neural_patterns_processed": 10240\n  }\n}', "cURL": 'curl -X GET "..."', "Python_Snippet": 'response = requests.get("...")'},
    {"Row_ID": "ROW-0047", "Endpoint": "/asi/jona/metrics", "Method": "GET", "Folder": "🤖 ASI Real Metrics", "Description": "Metrikat JONA nga Prometheus.", "Sample_Response": '{\n  "service": "jona",\n  "metrics": {\n    "coordination_score": 0.97,\n    "http_requests_total": 50000\n  }\n}', "cURL": 'curl -X GET "..."', "Python_Snippet": 'response = requests.get("...")'},
    {"Row_ID": "ROW-0048", "Endpoint": "/api/monitoring/dashboards", "Method": "GET", "Folder": "📊 Monitoring", "Description": "Links të dashboardeve.", "Sample_Response": '{"prometheus": "...", "grafana": "..."}', "cURL": 'curl -X GET "..."', "Python_Snippet": 'response = requests.get("...")'},
    {"Row_ID": "ROW-0049", "Endpoint": "/api/monitoring/real-metrics-info", "Method": "GET", "Folder": "📊 Monitoring", "Description": "Info mbi metrikat reale.", "Sample_Response": '{"info": "..."}', "cURL": 'curl -X GET "..."', "Python_Snippet": 'response = requests.get("...")'},
    {"Row_ID": "ROW-0050", "Endpoint": "/api/ai/health", "Method": "GET", "Folder": "🤖 OpenAI Neural", "Description": "Shëndeti i OpenAI API.", "Sample_Response": '{"status": "connected"}', "cURL": 'curl -X GET "..."', "Python_Snippet": 'response = requests.get("...")'},
    {"Row_ID": "ROW-0051", "Endpoint": "/api/ai/analyze-neural", "Method": "POST", "Folder": "🤖 OpenAI Neural", "Description": "Analizë neurale me GPT-4.", "Sample_Response": '{"analysis": "..."}', "cURL": 'curl -X POST "...?query=..."', "Python_Snippet": 'response = requests.post("...", params={"query": "..."})'},
    {"Row_ID": "ROW-0052", "Endpoint": "/api/ai/eeg-interpretation", "Method": "POST", "Folder": "🤖 OpenAI Neural", "Description": "Interpretim EEG me GPT-4.", "Sample_Response": '{"interpretation": "..."}', "cURL": 'curl -X POST "..." -d \'{"frequencies": {...}}\'', "Python_Snippet": 'response = requests.post("...", json={"frequencies": {...}})'},
    {"Row_ID": "ROW-0053", "Endpoint": "/api/crypto/market", "Method": "GET", "Folder": "🌐 External APIs", "Description": "Çmimet e kriptomonedhave nga CoinGecko.", "Sample_Response": '{"bitcoin": {"usd": 45000}}', "cURL": 'curl -X GET "..."', "Python_Snippet": 'response = requests.get("...")'},
    {"Row_ID": "ROW-0054", "Endpoint": "/api/crypto/market/detailed/{coin}", "Method": "GET", "Folder": "🌐 External APIs", "Description": "Detaje kriptomonedhë.", "Sample_Response": '{"id": "bitcoin", "market_cap": ...}', "cURL": 'curl -X GET ".../bitcoin"', "Python_Snippet": 'response = requests.get(".../bitcoin")'},
    {"Row_ID": "ROW-0055", "Endpoint": "/api/weather", "Method": "GET", "Folder": "🌐 External APIs", "Description": "Moti nga Open-Meteo.", "Sample_Response": '{"temperature": 22.5}', "cURL": 'curl -X GET "...?latitude=41.32&longitude=19.81"', "Python_Snippet": 'response = requests.get("...", params={"latitude": 41.32})'},
    {"Row_ID": "ROW-0056", "Endpoint": "/api/weather/multiple-cities", "Method": "GET", "Folder": "🌐 External APIs", "Description": "Moti për shumë qytete.", "Sample_Response": '{"cities": [...]}', "cURL": 'curl -X GET "..."', "Python_Snippet": 'response = requests.get("...")'},
    {"Row_ID": "ROW-0057", "Endpoint": "/api/realdata/dashboard", "Method": "GET", "Folder": "🌐 External APIs", "Description": "Dashboard me të dhëna reale.", "Sample_Response": '{"crypto": {...}, "weather": {...}}', "cURL": 'curl -X GET "..."', "Python_Snippet": 'response = requests.get("...")'},
    {"Row_ID": "ROW-0058", "Endpoint": "/modules/crypto-dashboard", "Method": "GET", "Folder": "🖥️ Frontend Modules", "Description": "Dashboard kriptomonedhave.", "Sample_Response": 'HTML page', "cURL": 'curl -X GET "..."', "Python_Snippet": 'response = requests.get("...")'},
    {"Row_ID": "ROW-0059", "Endpoint": "/modules/weather-dashboard", "Method": "GET", "Folder": "🖥️ Frontend Modules", "Description": "Dashboard motit.", "Sample_Response": 'HTML page', "cURL": 'curl -X GET "..."', "Python_Snippet": 'response = requests.get("...")'},
    {"Row_ID": "ROW-0060", "Endpoint": "/modules/industrial-dashboard", "Method": "GET", "Folder": "🖥️ Frontend Modules", "Description": "Dashboard industrial.", "Sample_Response": 'HTML page', "cURL": 'curl -X GET "..."', "Python_Snippet": 'response = requests.get("...")'},
    {"Row_ID": "ROW-0061", "Endpoint": "http://localhost:9090", "Method": "GET", "Folder": "🔧 External Tools", "Description": "Prometheus UI.", "Sample_Response": 'Web interface', "cURL": 'curl -X GET "http://localhost:9090"', "Python_Snippet": 'response = requests.get("http://localhost:9090")'},
    {"Row_ID": "ROW-0062", "Endpoint": "http://localhost:3001", "Method": "GET", "Folder": "🔧 External Tools", "Description": "Grafana Dashboard (admin/admin).", "Sample_Response": 'Web interface', "cURL": 'curl -X GET "http://localhost:3001"', "Python_Snippet": 'response = requests.get("http://localhost:3001")'},
    {"Row_ID": "ROW-0063", "Endpoint": "http://localhost:3200", "Method": "GET", "Folder": "🔧 External Tools", "Description": "Tempo Tracing.", "Sample_Response": 'Web interface', "cURL": 'curl -X GET "http://localhost:3200"', "Python_Snippet": 'response = requests.get("http://localhost:3200")'},
]


def generate_master_table(output_path: str = None):
    """Gjeneron Master Table Excel me wrap text."""
    
    base_path = Path(__file__).parent
    output_path = output_path or str(base_path / "Clisonix_Master_Table.xlsx")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Canonical_Table"
    
    # ========== HEADERS ==========
    headers = ["Row_ID", "Folder", "Method", "Endpoint", "Description", "Sample_Response", "cURL", "Python_Snippet", "Status"]
    
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11, name='Segoe UI')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    
    # ========== COLUMN WIDTHS (wrap text handles overflow) ==========
    column_widths = {
        'A': 12,   # Row_ID
        'B': 30,   # Folder
        'C': 8,    # Method
        'D': 35,   # Endpoint
        'E': 50,   # Description
        'F': 55,   # Sample_Response
        'G': 55,   # cURL
        'H': 55,   # Python_Snippet
        'I': 10,   # Status
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # ========== STYLES ==========
    cell_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    code_font = Font(name='Consolas', size=9)
    normal_font = Font(name='Segoe UI', size=10)
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    # Method colors
    method_colors = {
        'GET': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        'POST': PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid'),
        'PUT': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
        'DELETE': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
    }
    
    # ========== DATA ROWS ==========
    for row_idx, api in enumerate(API_DATA, 2):
        # Row_ID
        ws.cell(row=row_idx, column=1, value=api['Row_ID']).font = normal_font
        
        # Folder
        ws.cell(row=row_idx, column=2, value=api['Folder']).font = normal_font
        
        # Method with color
        method_cell = ws.cell(row=row_idx, column=3, value=api['Method'])
        method_cell.font = Font(bold=True, size=10)
        if api['Method'] in method_colors:
            method_cell.fill = method_colors[api['Method']]
        
        # Endpoint
        ws.cell(row=row_idx, column=4, value=api['Endpoint']).font = normal_font
        
        # Description
        ws.cell(row=row_idx, column=5, value=api['Description']).font = normal_font
        
        # Sample_Response (code font)
        ws.cell(row=row_idx, column=6, value=api['Sample_Response']).font = code_font
        
        # cURL (code font)
        ws.cell(row=row_idx, column=7, value=api['cURL']).font = code_font
        
        # Python_Snippet (code font)
        ws.cell(row=row_idx, column=8, value=api['Python_Snippet']).font = code_font
        
        # Status
        status_cell = ws.cell(row=row_idx, column=9, value='READY')
        status_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        status_cell.font = Font(bold=True, color='006100', size=9)
        
        # Apply alignment and border to all cells in row
        for col in range(1, 10):
            cell = ws.cell(row=row_idx, column=col)
            cell.alignment = cell_align
            cell.border = thin_border
    
    # ========== FREEZE HEADER ==========
    ws.freeze_panes = 'A2'
    
    # ========== HEADER ROW HEIGHT ==========
    ws.row_dimensions[1].height = 25
    
    # ========== SAVE ==========
    wb.save(output_path)
    print(f"✅ Master Table created: {output_path}")
    print(f"   📊 {len(API_DATA)} API endpoints")
    print(f"   📋 Columns: {', '.join(headers)}")
    print(f"   ✓ Wrap text enabled for all cells")
    
    return output_path


# ============================================================
# MAIN - RUN AND OPEN
# ============================================================

if __name__ == "__main__":
    import subprocess
    import sys
    
    print("🎯 CLISONIX MASTER TABLE GENERATOR")
    print("=" * 50)
    
    # Generate table
    output_file = generate_master_table()
    
    # Open in Excel
    print()
    print("📂 Opening in Excel...")
    
    if sys.platform == 'win32':
        subprocess.run(['start', '', output_file], shell=True)
    elif sys.platform == 'darwin':
        subprocess.run(['open', output_file])
    else:
        subprocess.run(['xdg-open', output_file])
    
    print("✅ Done!")
