"""
Shto kolonën 'Funksione Shtesë' në Excel Protocol Registry
"""
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter

# Load Excel
wb = load_workbook('c:/Users/Admin/Desktop/real_excel_report.xlsx')
ws = wb['Protocol Registry']

# Styles
wrap_align = Alignment(wrap_text=True, vertical='top', horizontal='left')
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True, size=11)

# Shto kolonën e re në header (kolona 18)
col_num = 18
header_cell = ws.cell(row=1, column=col_num, value='Funksione_Shtesë')
header_cell.alignment = wrap_align
header_cell.border = border
header_cell.fill = header_fill
header_cell.font = header_font

# Funksionet sipas kategorisë
extra_functions = {
    'Health': 'Retry (3x), Timeout (5s), Logging, Health Dashboard Integration',
    'Health & Status': 'Retry (3x), Timeout (5s), Logging, Prometheus Metrics, AlertManager',
    'ASI Trinity': 'Retry (3x), Timeout (10s), Security (JWT), Rate Limit (100/min), Audit Log',
    'Brain Engine': 'Retry (2x), Timeout (30s), Caching (Redis 60s), ML Pipeline, Async Queue',
    'ALBA Data': 'Streaming, WebSocket, Buffer (1000 events), Backpressure, Real-time Sync',
    'Reporting': 'Caching (5min), Export Queue, Large File Handling, Progress Tracking',
    'Billing': 'Idempotency Key, Retry (5x), Webhook Callback, Transaction Log, PCI Compliance',
    'External APIs': 'Rate Limit (10/min), Caching (15min), Fallback, Circuit Breaker',
    'Monitoring': 'No Auth (Internal), Metrics Aggregation, Dashboard Embed',
    'Cycle Engine': 'State Machine, Retry with Backoff, Event Sourcing, Saga Pattern',
    'Auth': 'Rate Limit (5/min), Brute Force Protection, MFA Support, Session Management',
    'Pulse System': 'Async, Message Queue, Retry (3x), Dead Letter Queue',
    'Docker': 'Admin Only, Audit Log, Rate Limit (30/min), Container Isolation',
    'SDK': 'Version Check, Auto-Update, Config Validation',
    'System': 'Retry (3x), Timeout (5s), Logging, Metrics'
}

# Plotëso funksionet për çdo rresht
for row in range(2, ws.max_row + 1):
    folder = ws.cell(row=row, column=4).value  # Folder column
    
    # Gjej funksionin e duhur
    func = 'Retry, Logging, Metrics'  # Default
    if folder:
        for key, value in extra_functions.items():
            if key.lower() in folder.lower():
                func = value
                break
    
    cell = ws.cell(row=row, column=col_num, value=func)
    cell.alignment = wrap_align
    cell.border = border

# Vendos width për kolonën e re
ws.column_dimensions[get_column_letter(col_num)].width = 45

wb.save('c:/Users/Admin/Desktop/real_excel_report.xlsx')
print(f'✅ Kolona "Funksione_Shtesë" u shtua me sukses!')
print(f'📊 Total rreshta: {ws.max_row}')
print(f'📋 Total kolona: {ws.max_column}')
