"""
Canonical Table - Infinite Excel Engine
========================================
Append-only table with frozen headers, state machine, and maturity gate.
Core principle: Data is never deleted, only appended with new states.

States: RAW → NORMALIZED → TEST → READY/FAIL → ARCHIVED
Maturity: IMMATURE → MATURING → MATURE
"""

import json
import uuid
from datetime import datetime
from typing import Dict, List, Optional, Any
from dataclasses import dataclass, field, asdict
from enum import Enum
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment


class Status(Enum):
    """Row status in the pipeline"""
    RAW = "RAW"
    NORMALIZED = "NORMALIZED"
    TEST = "TEST"
    READY = "READY"
    FAIL = "FAIL"
    ARCHIVED = "ARCHIVED"


class MaturityState(Enum):
    """Maturity level determining ML activation eligibility"""
    IMMATURE = "IMMATURE"
    MATURING = "MATURING"
    MATURE = "MATURE"


class SecurityStatus(Enum):
    """Security validation result"""
    PASS = "PASS"
    FAIL = "FAIL"
    PENDING = "PENDING"


class ValidationStatus(Enum):
    """Schema validation result"""
    PASS = "PASS"
    FAIL = "FAIL"
    PENDING = "PENDING"


class MLConfidence(Enum):
    """ML confidence level"""
    LOW = "LOW"
    MEDIUM = "MEDIUM"
    HIGH = "HIGH"


@dataclass
class CanonicalRow:
    """Single row in the canonical table"""
    # Identifiers
    row_id: str = field(default_factory=lambda: str(uuid.uuid4()))
    timestamp: str = field(default_factory=lambda: datetime.utcnow().isoformat())
    
    # Core fields
    cycle: int = 0
    layer: str = "L1"
    variant: str = "A"
    input_type: str = ""
    source_protocol: str = ""
    
    # Status fields
    status: str = Status.RAW.value
    security: str = SecurityStatus.PENDING.value
    schema_validation: str = ValidationStatus.PENDING.value
    
    # Execution fields
    agent: Optional[str] = None
    lab: Optional[str] = None
    lab_result: Optional[Dict] = None
    artifacts: Optional[List[str]] = None
    
    # Maturity
    maturity_state: str = MaturityState.IMMATURE.value
    maturity_score: float = 0.0
    
    # ML Overlay (populated only when MATURE)
    ml_score: float = 0.0
    ml_confidence: str = MLConfidence.LOW.value
    ml_suggested_agent: Optional[str] = None
    ml_suggested_lab: Optional[str] = None
    ml_anomaly_prob: float = 0.0
    
    # Enforcement
    enforcement_applied: bool = False
    compliance_status: Optional[str] = None
    
    # Metadata
    metadata: Dict = field(default_factory=dict)
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary"""
        return asdict(self)
    
    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> 'CanonicalRow':
        """Create from dictionary"""
        return cls(**{k: v for k, v in data.items() if k in cls.__dataclass_fields__})


class CanonicalTable:
    """
    Infinite Excel Engine - Append-only canonical table
    
    Features:
    - Append-only: rows are never deleted
    - Frozen headers
    - Color-coded status and maturity
    - State machine transitions
    - Maturity gate for ML activation
    """
    
    HEADERS = [
        'Row_ID', 'Timestamp', 'Cycle', 'Layer', 'Variant', 'Input_Type', 
        'Source_Protocol', 'Status', 'Security', 'Schema_Validation',
        'Agent', 'Lab', 'Lab_Result', 'Artifacts', 'Maturity_State', 'Maturity_Score',
        'ML_Score', 'ML_Confidence', 'ML_Suggested_Agent', 'ML_Suggested_Lab', 
        'ML_Anomaly_Prob', 'Enforcement_Applied', 'Compliance_Status', 'Metadata'
    ]
    
    STATUS_COLORS = {
        'RAW': 'FFCCCC',       # Light red
        'NORMALIZED': 'FFE5CC', # Light orange
        'TEST': 'FFF2CC',       # Light yellow
        'READY': 'CCFFCC',      # Light green
        'FAIL': 'FF6666',       # Red
        'ARCHIVED': 'CCCCCC'    # Gray
    }
    
    MATURITY_COLORS = {
        'IMMATURE': 'FFCCCC',   # Light red
        'MATURING': 'FFF2CC',   # Light yellow
        'MATURE': 'CCFFCC'      # Light green
    }
    
    ML_CONFIDENCE_COLORS = {
        'LOW': 'FFCCCC',        # Light red
        'MEDIUM': 'FFF2CC',     # Light yellow
        'HIGH': 'CCFFCC'        # Light green
    }
    
    def __init__(self, storage_path: str = "data/canonical_table.json", 
                 excel_path: str = "data/canonical_table.xlsx"):
        """Initialize canonical table with storage paths"""
        self.storage_path = Path(storage_path)
        self.excel_path = Path(excel_path)
        self.rows: List[CanonicalRow] = []
        
        # Ensure data directory exists
        self.storage_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Load existing data
        self._load()
    
    def _load(self) -> None:
        """Load existing data from JSON storage"""
        if self.storage_path.exists():
            try:
                with open(self.storage_path, 'r') as f:
                    data = json.load(f)
                    self.rows = [CanonicalRow.from_dict(row) for row in data]
            except (json.JSONDecodeError, KeyError):
                self.rows = []
    
    def _save(self) -> None:
        """Save data to JSON storage"""
        with open(self.storage_path, 'w') as f:
            json.dump([row.to_dict() for row in self.rows], f, indent=2, default=str)
    
    def append(self, row: CanonicalRow) -> CanonicalRow:
        """
        Append a row to the table (append-only, never delete)
        
        Args:
            row: CanonicalRow to append
            
        Returns:
            The appended row with updated timestamp
        """
        row.timestamp = datetime.utcnow().isoformat()
        self.rows.append(row)
        self._save()
        return row
    
    def update_row(self, row_id: str, updates: Dict[str, Any]) -> Optional[CanonicalRow]:
        """
        Update a row by creating a new version (append-only pattern)
        The old row remains, new row is appended with updates
        
        Args:
            row_id: ID of the row to update
            updates: Dictionary of fields to update
            
        Returns:
            New row with updates, or None if not found
        """
        # Find the latest version of the row
        original = None
        for row in reversed(self.rows):
            if row.row_id == row_id:
                original = row
                break
        
        if not original:
            return None
        
        # Create new version with updates
        new_data = original.to_dict()
        new_data.update(updates)
        new_data['row_id'] = str(uuid.uuid4())  # New ID for new version
        new_data['metadata']['previous_version'] = row_id
        
        new_row = CanonicalRow.from_dict(new_data)
        return self.append(new_row)
    
    def get_row(self, row_id: str) -> Optional[CanonicalRow]:
        """Get a row by ID"""
        for row in reversed(self.rows):
            if row.row_id == row_id:
                return row
        return None
    
    def get_mature_rows(self) -> List[CanonicalRow]:
        """Get all rows with MATURE maturity state (eligible for ML)"""
        return [row for row in self.rows if row.maturity_state == MaturityState.MATURE.value]
    
    def get_rows_by_status(self, status: Status) -> List[CanonicalRow]:
        """Get all rows with specific status"""
        return [row for row in self.rows if row.status == status.value]
    
    def export_to_excel(self, path: Optional[str] = None) -> str:
        """
        Export table to Excel with color coding and frozen headers
        
        Args:
            path: Optional custom path for Excel file
            
        Returns:
            Path to the created Excel file
        """
        excel_path = Path(path) if path else self.excel_path
        excel_path.parent.mkdir(parents=True, exist_ok=True)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Canonical Table"
        
        # Apply headers with styling
        for col_num, header in enumerate(self.HEADERS, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Freeze header row
        ws.freeze_panes = ws['A2']
        
        # Add data rows with color coding
        for row_num, row in enumerate(self.rows, 2):
            row_data = row.to_dict()
            
            for col_num, header in enumerate(self.HEADERS, 1):
                # Map header to row field
                field_name = header.lower()
                value = row_data.get(field_name, '')
                
                # Handle complex types
                if isinstance(value, (dict, list)):
                    value = json.dumps(value) if value else ''
                
                cell = ws.cell(row=row_num, column=col_num, value=value)
                
                # Apply color coding
                if header == 'Status':
                    color = self.STATUS_COLORS.get(value, 'FFFFFF')
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                elif header == 'Maturity_State':
                    color = self.MATURITY_COLORS.get(value, 'FFFFFF')
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                elif header == 'ML_Confidence':
                    color = self.ML_CONFIDENCE_COLORS.get(value, 'FFFFFF')
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        # Adjust column widths
        for col_num, header in enumerate(self.HEADERS, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = max(15, len(header) + 2)
        
        wb.save(excel_path)
        return str(excel_path)
    
    def __len__(self) -> int:
        return len(self.rows)
    
    def __iter__(self):
        return iter(self.rows)


def compute_maturity_state(row: Dict[str, Any]) -> str:
    """
    Compute maturity state based on row fields
    
    Maturity Gate Rule:
    MATURE = Security=PASS AND Validation=PASS AND Artifacts!=NULL AND Status=READY
    
    Args:
        row: Dictionary containing row data
        
    Returns:
        Maturity state: IMMATURE, MATURING, or MATURE
    """
    security = row.get('security', 'PENDING')
    validation = row.get('schema_validation', row.get('validation', 'PENDING'))
    artifacts = row.get('artifacts')
    status = row.get('status', 'RAW')
    
    # MATURE: All conditions met
    if (security == 'PASS' and 
        validation == 'PASS' and 
        artifacts and 
        status == 'READY'):
        return MaturityState.MATURE.value
    
    # MATURING: Security and validation passed, but not complete
    if security == 'PASS' and validation == 'PASS':
        return MaturityState.MATURING.value
    
    # IMMATURE: Anything else
    return MaturityState.IMMATURE.value


def append_row(row: Dict[str, Any], table: Optional[CanonicalTable] = None) -> CanonicalRow:
    """
    Convenience function to append a row to the canonical table
    
    Args:
        row: Dictionary containing row data
        table: Optional CanonicalTable instance (creates new if not provided)
        
    Returns:
        The appended CanonicalRow
    """
    if table is None:
        table = CanonicalTable()
    
    canonical_row = CanonicalRow.from_dict(row)
    return table.append(canonical_row)


# Singleton instance for global access
_global_table: Optional[CanonicalTable] = None

def get_global_table() -> CanonicalTable:
    """Get or create the global canonical table instance"""
    global _global_table
    if _global_table is None:
        _global_table = CanonicalTable()
    return _global_table
