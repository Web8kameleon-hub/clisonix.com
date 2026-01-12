"""
Live Excel Dashboard - Basic Version
Monitors pipeline_output.json and updates Excel with color coding.
"""

import json
import time
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Paths
excel_file = 'canonical_table_dashboard.xlsx'
json_file = 'pipeline_output.json'

# Colors for Status, Maturity, ML Confidence
status_colors = {
    'RAW': 'FFCCCC',      # Light red
    'TEST': 'FFF2CC',     # Light yellow
    'READY': 'CCFFCC',    # Light green
    'FAIL': 'FF6666'      # Red
}

maturity_colors = {
    'IMMATURE': 'FFCCCC',  # Light red
    'MATURING': 'FFF2CC',  # Light yellow
    'MATURE': 'CCFFCC'     # Light green
}

ml_conf_colors = {
    'LOW': 'FFCCCC',       # Light red
    'MEDIUM': 'FFF2CC',    # Light yellow
    'HIGH': 'CCFFCC'       # Light green
}


def get_fill(value: str, color_map: dict) -> PatternFill:
    """Get PatternFill for a value from color map."""
    color = color_map.get(value, 'FFFFFF')
    return PatternFill(start_color=color, end_color=color, fill_type='solid')


def run_live_dashboard(excel_path: str = None, json_path: str = None):
    """
    Run live Excel dashboard.
    Monitors JSON file for changes and updates Excel.
    """
    excel_path = excel_path or excel_file
    json_path = json_path or json_file
    
    # Load workbook
    wb = load_workbook(excel_path)
    ws = wb['Canonical Table']
    headers = [cell.value for cell in ws[1]]
    
    # Track last modified time
    last_mtime = 0
    
    print("Starting live Excel dashboard. Press Ctrl+C to stop.")
    print(f"Watching: {json_path}")
    print(f"Updating: {excel_path}")
    
    try:
        while True:
            if os.path.exists(json_path):
                mtime = os.path.getmtime(json_path)
                if mtime != last_mtime:
                    last_mtime = mtime
                    
                    with open(json_path) as f:
                        data = json.load(f)
                    
                    # Handle single dict or list of dicts
                    if isinstance(data, dict):
                        data = [data]
                    
                    next_row = ws.max_row + 1
                    
                    for row_data in data:
                        for col_num, header in enumerate(headers, 1):
                            value = row_data.get(header, '')
                            cell = ws.cell(row=next_row, column=col_num, value=value)
                            
                            # Apply color coding based on column
                            if header == 'Status':
                                cell.fill = get_fill(value, status_colors)
                            elif header == 'Maturity_State':
                                cell.fill = get_fill(value, maturity_colors)
                            elif header == 'ML_Confidence':
                                cell.fill = get_fill(value, ml_conf_colors)
                        
                        next_row += 1
                    
                    wb.save(excel_path)
                    print(f"Updated Excel with {len(data)} new row(s)")
            
            time.sleep(1)  # Check every second
            
    except KeyboardInterrupt:
        print("\nLive dashboard stopped.")
        wb.save(excel_path)


if __name__ == '__main__':
    import argparse
    
    parser = argparse.ArgumentParser(description='Live Excel Dashboard')
    parser.add_argument('--excel', type=str, default=excel_file, 
                        help='Path to Excel file')
    parser.add_argument('--json', type=str, default=json_file,
                        help='Path to JSON output file')
    
    args = parser.parse_args()
    run_live_dashboard(args.excel, args.json)
