"""
Live Excel Dashboard with Alerts
Monitors pipeline output, applies color coding, and triggers console alerts
for FAIL status and high ML anomaly probability.
"""

import json
import time
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.chart import BarChart, Reference, PieChart

# Paths
excel_file = 'canonical_table_dashboard.xlsx'
json_file = 'pipeline_output.json'

# Colors - FAIL is bright red for alert
status_colors = {
    'RAW': 'FFCCCC',
    'TEST': 'FFF2CC',
    'READY': 'CCFFCC',
    'FAIL': 'FF0000'  # Bright red for alert
}

maturity_colors = {
    'IMMATURE': 'FFCCCC',
    'MATURING': 'FFF2CC',
    'MATURE': 'CCFFCC'
}

ml_conf_colors = {
    'LOW': 'FFCCCC',
    'MEDIUM': 'FFF2CC',
    'HIGH': 'CCFFCC'
}

# Threshold for ML anomaly alert
ML_ANOMALY_THRESHOLD = 0.7


def get_fill(value: str, color_map: dict) -> PatternFill:
    """Get PatternFill for a value from color map."""
    color = color_map.get(value, 'FFFFFF')
    return PatternFill(start_color=color, end_color=color, fill_type='solid')


def print_alert(message: str, alert_type: str = "WARNING"):
    """Print formatted alert message."""
    colors = {
        "WARNING": "\033[93m",  # Yellow
        "ERROR": "\033[91m",    # Red
        "INFO": "\033[94m",     # Blue
        "RESET": "\033[0m"
    }
    color = colors.get(alert_type, colors["INFO"])
    reset = colors["RESET"]
    print(f"{color}[{alert_type}] {message}{reset}")


def run_live_dashboard_with_alerts(excel_path: str = None, 
                                    json_path: str = None,
                                    anomaly_threshold: float = None):
    """
    Run live Excel dashboard with alert system.
    
    Args:
        excel_path: Path to Excel dashboard file
        json_path: Path to JSON output file  
        anomaly_threshold: Threshold for ML anomaly alerts (default 0.7)
    """
    excel_path = excel_path or excel_file
    json_path = json_path or json_file
    threshold = anomaly_threshold or ML_ANOMALY_THRESHOLD
    
    # Load workbook
    wb = load_workbook(excel_path)
    ws = wb['Canonical Table']
    headers = [cell.value for cell in ws[1]]
    
    # Track last modified time
    last_mtime = 0
    
    # Alert counters
    fail_count = 0
    anomaly_count = 0
    
    print("=" * 60)
    print("LIVE EXCEL DASHBOARD WITH ALERTS")
    print("=" * 60)
    print(f"Watching: {json_path}")
    print(f"Updating: {excel_path}")
    print(f"ML Anomaly Threshold: {threshold}")
    print("Press Ctrl+C to stop.")
    print("=" * 60)
    
    try:
        while True:
            if os.path.exists(json_path):
                mtime = os.path.getmtime(json_path)
                if mtime != last_mtime:
                    last_mtime = mtime
                    
                    with open(json_path) as f:
                        data = json.load(f)
                    
                    if isinstance(data, dict):
                        data = [data]
                    
                    next_row = ws.max_row + 1
                    
                    for row_data in data:
                        for col_num, header in enumerate(headers, 1):
                            value = row_data.get(header, '')
                            cell = ws.cell(row=next_row, column=col_num, value=value)
                            
                            # Color coding with alerts
                            if header == 'Status':
                                cell.fill = get_fill(value, status_colors)
                                if value == 'FAIL':
                                    cell.font = Font(bold=True, color='FFFFFF')
                                    fail_count += 1
                                    print_alert(
                                        f"Row {next_row} Status = FAIL! (Total fails: {fail_count})",
                                        "ERROR"
                                    )
                                    
                            elif header == 'Maturity_State':
                                cell.fill = get_fill(value, maturity_colors)
                                
                            elif header == 'ML_Confidence':
                                cell.fill = get_fill(value, ml_conf_colors)
                                
                            elif header == 'ML_Anomaly_Prob' and value != '':
                                try:
                                    prob = float(value)
                                    if prob >= threshold:
                                        # Orange highlight for anomaly
                                        cell.fill = PatternFill(
                                            start_color='FF9900',
                                            end_color='FF9900',
                                            fill_type='solid'
                                        )
                                        anomaly_count += 1
                                        print_alert(
                                            f"Row {next_row} ML Anomaly probability = {prob:.3f} "
                                            f"(exceeds threshold {threshold}) "
                                            f"(Total anomalies: {anomaly_count})",
                                            "WARNING"
                                        )
                                except ValueError:
                                    pass
                        
                        next_row += 1
                    
                    # Save workbook after new data
                    wb.save(excel_path)
                    print(f"Updated Excel dashboard with {len(data)} new row(s) and alerts applied.")
            
            time.sleep(1)
            
    except KeyboardInterrupt:
        print("\n" + "=" * 60)
        print("DASHBOARD SUMMARY")
        print("=" * 60)
        print(f"Total FAIL alerts: {fail_count}")
        print(f"Total ML Anomaly alerts: {anomaly_count}")
        print("Live Excel dashboard stopped.")
        wb.save(excel_path)


if __name__ == '__main__':
    import argparse
    
    parser = argparse.ArgumentParser(description='Live Excel Dashboard with Alerts')
    parser.add_argument('--excel', type=str, default=excel_file,
                        help='Path to Excel file')
    parser.add_argument('--json', type=str, default=json_file,
                        help='Path to JSON output file')
    parser.add_argument('--threshold', type=float, default=ML_ANOMALY_THRESHOLD,
                        help='ML anomaly probability threshold (default: 0.7)')
    
    args = parser.parse_args()
    run_live_dashboard_with_alerts(args.excel, args.json, args.threshold)
