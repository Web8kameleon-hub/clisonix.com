"""
Live Excel Dashboard with Charts
Monitors pipeline_output.json, updates Excel with color coding AND refreshes charts.
"""

import json
import time
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.chart import BarChart, Reference, PieChart

# Paths
excel_file = 'canonical_table_dashboard.xlsx'
json_file = 'pipeline_output.json'

# Colors
status_colors = {
    'RAW': 'FFCCCC',
    'TEST': 'FFF2CC', 
    'READY': 'CCFFCC',
    'FAIL': 'FF6680'
}

maturity_colors = {
    'IMMATURE': 'FFCCCC',
    'MATURING': 'FFF2CC',
    'MATURE': 'CCFFCC'
}

ml_conf_colors = {
    'LOW': 'FFCCCC',
    'MEDIUM': 'FFF2CC',
    'HIGH': 'CCFFCC'
}


def get_fill(value: str, color_map: dict) -> PatternFill:
    """Get PatternFill for a value from color map."""
    color = color_map.get(value, 'FFFFFF')
    return PatternFill(start_color=color, end_color=color, fill_type='solid')


def update_charts(ws, max_row: int):
    """
    Update all charts in the worksheet.
    Recalculates counts and rebuilds charts.
    """
    # --- Status Bar Chart ---
    status_counts = {}
    for row in range(2, max_row + 1):
        val = ws.cell(row=row, column=6).value
        if val:
            status_counts[val] = status_counts.get(val, 0) + 1
    
    if not status_counts:
        return
    
    chart_start_row = max_row + 3
    
    # Write status counts
    for i, (status, count) in enumerate(status_counts.items(), start=chart_start_row):
        ws.cell(row=i, column=1, value=status)
        ws.cell(row=i, column=2, value=count)
    
    # Create Bar Chart
    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.style = 10
    bar_chart.title = "Status Counts"
    
    data_ref = Reference(ws, min_col=2, min_row=chart_start_row, 
                        max_row=chart_start_row + len(status_counts) - 1)
    cat_ref = Reference(ws, min_col=1, min_row=chart_start_row, 
                       max_row=chart_start_row + len(status_counts) - 1)
    bar_chart.add_data(data_ref, titles_from_data=False)
    bar_chart.set_categories(cat_ref)
    ws.add_chart(bar_chart, "H2")
    
    # --- Maturity Pie Chart ---
    maturity_counts = {}
    for row in range(2, max_row + 1):
        val = ws.cell(row=row, column=12).value
        if val:
            maturity_counts[val] = maturity_counts.get(val, 0) + 1
    
    if maturity_counts:
        pie_start_row = chart_start_row + len(status_counts) + 2
        
        for i, (maturity, count) in enumerate(maturity_counts.items(), start=pie_start_row):
            ws.cell(row=i, column=1, value=maturity)
            ws.cell(row=i, column=2, value=count)
        
        pie_chart = PieChart()
        pie_chart.title = "Maturity Distribution"
        
        data_ref = Reference(ws, min_col=2, min_row=pie_start_row, 
                            max_row=pie_start_row + len(maturity_counts) - 1)
        cat_ref = Reference(ws, min_col=1, min_row=pie_start_row, 
                           max_row=pie_start_row + len(maturity_counts) - 1)
        pie_chart.add_data(data_ref, titles_from_data=False)
        pie_chart.set_categories(cat_ref)
        ws.add_chart(pie_chart, "H20")
    
    # --- ML Confidence Pie Chart ---
    ml_counts = {}
    for row in range(2, max_row + 1):
        val = ws.cell(row=row, column=14).value
        if val:
            ml_counts[val] = ml_counts.get(val, 0) + 1
    
    if ml_counts:
        ml_start_row = pie_start_row + len(maturity_counts) + 2
        
        for i, (conf, count) in enumerate(ml_counts.items(), start=ml_start_row):
            ws.cell(row=i, column=1, value=conf)
            ws.cell(row=i, column=2, value=count)
        
        ml_chart = PieChart()
        ml_chart.title = "ML Confidence Distribution"
        
        data_ref = Reference(ws, min_col=2, min_row=ml_start_row, 
                            max_row=ml_start_row + len(ml_counts) - 1)
        cat_ref = Reference(ws, min_col=1, min_row=ml_start_row, 
                           max_row=ml_start_row + len(ml_counts) - 1)
        ml_chart.add_data(data_ref, titles_from_data=False)
        ml_chart.set_categories(cat_ref)
        ws.add_chart(ml_chart, "H40")


def run_live_dashboard_with_charts(excel_path: str = None, json_path: str = None):
    """
    Run live Excel dashboard with automatic chart updates.
    """
    excel_path = excel_path or excel_file
    json_path = json_path or json_file
    
    # Load workbook
    wb = load_workbook(excel_path)
    ws = wb['Canonical Table']
    headers = [cell.value for cell in ws[1]]
    
    # Track last modified time
    last_mtime = 0
    
    print("Starting live Excel dashboard with charts. Press Ctrl+C to stop.")
    print(f"Watching: {json_path}")
    print(f"Updating: {excel_path}")
    
    try:
        while True:
            if os.path.exists(json_path):
                mtime = os.path.getmtime(json_path)
                if mtime != last_mtime:
                    last_mtime = mtime
                    
                    with open(json_path) as f:
                        data = json.load(f)
                    
                    if isinstance(data, dict):
                        data = [data]
                    
                    next_row = ws.max_row + 1
                    
                    # Append new rows with color coding
                    for row_data in data:
                        for col_num, header in enumerate(headers, 1):
                            value = row_data.get(header, '')
                            cell = ws.cell(row=next_row, column=col_num, value=value)
                            
                            if header == 'Status':
                                cell.fill = get_fill(value, status_colors)
                            elif header == 'Maturity_State':
                                cell.fill = get_fill(value, maturity_colors)
                            elif header == 'ML_Confidence':
                                cell.fill = get_fill(value, ml_conf_colors)
                        
                        next_row += 1
                    
                    # Update charts with new data
                    update_charts(ws, ws.max_row)
                    
                    wb.save(excel_path)
                    print(f"Updated Excel dashboard with {len(data)} new row(s) and refreshed charts.")
            
            time.sleep(1)
            
    except KeyboardInterrupt:
        print("\nLive Excel dashboard stopped.")
        wb.save(excel_path)


if __name__ == '__main__':
    import argparse
    
    parser = argparse.ArgumentParser(description='Live Excel Dashboard with Charts')
    parser.add_argument('--excel', type=str, default=excel_file,
                        help='Path to Excel file')
    parser.add_argument('--json', type=str, default=json_file,
                        help='Path to JSON output file')
    
    args = parser.parse_args()
    run_live_dashboard_with_charts(args.excel, args.json)
