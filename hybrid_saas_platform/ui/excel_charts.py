"""
Excel Charts Generator
Creates Bar and Pie charts for Status, Maturity, and ML Confidence distributions.
"""

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference, PieChart


def create_dashboard_charts(excel_path: str = 'canonical_table_dashboard.xlsx',
                            output_path: str = 'canonical_table_with_charts.xlsx'):
    """
    Create Excel dashboard with charts from canonical table.
    
    Args:
        excel_path: Path to source Excel file
        output_path: Path to save dashboard with charts
    """
    # Load workbook
    wb = load_workbook(excel_path)
    ws = wb['Canonical Table']
    max_row = ws.max_row
    
    # --- 1. Bar Chart for Status Counts ---
    status_col = 6  # Status column (F)
    status_counts = {}
    
    # Count occurrences
    for row in range(2, max_row + 1):
        value = ws.cell(row=row, column=status_col).value
        if value:
            status_counts[value] = status_counts.get(value, 0) + 1
    
    if status_counts:
        # Write counts to a hidden table for chart
        chart_start_row = max_row + 3
        for i, (status, count) in enumerate(status_counts.items(), start=chart_start_row):
            ws.cell(row=i, column=1, value=status)
            ws.cell(row=i, column=2, value=count)
        
        # Create BarChart
        bar_chart = BarChart()
        bar_chart.type = "col"
        bar_chart.style = 10
        bar_chart.title = "Status Counts"
        bar_chart.y_axis.title = "Count"
        bar_chart.x_axis.title = "Status"
        
        data = Reference(ws, min_col=2, min_row=chart_start_row, 
                        max_row=chart_start_row + len(status_counts) - 1)
        cats = Reference(ws, min_col=1, min_row=chart_start_row, 
                        max_row=chart_start_row + len(status_counts) - 1)
        bar_chart.add_data(data, titles_from_data=False)
        bar_chart.set_categories(cats)
        bar_chart.shape = 4
        ws.add_chart(bar_chart, "H2")
        
        print(f"Status chart created with {len(status_counts)} categories")
    
    # --- 2. Pie Chart for Maturity_State ---
    maturity_col = 12  # Maturity_State column (L)
    maturity_counts = {}
    
    for row in range(2, max_row + 1):
        value = ws.cell(row=row, column=maturity_col).value
        if value:
            maturity_counts[value] = maturity_counts.get(value, 0) + 1
    
    if maturity_counts:
        # Write counts for pie chart
        pie_start_row = chart_start_row + len(status_counts) + 2
        for i, (maturity, count) in enumerate(maturity_counts.items(), start=pie_start_row):
            ws.cell(row=i, column=1, value=maturity)
            ws.cell(row=i, column=2, value=count)
        
        # Create PieChart
        pie_chart = PieChart()
        pie_chart.title = "Maturity Distribution"
        
        data = Reference(ws, min_col=2, min_row=pie_start_row, 
                        max_row=pie_start_row + len(maturity_counts) - 1)
        cats = Reference(ws, min_col=1, min_row=pie_start_row, 
                        max_row=pie_start_row + len(maturity_counts) - 1)
        pie_chart.add_data(data, titles_from_data=False)
        pie_chart.set_categories(cats)
        ws.add_chart(pie_chart, "H20")
        
        print(f"Maturity chart created with {len(maturity_counts)} categories")
    
    # --- 3. Pie Chart for ML_Confidence ---
    ml_conf_col = 14  # ML_Confidence column (N)
    ml_counts = {}
    
    for row in range(2, max_row + 1):
        value = ws.cell(row=row, column=ml_conf_col).value
        if value:
            ml_counts[value] = ml_counts.get(value, 0) + 1
    
    if ml_counts:
        ml_start_row = pie_start_row + len(maturity_counts) + 2
        for i, (conf, count) in enumerate(ml_counts.items(), start=ml_start_row):
            ws.cell(row=i, column=1, value=conf)
            ws.cell(row=i, column=2, value=count)
        
        # Create ML Confidence Pie Chart
        ml_chart = PieChart()
        ml_chart.title = "ML Confidence Distribution"
        
        data = Reference(ws, min_col=2, min_row=ml_start_row, 
                        max_row=ml_start_row + len(ml_counts) - 1)
        cats = Reference(ws, min_col=1, min_row=ml_start_row, 
                        max_row=ml_start_row + len(ml_counts) - 1)
        ml_chart.add_data(data, titles_from_data=False)
        ml_chart.set_categories(cats)
        ws.add_chart(ml_chart, "H40")
        
        print(f"ML Confidence chart created with {len(ml_counts)} categories")
    
    # Save workbook
    wb.save(output_path)
    print(f"\nExcel dashboard with charts saved to: {output_path}")
    
    return output_path


if __name__ == '__main__':
    import argparse
    
    parser = argparse.ArgumentParser(description='Create Excel Dashboard with Charts')
    parser.add_argument('--input', type=str, default='canonical_table_template.xlsx',
                        help='Path to source Excel file')
    parser.add_argument('--output', type=str, default='canonical_table_dashboard.xlsx',
                        help='Path to output dashboard file')
    
    args = parser.parse_args()
    create_dashboard_charts(args.input, args.output)
