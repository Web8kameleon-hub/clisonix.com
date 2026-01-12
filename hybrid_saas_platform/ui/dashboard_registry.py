"""
🎯 CLISONIX DASHBOARD REGISTRY
Ultra-Professional Dashboard Management System

Çdo dashboard që krijohet regjistrohet automatikisht në një tabelë qendrore Excel.
Tabela shërben si "Control Center" për të gjitha dashboard-et.

Features:
- Auto-registration i dashboard-eve
- Links direkte tek çdo dashboard
- Metadata (data krijimit, lloji, statusi)
- Ultra-wide columns (500-600px)
- Full wrap text support
- 600px row heights
"""

import os
import json
from datetime import datetime
from pathlib import Path

# Provoj openpyxl
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side,
        NamedStyle
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image
    from openpyxl.chart import BarChart, PieChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ openpyxl not installed. Run: pip install openpyxl")


# ============================================================
# CONFIGURATION
# ============================================================

REGISTRY_FILE = "Dashboard_Registry.xlsx"
DASHBOARDS_FOLDER = "dashboards"

# Ultra-wide column widths (in Excel units, ~7px per unit)
COLUMN_CONFIG = {
    'A': {'name': 'ID', 'width': 8},
    'B': {'name': 'Dashboard_Name', 'width': 50},
    'C': {'name': 'Type', 'width': 25},
    'D': {'name': 'Description', 'width': 70},  # ~500px
    'E': {'name': 'File_Path', 'width': 85},    # ~600px
    'F': {'name': 'Created_Date', 'width': 25},
    'G': {'name': 'Last_Updated', 'width': 25},
    'H': {'name': 'Status', 'width': 15},
    'I': {'name': 'Data_Source', 'width': 85},  # ~600px
    'J': {'name': 'Sample_Output', 'width': 85}, # ~600px
    'K': {'name': 'Notes', 'width': 85},         # ~600px
}

# Row height - AUTO (wrap text handles it)
ROW_HEIGHT = None  # Auto-adjust based on content

# Styles
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=12, name='Segoe UI')
CELL_FONT = Font(size=11, name='Consolas')
LINK_FONT = Font(size=11, name='Segoe UI', color='0563C1', underline='single')

CELL_ALIGNMENT = Alignment(
    horizontal='left',
    vertical='top',
    wrap_text=True,
    shrink_to_fit=False
)

HEADER_ALIGNMENT = Alignment(
    horizontal='center',
    vertical='center',
    wrap_text=True
)

THIN_BORDER = Border(
    left=Side(style='thin', color='B4B4B4'),
    right=Side(style='thin', color='B4B4B4'),
    top=Side(style='thin', color='B4B4B4'),
    bottom=Side(style='thin', color='B4B4B4')
)

# Status colors
STATUS_COLORS = {
    'Active': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
    'Draft': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
    'Archived': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
    'Testing': PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid'),
}

# Dashboard types
DASHBOARD_TYPES = {
    'live': '📊 Live Dashboard',
    'charts': '📈 Charts Dashboard',
    'alerts': '🚨 Alerts Dashboard',
    'api': '🔌 API Dashboard',
    'analytics': '📉 Analytics Dashboard',
    'monitoring': '🖥️ Monitoring Dashboard',
}


# ============================================================
# REGISTRY MANAGER CLASS
# ============================================================

class DashboardRegistry:
    """
    Manages the central dashboard registry Excel file.
    Auto-registers dashboards and maintains links.
    """
    
    def __init__(self, registry_path: str = None):
        self.base_path = Path(__file__).parent.parent.parent
        self.registry_path = registry_path or self.base_path / REGISTRY_FILE
        self.dashboards_path = self.base_path / DASHBOARDS_FOLDER
        
        # Ensure dashboards folder exists
        self.dashboards_path.mkdir(exist_ok=True)
        
        # Initialize or load registry
        if not os.path.exists(self.registry_path):
            self._create_registry()
        
    def _create_registry(self):
        """Creates the master registry Excel file with ultra-wide columns."""
        if not OPENPYXL_AVAILABLE:
            print("❌ Cannot create registry: openpyxl not available")
            return
            
        wb = Workbook()
        
        # ========== SHEET 1: Dashboard Index ==========
        ws = wb.active
        ws.title = "Dashboard_Index"
        
        # Set column widths
        for col_letter, config in COLUMN_CONFIG.items():
            ws.column_dimensions[col_letter].width = config['width']
            
        # Create header row
        for col_idx, (col_letter, config) in enumerate(COLUMN_CONFIG.items(), 1):
            cell = ws.cell(row=1, column=col_idx, value=config['name'])
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER
        
        # Set header row height
        ws.row_dimensions[1].height = 40
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # ========== SHEET 2: Dashboard Links ==========
        ws_links = wb.create_sheet("Quick_Links")
        ws_links.column_dimensions['A'].width = 10
        ws_links.column_dimensions['B'].width = 50
        ws_links.column_dimensions['C'].width = 100
        
        # Header
        for idx, header in enumerate(['#', 'Dashboard', 'Click to Open'], 1):
            cell = ws_links.cell(row=1, column=idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER
        
        # ========== SHEET 3: Statistics ==========
        ws_stats = wb.create_sheet("Statistics")
        ws_stats['A1'] = "📊 Dashboard Statistics"
        ws_stats['A1'].font = Font(bold=True, size=16)
        
        ws_stats['A3'] = "Total Dashboards:"
        ws_stats['B3'] = 0
        ws_stats['A4'] = "Active:"
        ws_stats['B4'] = 0
        ws_stats['A5'] = "Last Updated:"
        ws_stats['B5'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Save
        wb.save(self.registry_path)
        print(f"✅ Dashboard Registry created: {self.registry_path}")
        
    def register_dashboard(
        self,
        name: str,
        dashboard_type: str,
        description: str,
        file_path: str,
        data_source: str = "",
        sample_output: str = "",
        notes: str = "",
        status: str = "Active"
    ) -> int:
        """
        Registers a new dashboard in the registry.
        Returns the dashboard ID.
        """
        if not OPENPYXL_AVAILABLE:
            print("❌ Cannot register: openpyxl not available")
            return -1
            
        wb = load_workbook(self.registry_path)
        ws = wb["Dashboard_Index"]
        
        # Find next empty row
        next_row = ws.max_row + 1
        dashboard_id = next_row - 1  # ID starts from 1
        
        # Prepare data
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        type_display = DASHBOARD_TYPES.get(dashboard_type, f"📋 {dashboard_type}")
        
        row_data = [
            dashboard_id,           # A: ID
            name,                   # B: Dashboard_Name
            type_display,           # C: Type
            description,            # D: Description
            file_path,              # E: File_Path
            now,                    # F: Created_Date
            now,                    # G: Last_Updated
            status,                 # H: Status
            data_source,            # I: Data_Source
            sample_output,          # J: Sample_Output
            notes,                  # K: Notes
        ]
        
        # Write row
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=next_row, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER
            
            # Status color
            if col_idx == 8:  # Status column
                if status in STATUS_COLORS:
                    cell.fill = STATUS_COLORS[status]
            
            # Make file path a hyperlink
            if col_idx == 5 and file_path:  # File_Path column
                cell.font = LINK_FONT
                try:
                    cell.hyperlink = file_path
                except:
                    pass  # Hyperlink may fail for some paths
        
        # Set row height (600px ≈ 450 points)
        ws.row_dimensions[next_row].height = ROW_HEIGHT
        
        # Update Quick_Links sheet
        ws_links = wb["Quick_Links"]
        link_row = next_row
        ws_links.cell(row=link_row, column=1, value=dashboard_id)
        ws_links.cell(row=link_row, column=2, value=name)
        link_cell = ws_links.cell(row=link_row, column=3, value=f"📂 {file_path}")
        link_cell.font = LINK_FONT
        try:
            link_cell.hyperlink = file_path
        except:
            pass
        
        # Update Statistics
        ws_stats = wb["Statistics"]
        ws_stats['B3'] = dashboard_id  # Total
        ws_stats['B5'] = now  # Last Updated
        
        # Count active
        active_count = sum(
            1 for row in ws.iter_rows(min_row=2, max_col=8)
            if row[7].value == "Active"
        )
        ws_stats['B4'] = active_count
        
        # Save
        wb.save(self.registry_path)
        
        print(f"✅ Dashboard registered: #{dashboard_id} - {name}")
        return dashboard_id
    
    def list_dashboards(self) -> list:
        """Returns a list of all registered dashboards."""
        if not OPENPYXL_AVAILABLE:
            return []
            
        wb = load_workbook(self.registry_path)
        ws = wb["Dashboard_Index"]
        
        dashboards = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Has ID
                dashboards.append({
                    'id': row[0],
                    'name': row[1],
                    'type': row[2],
                    'description': row[3],
                    'file_path': row[4],
                    'created': row[5],
                    'updated': row[6],
                    'status': row[7],
                })
        
        return dashboards
    
    def update_dashboard(self, dashboard_id: int, **kwargs):
        """Updates a dashboard's metadata."""
        if not OPENPYXL_AVAILABLE:
            return False
            
        wb = load_workbook(self.registry_path)
        ws = wb["Dashboard_Index"]
        
        # Find row by ID
        target_row = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
            if row[0].value == dashboard_id:
                target_row = row_idx
                break
        
        if not target_row:
            print(f"❌ Dashboard #{dashboard_id} not found")
            return False
        
        # Update fields
        field_map = {
            'name': 2, 'type': 3, 'description': 4, 'file_path': 5,
            'status': 8, 'data_source': 9, 'sample_output': 10, 'notes': 11
        }
        
        for field, value in kwargs.items():
            if field in field_map:
                ws.cell(row=target_row, column=field_map[field], value=value)
        
        # Update last_updated
        ws.cell(row=target_row, column=7, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        
        wb.save(self.registry_path)
        print(f"✅ Dashboard #{dashboard_id} updated")
        return True


# ============================================================
# DASHBOARD CREATOR - Creates dashboards and auto-registers
# ============================================================

def create_data_dashboard(
    name: str,
    data_source: str,
    output_path: str = None,
    dashboard_type: str = "analytics",
    description: str = "",
    auto_register: bool = True
) -> str:
    """
    Creates a new data dashboard Excel file with ultra-wide columns
    and automatically registers it in the registry.
    
    Args:
        name: Dashboard name
        data_source: Path to data file (JSON, CSV, etc.)
        output_path: Where to save the dashboard
        dashboard_type: Type of dashboard
        description: Dashboard description
        auto_register: Whether to register in the registry
    
    Returns:
        Path to created dashboard
    """
    if not OPENPYXL_AVAILABLE:
        print("❌ openpyxl not available")
        return None
    
    # Determine output path
    base_path = Path(__file__).parent.parent.parent
    dashboards_folder = base_path / DASHBOARDS_FOLDER
    dashboards_folder.mkdir(exist_ok=True)
    
    if not output_path:
        safe_name = name.replace(" ", "_").replace("/", "-")
        output_path = dashboards_folder / f"{safe_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    
    # Ultra-wide column configuration for dashboard data
    dashboard_columns = {
        'A': {'name': 'ID', 'width': 10},
        'B': {'name': 'Metric', 'width': 40},
        'C': {'name': 'Value', 'width': 30},
        'D': {'name': 'Description', 'width': 70},      # ~500px
        'E': {'name': 'Details', 'width': 85},          # ~600px
        'F': {'name': 'Sample_Response', 'width': 85},  # ~600px
        'G': {'name': 'cURL_Command', 'width': 85},     # ~600px
        'H': {'name': 'Python_Code', 'width': 85},      # ~600px
        'I': {'name': 'Timestamp', 'width': 25},
        'J': {'name': 'Status', 'width': 15},
    }
    
    # Set column widths
    for col_letter, config in dashboard_columns.items():
        ws.column_dimensions[col_letter].width = config['width']
    
    # Create header row
    for col_idx, (col_letter, config) in enumerate(dashboard_columns.items(), 1):
        cell = ws.cell(row=1, column=col_idx, value=config['name'])
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = 'A2'
    
    # Load data if provided
    sample_output = ""
    if data_source and os.path.exists(data_source):
        try:
            if data_source.endswith('.json'):
                with open(data_source, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    
                # Convert JSON to rows
                if isinstance(data, list):
                    for idx, item in enumerate(data, 1):
                        row_num = idx + 1
                        
                        # Write data
                        ws.cell(row=row_num, column=1, value=idx)
                        
                        if isinstance(item, dict):
                            for col_idx, (key, value) in enumerate(list(item.items())[:9], 2):
                                cell = ws.cell(row=row_num, column=col_idx)
                                
                                # Format value
                                if isinstance(value, (dict, list)):
                                    cell.value = json.dumps(value, indent=2, ensure_ascii=False)
                                else:
                                    cell.value = str(value)
                                
                                cell.font = CELL_FONT
                                cell.alignment = CELL_ALIGNMENT
                                cell.border = THIN_BORDER
                        
                        # Set row height
                        ws.row_dimensions[row_num].height = ROW_HEIGHT
                
                sample_output = json.dumps(data[:3] if isinstance(data, list) else data, indent=2)[:500]
                
        except Exception as e:
            print(f"⚠️ Could not load data: {e}")
    
    # Add metadata sheet
    ws_meta = wb.create_sheet("Metadata")
    ws_meta['A1'] = "Dashboard Information"
    ws_meta['A1'].font = Font(bold=True, size=14)
    
    metadata = [
        ("Name", name),
        ("Type", dashboard_type),
        ("Description", description),
        ("Data Source", data_source),
        ("Created", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Author", "Clisonix Cloud System"),
    ]
    
    for idx, (key, value) in enumerate(metadata, 3):
        ws_meta.cell(row=idx, column=1, value=key).font = Font(bold=True)
        ws_meta.cell(row=idx, column=2, value=value)
    
    ws_meta.column_dimensions['A'].width = 20
    ws_meta.column_dimensions['B'].width = 80
    
    # Save dashboard
    wb.save(output_path)
    print(f"✅ Dashboard created: {output_path}")
    
    # Auto-register
    if auto_register:
        registry = DashboardRegistry()
        registry.register_dashboard(
            name=name,
            dashboard_type=dashboard_type,
            description=description or f"Data dashboard from {data_source}",
            file_path=str(output_path),
            data_source=str(data_source) if data_source else "",
            sample_output=sample_output,
            status="Active"
        )
    
    return str(output_path)


# ============================================================
# BATCH REGISTER EXISTING DASHBOARDS
# ============================================================

def register_existing_dashboards():
    """
    Scans for existing dashboard files and registers them.
    """
    registry = DashboardRegistry()
    base_path = Path(__file__).parent.parent.parent
    ui_path = Path(__file__).parent
    
    # Known dashboards from ui folder
    dashboards_to_register = [
        {
            'name': 'Live Dashboard',
            'type': 'live',
            'description': 'Real-time Excel dashboard that monitors pipeline_output.json with automatic color coding based on Status (PASS/FAIL), Maturity levels, and ML Confidence scores.',
            'file': ui_path / 'live_dashboard.py',
        },
        {
            'name': 'Excel Charts Dashboard',
            'type': 'charts',
            'description': 'Bar and Pie charts dashboard showing Status distribution, Maturity levels breakdown, and ML Confidence histogram.',
            'file': ui_path / 'excel_charts.py',
        },
        {
            'name': 'Live Dashboard with Charts',
            'type': 'live',
            'description': 'Combined live monitoring with auto-refreshing charts. Updates every 5 seconds.',
            'file': ui_path / 'live_dashboard_charts.py',
        },
        {
            'name': 'Alerts Dashboard',
            'type': 'alerts',
            'description': 'Console-based alert system. Triggers warnings for FAIL status or ML Anomaly detection.',
            'file': ui_path / 'live_dashboard_alerts.py',
        },
        {
            'name': 'XLWings Pop-up Dashboard',
            'type': 'alerts',
            'description': 'Windows pop-up notification dashboard using xlwings. Requires Excel to be installed.',
            'file': ui_path / 'xlwings_dashboard.py',
        },
        {
            'name': 'XLWings Dashboard with Log',
            'type': 'alerts',
            'description': 'Pop-up alerts with Alert_Log sheet. Maintains timestamp history of all alerts triggered.',
            'file': ui_path / 'xlwings_dashboard_log.py',
        },
        {
            'name': 'API Enricher',
            'type': 'api',
            'description': 'Enriches API canonical table with complete documentation including Sample_Response, cURL commands, and Python snippets for all 28+ Clisonix endpoints.',
            'file': ui_path / 'api_enricher.py',
        },
        {
            'name': 'Postman Merger',
            'type': 'api',
            'description': 'Merges multiple Postman collections into one mega-collection and exports to Excel canonical table format.',
            'file': ui_path / 'postman_merger.py',
        },
    ]
    
    registered_count = 0
    for dash in dashboards_to_register:
        if dash['file'].exists():
            registry.register_dashboard(
                name=dash['name'],
                dashboard_type=dash['type'],
                description=dash['description'],
                file_path=str(dash['file']),
                data_source="pipeline_output.json / Postman Collections",
                status="Active"
            )
            registered_count += 1
    
    print(f"\n🎯 Registered {registered_count} dashboards in the registry")
    print(f"📂 Registry file: {registry.registry_path}")
    
    return registered_count


# ============================================================
# MAIN
# ============================================================

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Clisonix Dashboard Registry")
    parser.add_argument('--init', action='store_true', help='Initialize registry and register existing dashboards')
    parser.add_argument('--list', action='store_true', help='List all registered dashboards')
    parser.add_argument('--create', type=str, help='Create new dashboard with given name')
    parser.add_argument('--data', type=str, help='Data source for new dashboard')
    parser.add_argument('--open', action='store_true', help='Open registry after operation')
    
    args = parser.parse_args()
    
    if args.init:
        print("🚀 Initializing Dashboard Registry...")
        registry = DashboardRegistry()
        register_existing_dashboards()
        
        if args.open:
            import subprocess
            subprocess.run(['start', str(registry.registry_path)], shell=True)
    
    elif args.list:
        registry = DashboardRegistry()
        dashboards = registry.list_dashboards()
        
        print("\n📊 REGISTERED DASHBOARDS")
        print("=" * 80)
        for dash in dashboards:
            print(f"  #{dash['id']:2} | {dash['name']:30} | {dash['type']:20} | {dash['status']}")
        print("=" * 80)
        print(f"Total: {len(dashboards)} dashboards")
    
    elif args.create:
        dashboard_path = create_data_dashboard(
            name=args.create,
            data_source=args.data or "",
            dashboard_type="analytics",
            description=f"Custom dashboard: {args.create}"
        )
        
        if args.open and dashboard_path:
            import subprocess
            subprocess.run(['start', dashboard_path], shell=True)
    
    else:
        # Default: init and open
        print("🚀 Clisonix Dashboard Registry")
        print("=" * 50)
        print("Usage:")
        print("  --init     Initialize and register dashboards")
        print("  --list     List all dashboards")
        print("  --create   Create new dashboard")
        print("  --open     Open file after operation")
        print()
        print("Example:")
        print("  python dashboard_registry.py --init --open")
