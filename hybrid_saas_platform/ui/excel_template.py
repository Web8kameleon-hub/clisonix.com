"""
Excel Template Generator
========================
Creates the canonical table Excel template with color coding.
Based on the user's openpyxl specification.
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from typing import List, Dict, Any


def create_canonical_table_template(output_path: str = "canonical_table_template.xlsx",
                                   sample_rows: List[Dict[str, Any]] = None) -> str:
    """
    Create the canonical table Excel template with:
    - Frozen header row
    - Color coding for Status, Maturity, ML Confidence
    - Professional styling
    
    Args:
        output_path: Path for the output Excel file
        sample_rows: Optional sample data rows
        
    Returns:
        Path to created file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Canonical Table"
    
    # Define headers
    headers = [
        'Row_ID', 'Timestamp', 'Cycle', 'Layer', 'Variant', 'Input_Type',
        'Source_Protocol', 'Status', 'Security', 'Schema_Validation',
        'Agent', 'Lab', 'Artifacts', 'Maturity_State',
        'ML_Score', 'ML_Confidence', 'ML_Suggested_Agent', 'ML_Suggested_Lab',
        'ML_Anomaly_Prob', 'Enforcement_Applied', 'Compliance_Status'
    ]
    
    # Define styles
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Color coding definitions
    status_colors = {
        'RAW': 'FFCCCC',        # Light red
        'NORMALIZED': 'FFE5CC', # Light orange
        'TEST': 'FFF2CC',       # Light yellow
        'READY': 'CCFFCC',      # Light green
        'FAIL': 'FF6666',       # Red
        'ARCHIVED': 'CCCCCC'    # Gray
    }
    
    maturity_colors = {
        'IMMATURE': 'FFCCCC',   # Light red
        'MATURING': 'FFF2CC',   # Light yellow
        'MATURE': 'CCFFCC'      # Light green
    }
    
    ml_confidence_colors = {
        'LOW': 'FFCCCC',        # Light red
        'MEDIUM': 'FFF2CC',     # Light yellow
        'HIGH': 'CCFFCC'        # Light green
    }
    
    security_colors = {
        'PASS': 'CCFFCC',       # Light green
        'FAIL': 'FF6666',       # Red
        'PENDING': 'FFF2CC'     # Light yellow
    }
    
    # Apply headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Freeze header row
    ws.freeze_panes = ws['A2']
    
    # Default sample rows if none provided
    if sample_rows is None:
        sample_rows = [
            {
                'row_id': 'demo-001',
                'timestamp': '2025-01-11T10:00:00',
                'cycle': 1,
                'layer': 'L1',
                'variant': 'A',
                'input_type': 'grain',
                'source_protocol': 'REST',
                'status': 'RAW',
                'security': 'PENDING',
                'schema_validation': 'PENDING',
                'agent': None,
                'lab': None,
                'artifacts': None,
                'maturity_state': 'IMMATURE',
                'ml_score': 0.0,
                'ml_confidence': 'LOW',
                'ml_suggested_agent': None,
                'ml_suggested_lab': None,
                'ml_anomaly_prob': 0.0,
                'enforcement_applied': False,
                'compliance_status': None
            },
            {
                'row_id': 'demo-002',
                'timestamp': '2025-01-11T10:05:00',
                'cycle': 1,
                'layer': 'L1',
                'variant': 'A',
                'input_type': 'grain',
                'source_protocol': 'REST',
                'status': 'TEST',
                'security': 'PASS',
                'schema_validation': 'PASS',
                'agent': 'agent-data-processing',
                'lab': 'lab-data-validation',
                'artifacts': None,
                'maturity_state': 'MATURING',
                'ml_score': 0.0,
                'ml_confidence': 'LOW',
                'ml_suggested_agent': None,
                'ml_suggested_lab': None,
                'ml_anomaly_prob': 0.0,
                'enforcement_applied': False,
                'compliance_status': None
            },
            {
                'row_id': 'demo-003',
                'timestamp': '2025-01-11T10:10:00',
                'cycle': 1,
                'layer': 'L1',
                'variant': 'A',
                'input_type': 'grain',
                'source_protocol': 'REST',
                'status': 'READY',
                'security': 'PASS',
                'schema_validation': 'PASS',
                'agent': 'agent-data-processing',
                'lab': 'lab-data-validation',
                'artifacts': 'artifact-001,artifact-002',
                'maturity_state': 'MATURE',
                'ml_score': 0.85,
                'ml_confidence': 'HIGH',
                'ml_suggested_agent': 'agent-data-processing',
                'ml_suggested_lab': 'lab-transformation',
                'ml_anomaly_prob': 0.05,
                'enforcement_applied': True,
                'compliance_status': 'COMPLIANT'
            },
            {
                'row_id': 'demo-004',
                'timestamp': '2025-01-11T10:15:00',
                'cycle': 2,
                'layer': 'L2',
                'variant': 'B',
                'input_type': 'order',
                'source_protocol': 'GraphQL',
                'status': 'FAIL',
                'security': 'FAIL',
                'schema_validation': 'PASS',
                'agent': None,
                'lab': None,
                'artifacts': None,
                'maturity_state': 'IMMATURE',
                'ml_score': 0.0,
                'ml_confidence': 'LOW',
                'ml_suggested_agent': None,
                'ml_suggested_lab': None,
                'ml_anomaly_prob': 0.0,
                'enforcement_applied': False,
                'compliance_status': None
            }
        ]
    
    # Add data rows
    for row_num, row_data in enumerate(sample_rows, 2):
        for col_num, header in enumerate(headers, 1):
            # Get value using lowercase header as key
            key = header.lower()
            value = row_data.get(key, '')
            
            # Handle None values
            if value is None:
                value = ''
            
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Apply color coding
            if header == 'Status' and value in status_colors:
                cell.fill = PatternFill(
                    start_color=status_colors[value],
                    end_color=status_colors[value],
                    fill_type="solid"
                )
            elif header == 'Maturity_State' and value in maturity_colors:
                cell.fill = PatternFill(
                    start_color=maturity_colors[value],
                    end_color=maturity_colors[value],
                    fill_type="solid"
                )
            elif header == 'ML_Confidence' and value in ml_confidence_colors:
                cell.fill = PatternFill(
                    start_color=ml_confidence_colors[value],
                    end_color=ml_confidence_colors[value],
                    fill_type="solid"
                )
            elif header in ['Security', 'Schema_Validation'] and value in security_colors:
                cell.fill = PatternFill(
                    start_color=security_colors[value],
                    end_color=security_colors[value],
                    fill_type="solid"
                )
    
    # Adjust column widths
    column_widths = {
        'Row_ID': 15,
        'Timestamp': 20,
        'Cycle': 8,
        'Layer': 8,
        'Variant': 10,
        'Input_Type': 15,
        'Source_Protocol': 15,
        'Status': 12,
        'Security': 12,
        'Schema_Validation': 18,
        'Agent': 25,
        'Lab': 20,
        'Artifacts': 25,
        'Maturity_State': 15,
        'ML_Score': 10,
        'ML_Confidence': 15,
        'ML_Suggested_Agent': 25,
        'ML_Suggested_Lab': 20,
        'ML_Anomaly_Prob': 15,
        'Enforcement_Applied': 18,
        'Compliance_Status': 18
    }
    
    for col_num, header in enumerate(headers, 1):
        col_letter = ws.cell(row=1, column=col_num).column_letter
        ws.column_dimensions[col_letter].width = column_widths.get(header, 15)
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    # Save workbook
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    output = create_canonical_table_template("data/canonical_table_template.xlsx")
    print(f"✅ Created Excel template: {output}")
