"""
Postman Collection Generator from Excel Canonical Table
Generates Postman Collection v2.1.0 from API specifications in Excel.
"""

import json
import os
from datetime import datetime
from typing import Dict, List, Any, Optional

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

from openpyxl import load_workbook


def parse_json_field(value: str) -> Dict:
    """Safely parse JSON string field from Excel."""
    if not value or pd.isna(value) if PANDAS_AVAILABLE else not value:
        return {}
    try:
        if isinstance(value, str):
            return json.loads(value)
        return {}
    except (json.JSONDecodeError, TypeError):
        return {}


def generate_postman_collection(
    excel_file: str,
    output_file: str = None,
    collection_name: str = "Ultra Canonical API Collection",
    base_url: str = "https://api.yourdomain.com",
    sheet_name: str = "Canonical Table"
) -> str:
    """
    Generate Postman Collection from Excel API specifications.
    
    Args:
        excel_file: Path to Excel file with API specs
        output_file: Output JSON file path (default: auto-generated)
        collection_name: Name for the Postman collection
        base_url: Base URL for API endpoints
        sheet_name: Excel sheet name to read
        
    Returns:
        Path to generated Postman collection JSON file
    """
    
    if not os.path.exists(excel_file):
        raise FileNotFoundError(f"Excel file not found: {excel_file}")
    
    # Read Excel with pandas if available, otherwise openpyxl
    if PANDAS_AVAILABLE:
        df = pd.read_excel(excel_file, sheet_name=sheet_name)
        rows = df.to_dict('records')
    else:
        wb = load_workbook(excel_file)
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            rows.append(row_dict)
    
    # Generate unique collection ID
    collection_id = f"ultra-collection-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    
    # Postman Collection skeleton v2.1.0
    postman_collection = {
        "info": {
            "name": collection_name,
            "_postman_id": collection_id,
            "description": f"Generated automatically from Excel API spec\nSource: {excel_file}\nGenerated: {datetime.now().isoformat()}",
            "schema": "https://schema.getpostman.com/json/collection/v2.1.0/collection.json"
        },
        "item": [],
        "variable": [
            {"key": "base_url", "value": base_url, "type": "string"},
            {"key": "api_key", "value": "", "type": "string"},
            {"key": "auth_token", "value": "", "type": "string"}
        ],
        "auth": {
            "type": "bearer",
            "bearer": [{"key": "token", "value": "{{auth_token}}", "type": "string"}]
        }
    }
    
    # Group items by folder (based on first path segment)
    folders: Dict[str, List[Dict]] = {}
    
    for row in rows:
        # Extract fields with defaults
        endpoint = row.get("Endpoint", "/") or "/"
        method = str(row.get("Method", "GET") or "GET").upper()
        headers = row.get("Headers", "{}")
        params = row.get("Query_Params", "{}")
        body = row.get("Body", "")
        sample_response = row.get("Sample_Response", "")
        description = row.get("Description", "")
        auth_type = row.get("Auth_Type", "bearer")
        
        # Additional fields from canonical table
        row_id = row.get("Row_ID", "")
        protocol = row.get("Protocol", "REST")
        status = row.get("Status", "")
        
        # Parse JSON fields
        headers_dict = parse_json_field(headers)
        params_dict = parse_json_field(params)
        
        # Add default headers if not present
        if "Content-Type" not in headers_dict:
            headers_dict["Content-Type"] = "application/json"
        if "Accept" not in headers_dict:
            headers_dict["Accept"] = "application/json"
        
        # Build path segments
        path_segments = [p for p in endpoint.strip("/").split("/") if p]
        
        # Determine folder name (first path segment or "root")
        folder_name = path_segments[0] if path_segments else "root"
        
        # Create request object
        request_obj = {
            "method": method,
            "header": [
                {"key": k, "value": str(v), "type": "text"} 
                for k, v in headers_dict.items()
            ],
            "url": {
                "raw": "{{base_url}}" + endpoint,
                "host": ["{{base_url}}"],
                "path": path_segments,
                "query": [
                    {"key": k, "value": str(v), "disabled": False} 
                    for k, v in params_dict.items()
                ]
            },
            "description": f"{description}\n\nRow ID: {row_id}\nProtocol: {protocol}\nStatus: {status}"
        }
        
        # Add body if present and method supports it
        if body and method in ["POST", "PUT", "PATCH"]:
            request_obj["body"] = {
                "mode": "raw",
                "raw": body if isinstance(body, str) else json.dumps(body, indent=2),
                "options": {
                    "raw": {"language": "json"}
                }
            }
        
        # Create item for Postman
        item = {
            "name": f"{method} {endpoint}",
            "request": request_obj,
            "response": []
        }
        
        # Add sample response if present
        if sample_response:
            item["response"].append({
                "name": "Sample Response",
                "originalRequest": request_obj,
                "status": "OK",
                "code": 200,
                "_postman_previewlanguage": "json",
                "header": [
                    {"key": "Content-Type", "value": "application/json"}
                ],
                "body": sample_response if isinstance(sample_response, str) else json.dumps(sample_response, indent=2)
            })
        
        # Add to folder
        if folder_name not in folders:
            folders[folder_name] = []
        folders[folder_name].append(item)
    
    # Build folder structure
    for folder_name, items in folders.items():
        folder = {
            "name": folder_name.capitalize(),
            "description": f"API endpoints for /{folder_name}",
            "item": items
        }
        postman_collection["item"].append(folder)
    
    # Generate output filename if not provided
    if not output_file:
        base_name = os.path.splitext(os.path.basename(excel_file))[0]
        output_file = f"{base_name}_postman_collection.json"
    
    # Save JSON
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(postman_collection, f, indent=2, ensure_ascii=False)
    
    print(f"✅ Postman collection created: {output_file}")
    print(f"   Total endpoints: {sum(len(items) for items in folders.values())}")
    print(f"   Folders: {list(folders.keys())}")
    
    return output_file


def add_api_columns_to_excel(excel_file: str, output_file: str = None) -> str:
    """
    Add API-specific columns to an existing Excel canonical table.
    
    Adds columns: Method, Headers, Query_Params, Body, Sample_Response, Description, Auth_Type
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = load_workbook(excel_file)
    ws = wb['Canonical Table']
    
    # Get existing headers
    existing_headers = [cell.value for cell in ws[1]]
    
    # New API columns to add
    api_columns = [
        'Method', 'Headers', 'Query_Params', 'Body', 
        'Sample_Response', 'Description', 'Auth_Type'
    ]
    
    # Header style
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    # Add new columns
    next_col = len(existing_headers) + 1
    added = []
    
    for col_name in api_columns:
        if col_name not in existing_headers:
            cell = ws.cell(row=1, column=next_col, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[cell.column_letter].width = 20
            added.append(col_name)
            next_col += 1
    
    # Add default values for existing rows
    if added:
        for row in range(2, ws.max_row + 1):
            for col_idx, col_name in enumerate(added, start=len(existing_headers) + 1):
                if col_name == 'Method':
                    ws.cell(row=row, column=col_idx, value='GET')
                elif col_name == 'Headers':
                    ws.cell(row=row, column=col_idx, value='{}')
                elif col_name == 'Query_Params':
                    ws.cell(row=row, column=col_idx, value='{}')
                elif col_name == 'Auth_Type':
                    ws.cell(row=row, column=col_idx, value='bearer')
    
    # Save
    output_file = output_file or excel_file
    wb.save(output_file)
    
    if added:
        print(f"✅ Added API columns to Excel: {added}")
    else:
        print("ℹ️ All API columns already exist")
    
    return output_file


if __name__ == '__main__':
    import argparse
    
    parser = argparse.ArgumentParser(description='Generate Postman Collection from Excel')
    parser.add_argument('--excel', type=str, default='canonical_table_dashboard.xlsx',
                        help='Path to Excel file with API specs')
    parser.add_argument('--output', type=str, default=None,
                        help='Output Postman collection JSON file')
    parser.add_argument('--name', type=str, default='Ultra Canonical API Collection',
                        help='Collection name')
    parser.add_argument('--base-url', type=str, default='https://api.yourdomain.com',
                        help='Base URL for API endpoints')
    parser.add_argument('--add-columns', action='store_true',
                        help='Add API columns to Excel first')
    
    args = parser.parse_args()
    
    # Add API columns if requested
    if args.add_columns:
        add_api_columns_to_excel(args.excel)
    
    # Generate Postman collection
    generate_postman_collection(
        excel_file=args.excel,
        output_file=args.output,
        collection_name=args.name,
        base_url=args.base_url
    )
