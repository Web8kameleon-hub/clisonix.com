"""
Postman Collection Merger & Excel Exporter
Merges multiple Postman collections into one unified collection
and exports to Excel canonical table format.
"""

import json
import os
from datetime import datetime
from typing import Dict, List, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


def load_postman_collection(file_path: str) -> Dict:
    """Load a Postman collection from JSON file."""
    with open(file_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def extract_requests_from_items(items: List[Dict], parent_folder: str = "") -> List[Dict]:
    """
    Recursively extract all requests from Postman collection items.
    
    Args:
        items: List of Postman items (can be folders or requests)
        parent_folder: Parent folder name for nested items
        
    Returns:
        List of flattened request dictionaries
    """
    requests = []
    
    for item in items:
        # Get folder name
        folder_name = item.get('name', 'Unknown')
        full_path = f"{parent_folder}/{folder_name}" if parent_folder else folder_name
        
        # Check if this is a folder (has nested items)
        if 'item' in item:
            # Recurse into folder
            nested_requests = extract_requests_from_items(item['item'], full_path)
            requests.extend(nested_requests)
        elif 'request' in item:
            # This is a request
            req = item['request']
            
            # Extract URL
            url = req.get('url', {})
            if isinstance(url, str):
                raw_url = url
                endpoint = url.split('}}')[-1] if '}}' in url else url
            else:
                raw_url = url.get('raw', '')
                path = url.get('path', [])
                endpoint = '/' + '/'.join(path) if path else '/'
            
            # Extract method
            method = req.get('method', 'GET')
            
            # Extract headers
            headers = {}
            for h in req.get('header', []):
                if isinstance(h, dict):
                    headers[h.get('key', '')] = h.get('value', '')
            
            # Extract query params
            query_params = {}
            if isinstance(url, dict):
                for q in url.get('query', []):
                    if isinstance(q, dict):
                        query_params[q.get('key', '')] = q.get('value', '')
            
            # Extract body
            body = req.get('body', {})
            if isinstance(body, dict):
                body_raw = body.get('raw', '')
            else:
                body_raw = str(body)
            
            # Extract description
            description = req.get('description', '') or item.get('description', '')
            
            # Extract sample response
            responses = item.get('response', [])
            sample_response = ''
            if responses and len(responses) > 0:
                first_resp = responses[0]
                if isinstance(first_resp, dict):
                    sample_response = first_resp.get('body', '')
            
            requests.append({
                'name': item.get('name', ''),
                'folder': full_path,
                'method': method,
                'endpoint': endpoint,
                'raw_url': raw_url,
                'headers': json.dumps(headers) if headers else '{}',
                'query_params': json.dumps(query_params) if query_params else '{}',
                'body': body_raw,
                'description': description,
                'sample_response': sample_response
            })
    
    return requests


def merge_postman_collections(
    collection_files: List[str],
    output_file: str = 'merged_postman_collection.json',
    collection_name: str = 'Clisonix Ultra Mega Collection'
) -> Dict:
    """
    Merge multiple Postman collections into one.
    
    Args:
        collection_files: List of paths to Postman collection JSON files
        output_file: Output path for merged collection
        collection_name: Name for the merged collection
        
    Returns:
        Merged collection dictionary
    """
    merged = {
        "info": {
            "name": collection_name,
            "_postman_id": f"merged-{datetime.now().strftime('%Y%m%d%H%M%S')}",
            "description": f"Merged collection from {len(collection_files)} sources\nGenerated: {datetime.now().isoformat()}",
            "schema": "https://schema.getpostman.com/json/collection/v2.1.0/collection.json"
        },
        "item": [],
        "variable": [],
        "auth": {
            "type": "bearer",
            "bearer": [{"key": "token", "value": "{{auth_token}}", "type": "string"}]
        }
    }
    
    all_variables = {}
    
    for file_path in collection_files:
        if not os.path.exists(file_path):
            print(f"⚠️ File not found: {file_path}")
            continue
            
        try:
            collection = load_postman_collection(file_path)
            source_name = collection.get('info', {}).get('name', os.path.basename(file_path))
            
            # Create folder for this collection's items
            folder = {
                "name": f"📁 {source_name}",
                "description": f"Imported from: {os.path.basename(file_path)}",
                "item": collection.get('item', [])
            }
            merged['item'].append(folder)
            
            # Collect variables
            for var in collection.get('variable', []):
                if isinstance(var, dict):
                    key = var.get('key', '')
                    if key and key not in all_variables:
                        all_variables[key] = var
            
            print(f"✅ Merged: {source_name} ({len(collection.get('item', []))} folders)")
            
        except Exception as e:
            print(f"❌ Error loading {file_path}: {e}")
    
    # Add all unique variables
    merged['variable'] = list(all_variables.values())
    
    # Save merged collection
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(merged, f, indent=2, ensure_ascii=False)
    
    print(f"\n✅ Merged collection saved: {output_file}")
    print(f"   Total source folders: {len(merged['item'])}")
    print(f"   Total variables: {len(merged['variable'])}")
    
    return merged


def export_postman_to_excel(
    collection_files: List[str],
    output_file: str = 'postman_canonical_table.xlsx'
) -> str:
    """
    Export Postman collections to Excel canonical table format.
    
    Args:
        collection_files: List of paths to Postman collection JSON files
        output_file: Output Excel file path
        
    Returns:
        Path to generated Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Canonical Table'
    
    # Headers matching canonical table + API specific
    headers = [
        'Row_ID', 'Timestamp', 'Source_Collection', 'Folder',
        'Protocol', 'Method', 'Endpoint', 'Raw_URL',
        'Headers', 'Query_Params', 'Body', 'Description',
        'Sample_Response', 'Status', 'Maturity_State'
    ]
    
    # Header styles
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 20
    
    # Collect all requests from all collections
    row_num = 2
    row_id = 1
    
    for file_path in collection_files:
        if not os.path.exists(file_path):
            print(f"⚠️ File not found: {file_path}")
            continue
            
        try:
            collection = load_postman_collection(file_path)
            source_name = collection.get('info', {}).get('name', os.path.basename(file_path))
            
            # Extract all requests
            requests = extract_requests_from_items(collection.get('item', []))
            
            for req in requests:
                row_data = [
                    f'ROW-{row_id:04d}',                         # Row_ID
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'), # Timestamp
                    source_name,                                  # Source_Collection
                    req['folder'],                                # Folder
                    'REST',                                       # Protocol
                    req['method'],                                # Method
                    req['endpoint'],                              # Endpoint
                    req['raw_url'],                               # Raw_URL
                    req['headers'],                               # Headers
                    req['query_params'],                          # Query_Params
                    req['body'][:500] if req['body'] else '',    # Body (truncated)
                    req['description'][:300] if req['description'] else '',  # Description
                    req['sample_response'][:500] if req['sample_response'] else '',  # Sample_Response
                    'READY',                                      # Status
                    'MATURE'                                      # Maturity_State
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col, value=value)
                    cell.border = thin_border
                    
                    # Color code status and maturity
                    if col == 14:  # Status
                        cell.fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
                    elif col == 15:  # Maturity
                        cell.fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
                
                row_num += 1
                row_id += 1
            
            print(f"✅ Exported: {source_name} ({len(requests)} requests)")
            
        except Exception as e:
            print(f"❌ Error processing {file_path}: {e}")
    
    # Adjust column widths
    ws.column_dimensions['G'].width = 40  # Endpoint
    ws.column_dimensions['H'].width = 50  # Raw_URL
    ws.column_dimensions['L'].width = 50  # Description
    
    # Save
    wb.save(output_file)
    
    print(f"\n✅ Excel canonical table saved: {output_file}")
    print(f"   Total API endpoints: {row_id - 1}")
    
    return output_file


if __name__ == '__main__':
    import argparse
    
    parser = argparse.ArgumentParser(description='Merge Postman Collections & Export to Excel')
    parser.add_argument('--collections', nargs='+', required=True,
                        help='Paths to Postman collection JSON files')
    parser.add_argument('--merge-output', type=str, default='merged_postman_collection.json',
                        help='Output path for merged collection')
    parser.add_argument('--excel-output', type=str, default='postman_canonical_table.xlsx',
                        help='Output path for Excel canonical table')
    parser.add_argument('--name', type=str, default='Clisonix Ultra Mega Collection',
                        help='Name for merged collection')
    parser.add_argument('--merge-only', action='store_true',
                        help='Only merge, skip Excel export')
    parser.add_argument('--excel-only', action='store_true',
                        help='Only export to Excel, skip merge')
    
    args = parser.parse_args()
    
    if not args.excel_only:
        merge_postman_collections(
            args.collections,
            args.merge_output,
            args.name
        )
    
    if not args.merge_only:
        export_postman_to_excel(
            args.collections,
            args.excel_output
        )
