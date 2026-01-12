"""
API Documentation Enricher
Adds complete Sample_Response, Description, cURL commands, and Python snippets
for all Clisonix Cloud API endpoints.
"""

import json
from datetime import datetime

# Complete API documentation with sample responses
API_DOCUMENTATION = {
    # ========== HEALTH & STATUS ==========
    "/health": {
        "method": "GET",
        "description": "Kontrollon gjendjen e shëndetit të sistemit. Kthen statusin e përgjithshëm të platformës Clisonix Cloud duke përfshirë versionin, uptime dhe statusin e shërbimeve të brendshme.",
        "sample_response": {
            "service": "clisonix-cloud",
            "status": "healthy",
            "version": "2.1.0",
            "uptime_seconds": 86400,
            "timestamp": "2026-01-11T10:00:00Z",
            "components": {
                "database": "healthy",
                "redis": "healthy",
                "prometheus": "healthy"
            }
        },
        "curl": 'curl -X GET "https://api.clisonix.com/health" -H "Authorization: Bearer $TOKEN"',
        "python": '''import requests

response = requests.get(
    "https://api.clisonix.com/health",
    headers={"Authorization": f"Bearer {TOKEN}"}
)
print(response.json())'''
    },
    
    "/status": {
        "method": "GET",
        "description": "Merr statusin e detajuar të sistemit duke përfshirë metrikat e CPU, RAM, disk dhe numrin e kërkesave aktive.",
        "sample_response": {
            "status": "operational",
            "cpu_percent": 23.5,
            "memory_percent": 45.2,
            "disk_percent": 67.8,
            "active_requests": 42,
            "total_requests_24h": 15420,
            "average_response_time_ms": 125,
            "timestamp": "2026-01-11T10:00:00Z"
        },
        "curl": 'curl -X GET "https://api.clisonix.com/status" -H "Authorization: Bearer $TOKEN"',
        "python": '''import requests

response = requests.get(
    "https://api.clisonix.com/status",
    headers={"Authorization": f"Bearer {TOKEN}"}
)
print(response.json())'''
    },
    
    "/api/system-status": {
        "method": "GET",
        "description": "Kthen statusin e plotë të sistemit industrial duke përfshirë të gjitha shërbimet, metrikat dhe alarmet aktive.",
        "sample_response": {
            "system": "clisonix-industrial",
            "status": "running",
            "services": {
                "alba_network": {"status": "active", "latency_ms": 12},
                "albi_neural": {"status": "active", "goroutines": 156},
                "jona_coordinator": {"status": "active", "score": 0.95}
            },
            "alerts": [],
            "last_check": "2026-01-11T10:00:00Z"
        },
        "curl": 'curl -X GET "https://api.clisonix.com/api/system-status" -H "Authorization: Bearer $TOKEN"',
        "python": '''import requests

response = requests.get(
    "https://api.clisonix.com/api/system-status",
    headers={"Authorization": f"Bearer {TOKEN}"}
)
print(response.json())'''
    },
    
    "/db/ping": {
        "method": "GET",
        "description": "Teston lidhjen me bazën e të dhënave PostgreSQL. Kthen kohën e përgjigjes dhe statusin e koneksionit.",
        "sample_response": {
            "database": "postgresql",
            "status": "connected",
            "ping_ms": 2.3,
            "version": "PostgreSQL 15.2",
            "connections": {"active": 12, "max": 100},
            "timestamp": "2026-01-11T10:00:00Z"
        },
        "curl": 'curl -X GET "https://api.clisonix.com/db/ping" -H "Authorization: Bearer $TOKEN"',
        "python": '''import requests

response = requests.get(
    "https://api.clisonix.com/db/ping",
    headers={"Authorization": f"Bearer {TOKEN}"}
)
print(response.json())'''
    },
    
    "/redis/ping": {
        "method": "GET",
        "description": "Teston lidhjen me Redis cache. Kthen statusin, memorien e përdorur dhe numrin e çelësave.",
        "sample_response": {
            "redis": "connected",
            "ping": "PONG",
            "latency_ms": 0.8,
            "memory_used_mb": 256,
            "keys_count": 15420,
            "version": "7.2.0",
            "timestamp": "2026-01-11T10:00:00Z"
        },
        "curl": 'curl -X GET "https://api.clisonix.com/redis/ping" -H "Authorization: Bearer $TOKEN"',
        "python": '''import requests

response = requests.get(
    "https://api.clisonix.com/redis/ping",
    headers={"Authorization": f"Bearer {TOKEN}"}
)
print(response.json())'''
    },
    
    # ========== ASK & NEURAL SYMPHONY ==========
    "/api/ask": {
        "method": "POST",
        "description": "Endpoint i inteligjencës artificiale për pyetje-përgjigje. Pranon pyetje në gjuhë natyrore dhe kthen përgjigje të gjeneruara nga AI me kontekst të plotë.",
        "sample_response": {
            "answer": "Clisonix është një platformë e avancuar neuro-audio që kombinon përpunimin EEG me gjenerimin e muzikës brain-sync.",
            "confidence": 0.95,
            "sources": ["internal_docs", "knowledge_base"],
            "processing_time_ms": 450,
            "tokens_used": 125,
            "timestamp": "2026-01-11T10:00:00Z"
        },
        "curl": '''curl -X POST "https://api.clisonix.com/api/ask" \\
  -H "Authorization: Bearer $TOKEN" \\
  -H "Content-Type: application/json" \\
  -d '{"question": "Çfarë është Clisonix?", "context": "Platformë neuro-audio", "include_details": true}' ''',
        "python": '''import requests

response = requests.post(
    "https://api.clisonix.com/api/ask",
    headers={
        "Authorization": f"Bearer {TOKEN}",
        "Content-Type": "application/json"
    },
    json={
        "question": "Çfarë është Clisonix?",
        "context": "Platformë neuro-audio dhe EEG",
        "include_details": True
    }
)
print(response.json())'''
    },
    
    "/neural-symphony": {
        "method": "GET",
        "description": "Gjeneron muzikë brain-sync në kohë reale bazuar në parametrat e specifikuar. Kthen stream audio WAV.",
        "sample_response": {
            "type": "audio/wav",
            "duration_seconds": 300,
            "frequency_hz": 40,
            "wave_type": "binaural",
            "status": "streaming",
            "stream_url": "/neural-symphony/stream/abc123"
        },
        "curl": 'curl -X GET "https://api.clisonix.com/neural-symphony" -H "Authorization: Bearer $TOKEN" -H "Accept: audio/wav" --output symphony.wav',
        "python": '''import requests

response = requests.get(
    "https://api.clisonix.com/neural-symphony",
    headers={
        "Authorization": f"Bearer {TOKEN}",
        "Accept": "audio/wav"
    },
    stream=True
)

with open("symphony.wav", "wb") as f:
    for chunk in response.iter_content(chunk_size=8192):
        f.write(chunk)'''
    },
    
    # ========== UPLOADS & PROCESSING ==========
    "/api/uploads/eeg/process": {
        "method": "POST",
        "description": "Ngarkon dhe përpunon skedarë EEG (formatet .edf, .bdf, .fif). Ekstrakon kanalet, frekuencat dhe gjeneron raport analitik.",
        "sample_response": {
            "file_id": "eeg-upload-001",
            "filename": "brain_scan_001.edf",
            "status": "processed",
            "channels": 64,
            "duration_seconds": 3600,
            "sampling_rate_hz": 256,
            "bands": {
                "delta": {"power": 0.25, "range": "0.5-4 Hz"},
                "theta": {"power": 0.30, "range": "4-8 Hz"},
                "alpha": {"power": 0.20, "range": "8-13 Hz"},
                "beta": {"power": 0.15, "range": "13-30 Hz"},
                "gamma": {"power": 0.10, "range": "30-100 Hz"}
            },
            "report_url": "/reports/eeg-upload-001.pdf",
            "timestamp": "2026-01-11T10:00:00Z"
        },
        "curl": '''curl -X POST "https://api.clisonix.com/api/uploads/eeg/process" \\
  -H "Authorization: Bearer $TOKEN" \\
  -F "file=@brain_scan.edf"''',
        "python": '''import requests

with open("brain_scan.edf", "rb") as f:
    response = requests.post(
        "https://api.clisonix.com/api/uploads/eeg/process",
        headers={"Authorization": f"Bearer {TOKEN}"},
        files={"file": f}
    )
print(response.json())'''
    },
    
    "/api/uploads/audio/process": {
        "method": "POST",
        "description": "Ngarkon dhe përpunon skedarë audio për analizë. Ekstrakon karakteristika audio, detekton emocione dhe gjeneron metadata.",
        "sample_response": {
            "file_id": "audio-upload-001",
            "filename": "recording.wav",
            "status": "processed",
            "duration_seconds": 180,
            "sample_rate_hz": 44100,
            "channels": 2,
            "analysis": {
                "tempo_bpm": 120,
                "key": "C major",
                "mood": "calm",
                "energy": 0.6,
                "valence": 0.7
            },
            "timestamp": "2026-01-11T10:00:00Z"
        },
        "curl": '''curl -X POST "https://api.clisonix.com/api/uploads/audio/process" \\
  -H "Authorization: Bearer $TOKEN" \\
  -F "file=@recording.wav"''',
        "python": '''import requests

with open("recording.wav", "rb") as f:
    response = requests.post(
        "https://api.clisonix.com/api/uploads/audio/process",
        headers={"Authorization": f"Bearer {TOKEN}"},
        files={"file": f}
    )
print(response.json())'''
    },
    
    # ========== BILLING ==========
    "/billing/paypal/order": {
        "method": "POST",
        "description": "Krijon një porosi PayPal për pagesë. Kthen ID-në e porosisë dhe URL-në për aprovim.",
        "sample_response": {
            "order_id": "PAYPAL-ORDER-5AB12345XY",
            "status": "CREATED",
            "intent": "CAPTURE",
            "amount": {"currency_code": "EUR", "value": "10.00"},
            "approval_url": "https://www.paypal.com/checkoutnow?token=5AB12345XY",
            "created_at": "2026-01-11T10:00:00Z"
        },
        "curl": '''curl -X POST "https://api.clisonix.com/billing/paypal/order" \\
  -H "Authorization: Bearer $TOKEN" \\
  -H "Content-Type: application/json" \\
  -d '{"intent": "CAPTURE", "purchase_units": [{"amount": {"currency_code": "EUR", "value": "10.00"}}]}'  ''',
        "python": '''import requests

response = requests.post(
    "https://api.clisonix.com/billing/paypal/order",
    headers={
        "Authorization": f"Bearer {TOKEN}",
        "Content-Type": "application/json"
    },
    json={
        "intent": "CAPTURE",
        "purchase_units": [{
            "amount": {"currency_code": "EUR", "value": "10.00"}
        }]
    }
)
print(response.json())'''
    },
    
    "/billing/stripe/payment-intent": {
        "method": "POST",
        "description": "Krijon një Stripe Payment Intent për pagesë me kartë ose SEPA. Kthen client_secret për frontend.",
        "sample_response": {
            "id": "pi_1234567890",
            "client_secret": "pi_1234567890_secret_abc123",
            "status": "requires_payment_method",
            "amount": 1000,
            "currency": "eur",
            "payment_method_types": ["card", "sepa_debit"],
            "created": 1704963600
        },
        "curl": '''curl -X POST "https://api.clisonix.com/billing/stripe/payment-intent" \\
  -H "Authorization: Bearer $TOKEN" \\
  -H "Content-Type: application/json" \\
  -d '{"amount": 1000, "currency": "eur", "payment_method_types": ["sepa_debit"]}'  ''',
        "python": '''import requests

response = requests.post(
    "https://api.clisonix.com/billing/stripe/payment-intent",
    headers={
        "Authorization": f"Bearer {TOKEN}",
        "Content-Type": "application/json"
    },
    json={
        "amount": 1000,
        "currency": "eur",
        "payment_method_types": ["sepa_debit"],
        "description": "Clisonix subscription"
    }
)
print(response.json())'''
    },
    
    # ========== ASI TRINITY ==========
    "/asi/status": {
        "method": "GET",
        "description": "Merr statusin e ASI Trinity (ALBA, ALBI, JONA). Kthen gjendjen e secilit agjent dhe koordinimin ndërmjet tyre.",
        "sample_response": {
            "trinity_status": "active",
            "agents": {
                "alba": {
                    "status": "online",
                    "role": "Network Manager",
                    "latency_ms": 12,
                    "requests_per_second": 150
                },
                "albi": {
                    "status": "online", 
                    "role": "Neural Processor",
                    "goroutines": 256,
                    "neural_patterns": 1024
                },
                "jona": {
                    "status": "online",
                    "role": "Coordinator",
                    "coordination_score": 0.97,
                    "http_requests_total": 50000
                }
            },
            "coordination_level": "synchronized",
            "timestamp": "2026-01-11T10:00:00Z"
        },
        "curl": 'curl -X GET "https://api.clisonix.com/asi/status" -H "Authorization: Bearer $TOKEN"',
        "python": '''import requests

response = requests.get(
    "https://api.clisonix.com/asi/status",
    headers={"Authorization": f"Bearer {TOKEN}"}
)
print(response.json())'''
    },
    
    "/asi/health": {
        "method": "GET",
        "description": "Kontrollon shëndetin e ASI Trinity. Kthen health check të detajuar për secilin komponent dhe alarme aktive.",
        "sample_response": {
            "overall_health": "healthy",
            "components": {
                "alba_network": {
                    "status": "healthy",
                    "cpu_usage": 23.5,
                    "memory_mb": 512,
                    "network_latency_ms": 5
                },
                "albi_neural": {
                    "status": "healthy",
                    "cpu_usage": 45.2,
                    "memory_mb": 1024,
                    "gc_pause_ms": 2.3
                },
                "jona_coordinator": {
                    "status": "healthy",
                    "cpu_usage": 15.8,
                    "memory_mb": 256,
                    "active_connections": 42
                }
            },
            "alerts": [],
            "last_check": "2026-01-11T10:00:00Z"
        },
        "curl": 'curl -X GET "https://api.clisonix.com/asi/health" -H "Authorization: Bearer $TOKEN"',
        "python": '''import requests

response = requests.get(
    "https://api.clisonix.com/asi/health",
    headers={"Authorization": f"Bearer {TOKEN}"}
)
print(response.json())'''
    },
    
    "/asi/execute": {
        "method": "POST",
        "description": "Ekzekuton një komandë në ASI Trinity. Mund të dërgojë komanda tek ALBA, ALBI ose JONA.",
        "sample_response": {
            "command": "status",
            "agent": "trinity",
            "result": {
                "executed": True,
                "output": "All agents operational",
                "execution_time_ms": 125
            },
            "timestamp": "2026-01-11T10:00:00Z"
        },
        "curl": '''curl -X POST "https://api.clisonix.com/asi/execute" \\
  -H "Authorization: Bearer $TOKEN" \\
  -H "Content-Type: application/json" \\
  -d '{"command": "status", "agent": "trinity", "parameters": {}}'  ''',
        "python": '''import requests

response = requests.post(
    "https://api.clisonix.com/asi/execute",
    headers={
        "Authorization": f"Bearer {TOKEN}",
        "Content-Type": "application/json"
    },
    json={
        "command": "status",
        "agent": "trinity",
        "parameters": {}
    }
)
print(response.json())'''
    },
    
    # ========== ALBA SPECIFIC ==========
    "/api/alba/health": {
        "method": "GET",
        "description": "Kontrollon shëndetin e ALBA Network Manager. Kthen metrikat e rrjetit, latency dhe statusin e lidhjeve.",
        "sample_response": {
            "service": "alba-network",
            "status": "healthy",
            "version": "3.2.1",
            "metrics": {
                "cpu_percent": 18.5,
                "memory_mb": 384,
                "network_in_mbps": 125.4,
                "network_out_mbps": 89.2,
                "active_connections": 156,
                "latency_p50_ms": 5,
                "latency_p95_ms": 12,
                "latency_p99_ms": 25
            },
            "uptime_seconds": 864000,
            "last_restart": "2026-01-01T00:00:00Z",
            "timestamp": "2026-01-11T10:00:00Z"
        },
        "curl": 'curl -X GET "https://api.clisonix.com/api/alba/health" -H "Authorization: Bearer $TOKEN"',
        "python": '''import requests

response = requests.get(
    "https://api.clisonix.com/api/alba/health",
    headers={"Authorization": f"Bearer {TOKEN}"}
)
print(response.json())'''
    },
    
    "/asi/alba/metrics": {
        "method": "GET",
        "description": "Merr metrikat real-time të ALBA nga Prometheus. Përfshin CPU, Memory, Network Latency.",
        "sample_response": {
            "service": "alba",
            "source": "prometheus",
            "metrics": {
                "cpu_usage_percent": 18.5,
                "memory_used_bytes": 402653184,
                "memory_percent": 12.3,
                "network_receive_bytes_total": 1073741824,
                "network_transmit_bytes_total": 536870912,
                "http_requests_total": 125000,
                "http_request_duration_seconds_p50": 0.005,
                "http_request_duration_seconds_p99": 0.025
            },
            "scraped_at": "2026-01-11T10:00:00Z"
        },
        "curl": 'curl -X GET "https://api.clisonix.com/asi/alba/metrics" -H "Authorization: Bearer $TOKEN"',
        "python": '''import requests

response = requests.get(
    "https://api.clisonix.com/asi/alba/metrics",
    headers={"Authorization": f"Bearer {TOKEN}"}
)
print(response.json())'''
    },
    
    # ========== ALBI SPECIFIC ==========
    "/asi/albi/metrics": {
        "method": "GET",
        "description": "Merr metrikat real-time të ALBI Neural nga Prometheus. Përfshin Goroutines, Neural Patterns, GC Operations.",
        "sample_response": {
            "service": "albi",
            "source": "prometheus",
            "metrics": {
                "goroutines": 256,
                "gc_pause_total_seconds": 0.125,
                "go_memstats_alloc_bytes": 536870912,
                "neural_patterns_processed": 10240,
                "neural_inference_duration_ms": 45,
                "model_accuracy": 0.97
            },
            "scraped_at": "2026-01-11T10:00:00Z"
        },
        "curl": 'curl -X GET "https://api.clisonix.com/asi/albi/metrics" -H "Authorization: Bearer $TOKEN"',
        "python": '''import requests

response = requests.get(
    "https://api.clisonix.com/asi/albi/metrics",
    headers={"Authorization": f"Bearer {TOKEN}"}
)
print(response.json())'''
    },
    
    # ========== JONA SPECIFIC ==========
    "/asi/jona/metrics": {
        "method": "GET",
        "description": "Merr metrikat real-time të JONA Coordinator nga Prometheus. Përfshin HTTP Requests, Coordination Score.",
        "sample_response": {
            "service": "jona",
            "source": "prometheus",
            "metrics": {
                "http_requests_total": 50000,
                "http_request_duration_seconds_avg": 0.125,
                "coordination_score": 0.97,
                "active_workflows": 12,
                "completed_workflows_24h": 1250,
                "error_rate_percent": 0.02
            },
            "scraped_at": "2026-01-11T10:00:00Z"
        },
        "curl": 'curl -X GET "https://api.clisonix.com/asi/jona/metrics" -H "Authorization: Bearer $TOKEN"',
        "python": '''import requests

response = requests.get(
    "https://api.clisonix.com/asi/jona/metrics",
    headers={"Authorization": f"Bearer {TOKEN}"}
)
print(response.json())'''
    },
    
    # ========== BRAIN ENGINE ==========
    "/brain/youtube/insight": {
        "method": "GET",
        "description": "Analizon një video YouTube dhe kthen insights, transkript dhe analiza sentimentesh.",
        "sample_response": {
            "video_id": "dQw4w9WgXcQ",
            "title": "Video Title",
            "duration_seconds": 212,
            "insights": {
                "summary": "Përmbledhje e videos...",
                "key_topics": ["topic1", "topic2"],
                "sentiment": "positive",
                "engagement_score": 0.85
            },
            "transcript_available": True,
            "timestamp": "2026-01-11T10:00:00Z"
        },
        "curl": 'curl -X GET "https://api.clisonix.com/brain/youtube/insight?video_id=dQw4w9WgXcQ" -H "Authorization: Bearer $TOKEN"',
        "python": '''import requests

response = requests.get(
    "https://api.clisonix.com/brain/youtube/insight",
    headers={"Authorization": f"Bearer {TOKEN}"},
    params={"video_id": "dQw4w9WgXcQ"}
)
print(response.json())'''
    },
    
    "/brain/music/brainsync": {
        "method": "POST",
        "description": "Gjeneron muzikë brain-sync bazuar në skedarin EEG dhe modalitetin e kërkuar (relax, focus, sleep).",
        "sample_response": {
            "job_id": "brainsync-job-001",
            "mode": "relax",
            "status": "processing",
            "estimated_duration_seconds": 300,
            "frequencies": {
                "carrier_hz": 200,
                "binaural_beat_hz": 10,
                "target_brainwave": "alpha"
            },
            "download_url": "/brain/music/download/brainsync-job-001",
            "timestamp": "2026-01-11T10:00:00Z"
        },
        "curl": '''curl -X POST "https://api.clisonix.com/brain/music/brainsync?mode=relax" \\
  -H "Authorization: Bearer $TOKEN" \\
  -F "file=@eeg_data.edf"''',
        "python": '''import requests

with open("eeg_data.edf", "rb") as f:
    response = requests.post(
        "https://api.clisonix.com/brain/music/brainsync",
        headers={"Authorization": f"Bearer {TOKEN}"},
        params={"mode": "relax"},
        files={"file": f}
    )
print(response.json())'''
    }
}


def enrich_excel_with_documentation(excel_file: str, output_file: str = None):
    """
    Enriches the Excel canonical table with complete documentation.
    Fixed wrap text and column widths for proper display.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    
    wb = load_workbook(excel_file)
    ws = wb['Canonical Table']
    
    # Get headers
    headers = [cell.value for cell in ws[1]]
    
    # Find column indices
    endpoint_col = headers.index('Endpoint') + 1 if 'Endpoint' in headers else None
    desc_col = headers.index('Description') + 1 if 'Description' in headers else None
    sample_col = headers.index('Sample_Response') + 1 if 'Sample_Response' in headers else None
    
    if not endpoint_col:
        print("❌ Endpoint column not found")
        return
    
    # Add new columns if needed
    curl_col = None
    python_col = None
    
    if 'cURL' not in headers:
        curl_col = len(headers) + 1
        ws.cell(row=1, column=curl_col, value='cURL')
        headers.append('cURL')
    else:
        curl_col = headers.index('cURL') + 1
        
    if 'Python_Snippet' not in headers:
        python_col = len(headers) + 1
        ws.cell(row=1, column=python_col, value='Python_Snippet')
        headers.append('Python_Snippet')
    else:
        python_col = headers.index('Python_Snippet') + 1
    
    # Style definitions
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12, name='Segoe UI')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    code_font = Font(size=10, name='Consolas')
    thin_border = Border(
        left=Side(style='thin', color='B4B4B4'),
        right=Side(style='thin', color='B4B4B4'),
        top=Side(style='thin', color='B4B4B4'),
        bottom=Side(style='thin', color='B4B4B4')
    )
    
    # ULTRA-WIDE COLUMN WIDTHS (Excel units: 1 unit ≈ 7 pixels)
    # 500px ≈ 71 units, 600px ≈ 85 units
    column_widths = {
        'A': 10,   # ID
        'B': 45,   # Endpoint
        'C': 12,   # Method
        'D': 71,   # Description (~500px)
        'E': 85,   # Sample_Response (~600px)
        'F': 85,   # cURL (~600px)
        'G': 85,   # Python_Snippet (~600px)
        'H': 40,   # Extra columns
        'I': 40,
        'J': 40,
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Style header row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Update rows with documentation
    enriched_count = 0
    
    for row in range(2, ws.max_row + 1):
        endpoint = ws.cell(row=row, column=endpoint_col).value
        
        if endpoint and endpoint in API_DOCUMENTATION:
            doc = API_DOCUMENTATION[endpoint]
            
            # Update Description with wrap text
            if desc_col:
                cell = ws.cell(row=row, column=desc_col, value=doc['description'])
                cell.alignment = cell_alignment
                cell.border = thin_border
                cell.font = Font(size=11, name='Segoe UI')
            
            # Update Sample_Response with wrap text (CODE FONT)
            if sample_col:
                sample_json = json.dumps(doc['sample_response'], indent=2, ensure_ascii=False)
                cell = ws.cell(row=row, column=sample_col, value=sample_json)
                cell.alignment = cell_alignment
                cell.border = thin_border
                cell.font = code_font
            
            # Add cURL with wrap text (CODE FONT)
            if curl_col:
                cell = ws.cell(row=row, column=curl_col, value=doc['curl'])
                cell.alignment = cell_alignment
                cell.border = thin_border
                cell.font = code_font
            
            # Add Python with wrap text (CODE FONT)
            if python_col:
                cell = ws.cell(row=row, column=python_col, value=doc['python'])
                cell.alignment = cell_alignment
                cell.border = thin_border
                cell.font = code_font
            
            enriched_count += 1
        
        # Apply formatting to all cells in this row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            if cell.alignment.wrap_text != True:
                cell.alignment = cell_alignment
            if not cell.border.left.style:
                cell.border = thin_border
    
    # Set row height - AUTO (let wrap text handle it)
    # Excel will auto-adjust based on content with wrap_text=True
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = None  # Auto height
    
    # Header row height
    ws.row_dimensions[1].height = 30
    
    # Freeze header
    ws.freeze_panes = 'A2'
    
    # Save
    output_file = output_file or excel_file
    wb.save(output_file)
    
    print(f"✅ Enriched {enriched_count} API endpoints with complete documentation")
    print(f"   Saved to: {output_file}")
    
    return output_file


def generate_api_docs_json(output_file: str = 'api_documentation.json'):
    """Export all API documentation to JSON file."""
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(API_DOCUMENTATION, f, indent=2, ensure_ascii=False)
    print(f"✅ API documentation exported to: {output_file}")
    return output_file


def generate_curl_commands(output_file: str = 'curl_commands.sh'):
    """Generate a shell script with all cURL commands."""
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write("#!/bin/bash\n")
        f.write("# Clisonix Cloud API - cURL Commands\n")
        f.write("# Generated: " + datetime.now().isoformat() + "\n\n")
        f.write("# Set your token first:\n")
        f.write("# export TOKEN='your-auth-token'\n\n")
        
        for endpoint, doc in API_DOCUMENTATION.items():
            f.write(f"# {endpoint} - {doc['method']}\n")
            f.write(f"# {doc['description'][:80]}...\n")
            f.write(doc['curl'] + "\n\n")
    
    print(f"✅ cURL commands exported to: {output_file}")
    return output_file


def generate_python_client(output_file: str = 'clisonix_api_client.py'):
    """Generate a Python API client with all endpoints."""
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write('"""\nClisonix Cloud API Client\n')
        f.write(f'Generated: {datetime.now().isoformat()}\n"""\n\n')
        f.write('import requests\nfrom typing import Dict, Any, Optional\n\n')
        f.write('class ClisonixClient:\n')
        f.write('    def __init__(self, base_url: str = "https://api.clisonix.com", token: str = ""):\n')
        f.write('        self.base_url = base_url\n')
        f.write('        self.token = token\n')
        f.write('        self.headers = {"Authorization": f"Bearer {token}"}\n\n')
        
        for endpoint, doc in API_DOCUMENTATION.items():
            method_name = endpoint.replace('/', '_').replace('-', '_').strip('_')
            
            f.write(f'    def {method_name}(self')
            if doc['method'] == 'POST':
                f.write(', data: Dict = None, files: Dict = None')
            f.write(') -> Dict:\n')
            f.write(f'        """{doc["description"][:100]}..."""\n')
            
            if doc['method'] == 'GET':
                f.write(f'        response = requests.get(f"{{self.base_url}}{endpoint}", headers=self.headers)\n')
            else:
                f.write(f'        response = requests.post(f"{{self.base_url}}{endpoint}", headers=self.headers, json=data, files=files)\n')
            
            f.write('        return response.json()\n\n')
    
    print(f"✅ Python client exported to: {output_file}")
    return output_file


if __name__ == '__main__':
    import argparse
    
    parser = argparse.ArgumentParser(description='API Documentation Enricher')
    parser.add_argument('--excel', type=str, help='Excel file to enrich')
    parser.add_argument('--output', type=str, help='Output file')
    parser.add_argument('--export-json', action='store_true', help='Export API docs to JSON')
    parser.add_argument('--export-curl', action='store_true', help='Export cURL commands')
    parser.add_argument('--export-python', action='store_true', help='Generate Python client')
    parser.add_argument('--all', action='store_true', help='Run all exports')
    
    args = parser.parse_args()
    
    if args.excel:
        enrich_excel_with_documentation(args.excel, args.output)
    
    if args.export_json or args.all:
        generate_api_docs_json()
    
    if args.export_curl or args.all:
        generate_curl_commands()
    
    if args.export_python or args.all:
        generate_python_client()
    
    if not any([args.excel, args.export_json, args.export_curl, args.export_python, args.all]):
        print("Usage examples:")
        print("  python api_enricher.py --excel clisonix-canonical-table.xlsx")
        print("  python api_enricher.py --all")
        print("  python api_enricher.py --export-json --export-curl --export-python")
