"""
CLISONIX EXCEL MICROSERVICE
Shërbim i izoluar për operacione Excel
Port: 8002
- Export/Import Excel
- Validation
- =PY() formula support
- Office Scripts integration
"""

from fastapi import FastAPI, Response, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional
import logging
import os
import json
import httpx

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s | %(levelname)s | %(name)s | %(message)s')
logger = logging.getLogger("excel-service")

# Core API URL
CORE_API_URL = os.environ.get("CORE_API_URL", "http://clisonix-core:8000")

app = FastAPI(
    title="Clisonix Excel Service",
    description="Isolated microservice for Excel operations, formulas, and Office Scripts",
    version="2.0.0",
    docs_url="/docs"
)

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Directories
EXCEL_DIR = Path("./excel_files")
EXCEL_DIR.mkdir(exist_ok=True)
UPLOADS_DIR = Path("./uploads")
UPLOADS_DIR.mkdir(exist_ok=True)

# Try to import openpyxl
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl not installed")

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    logger.warning("pandas not installed")


async def fetch_from_core(endpoint: str) -> Dict[str, Any]:
    """Fetch data from Core API"""
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            response = await client.get(f"{CORE_API_URL}{endpoint}")
            if response.status_code == 200:
                return response.json()
            else:
                logger.warning(f"Core API returned {response.status_code}")
                return {"error": f"Core API error: {response.status_code}"}
    except Exception as e:
        logger.error(f"Failed to fetch from Core API: {e}")
        return {"error": str(e), "fallback": True}


@app.get("/")
async def root():
    """Root endpoint"""
    return {
        "service": "Clisonix Excel Service",
        "version": "2.0.0",
        "excel_available": EXCEL_AVAILABLE,
        "pandas_available": PANDAS_AVAILABLE,
        "endpoints": [
            "/health",
            "/api/excel/generate",
            "/api/excel/validate",
            "/api/excel/parse",
            "/api/excel/formulas",
            "/api/excel/office-scripts"
        ]
    }


@app.get("/health")
async def health():
    """Health check"""
    return {
        "status": "healthy",
        "service": "excel-service",
        "excel_available": EXCEL_AVAILABLE,
        "pandas_available": PANDAS_AVAILABLE,
        "timestamp": datetime.now().isoformat()
    }


@app.get("/api/excel/generate")
async def generate_excel(
    title: str = "Clisonix Data Export",
    include_metrics: bool = True
):
    """Generate Excel file with system data"""
    if not EXCEL_AVAILABLE:
        return JSONResponse(
            status_code=503,
            content={"error": "Excel generation not available", "install": "pip install openpyxl"}
        )
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Export"
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        
        # Title row
        ws.merge_cells('A1:D1')
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=16)
        
        # Headers
        headers = ["Metric", "Value", "Status", "Timestamp"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Fetch metrics from Core API
        if include_metrics:
            metrics_data = await fetch_from_core("/api/metrics")
            row = 4
            for key, value in metrics_data.items():
                if key != "timestamp":
                    ws.cell(row=row, column=1, value=key)
                    ws.cell(row=row, column=2, value=str(value))
                    ws.cell(row=row, column=3, value="✓ OK")
                    ws.cell(row=row, column=4, value=datetime.now().strftime("%Y-%m-%d %H:%M"))
                    row += 1
        
        # Auto-width
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 20
        
        # Save
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"clisonix_export_{timestamp}.xlsx"
        filepath = EXCEL_DIR / filename
        wb.save(str(filepath))
        
        with open(filepath, 'rb') as f:
            content = f.read()
        
        logger.info(f"✓ Excel generated: {filename}")
        
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Excel generation failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/excel/validate")
async def validate_excel(file: UploadFile = File(...)):
    """Validate uploaded Excel file structure"""
    if not EXCEL_AVAILABLE:
        return JSONResponse(
            status_code=503,
            content={"error": "Excel validation not available"}
        )
    
    try:
        # Save uploaded file
        filepath = UPLOADS_DIR / file.filename
        with open(filepath, 'wb') as f:
            content = await file.read()
            f.write(content)
        
        # Load and validate
        wb = load_workbook(str(filepath))
        
        validation = {
            "filename": file.filename,
            "valid": True,
            "sheets": [],
            "total_rows": 0,
            "total_columns": 0,
            "issues": []
        }
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_info = {
                "name": sheet_name,
                "rows": ws.max_row,
                "columns": ws.max_column,
                "has_data": ws.max_row > 0 and ws.max_column > 0
            }
            validation["sheets"].append(sheet_info)
            validation["total_rows"] += ws.max_row
            validation["total_columns"] = max(validation["total_columns"], ws.max_column)
        
        # Cleanup
        filepath.unlink()
        
        return validation
        
    except Exception as e:
        logger.error(f"Excel validation failed: {e}")
        return {
            "filename": file.filename,
            "valid": False,
            "error": str(e)
        }


@app.post("/api/excel/parse")
async def parse_excel(file: UploadFile = File(...), sheet: str = None):
    """Parse Excel file and return JSON data"""
    if not EXCEL_AVAILABLE:
        return JSONResponse(
            status_code=503,
            content={"error": "Excel parsing not available"}
        )
    
    try:
        filepath = UPLOADS_DIR / file.filename
        with open(filepath, 'wb') as f:
            content = await file.read()
            f.write(content)
        
        wb = load_workbook(str(filepath))
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        
        # Parse data
        data = []
        headers = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                headers = [str(cell) if cell else f"Column_{i}" for i, cell in enumerate(row)]
            else:
                row_dict = {}
                for col_idx, cell in enumerate(row):
                    if col_idx < len(headers):
                        row_dict[headers[col_idx]] = cell
                if any(row_dict.values()):
                    data.append(row_dict)
        
        filepath.unlink()
        
        return {
            "filename": file.filename,
            "sheet": ws.title,
            "headers": headers,
            "rows": len(data),
            "data": data[:100]  # Limit to first 100 rows
        }
        
    except Exception as e:
        logger.error(f"Excel parsing failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/excel/formulas")
async def get_formulas():
    """List supported Excel formulas and =PY() integration"""
    return {
        "supported_formulas": [
            {"name": "=PY()", "description": "Execute Python code in Excel", "example": '=PY("import pandas as pd; result = df.sum()")'},
            {"name": "=CLISONIX.STATUS()", "description": "Get system status", "example": "=CLISONIX.STATUS()"},
            {"name": "=CLISONIX.METRICS()", "description": "Get live metrics", "example": "=CLISONIX.METRICS()"},
            {"name": "=CLISONIX.API(endpoint)", "description": "Call API endpoint", "example": '=CLISONIX.API("/api/status")'}
        ],
        "office_scripts": [
            {"name": "refreshData", "description": "Refresh all data connections"},
            {"name": "exportToPDF", "description": "Export current sheet to PDF"},
            {"name": "syncWithAPI", "description": "Sync data with Clisonix API"}
        ]
    }


@app.get("/api/excel/office-scripts")
async def get_office_scripts():
    """Get Office Scripts templates"""
    scripts = [
        {
            "name": "refreshClisonixData",
            "description": "Refresh data from Clisonix API",
            "script": """
async function main(workbook: ExcelScript.Workbook) {
    const sheet = workbook.getActiveWorksheet();
    const response = await fetch('https://clisonix.com/api/metrics');
    const data = await response.json();
    
    let row = 2;
    for (const [key, value] of Object.entries(data)) {
        sheet.getRange(\`A\${row}\`).setValue(key);
        sheet.getRange(\`B\${row}\`).setValue(String(value));
        row++;
    }
}
"""
        },
        {
            "name": "exportReport",
            "description": "Export formatted report",
            "script": """
async function main(workbook: ExcelScript.Workbook) {
    const sheet = workbook.getActiveWorksheet();
    const range = sheet.getUsedRange();
    range.getFormat().autofitColumns();
    console.log('Report exported successfully');
}
"""
        }
    ]
    return {"scripts": scripts}


@app.post("/api/excel/run-script")
async def run_office_script(script_name: str):
    """Trigger Office Script execution (placeholder)"""
    return {
        "status": "queued",
        "script": script_name,
        "message": "Office Scripts run in Excel client, not server-side",
        "timestamp": datetime.now().isoformat()
    }


@app.get("/api/excel/templates")
async def get_templates():
    """List available Excel templates"""
    return {
        "templates": [
            {"id": "metrics", "name": "Metrics Dashboard", "description": "System metrics with charts"},
            {"id": "sla", "name": "SLA Report", "description": "SLA tracking and compliance"},
            {"id": "api-catalog", "name": "API Catalog", "description": "All 71 API endpoints documented"},
            {"id": "functions", "name": "Functions Registry", "description": "Python and Office Script functions"}
        ]
    }


if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8002))
    logger.info(f"Starting Excel Service on port {port}")
    uvicorn.run(app, host="0.0.0.0", port=port)
