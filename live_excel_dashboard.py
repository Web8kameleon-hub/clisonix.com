"""
🔴 LIVE EXCEL DASHBOARD - Jo Template e Ngrirë!
Ky Excel përditësohet AUTOMATIKISHT çdo 30 sekonda me të dhëna reale nga API-të

Përdorimi:
  python live_excel_dashboard.py

Excel do të hapet dhe do të përditësohet LIVE!
"""
import requests
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
import threading
import os

# Config
EXCEL_PATH = 'c:/Users/Admin/Desktop/LIVE_API_DASHBOARD.xlsx'
BASE_URL = 'https://clisonix.com'
REFRESH_INTERVAL = 30  # sekonda

# Styles
wrap_align = Alignment(wrap_text=True, vertical='center', horizontal='left')
center_align = Alignment(wrap_text=True, vertical='center', horizontal='center')
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True, size=11)
online_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
offline_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
degraded_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
live_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
live_font = Font(color='FFFFFF', bold=True, size=14)

# API Endpoints to monitor
ENDPOINTS = [
    {'name': 'Reporting Dashboard', 'endpoint': '/api/reporting/dashboard', 'method': 'GET', 'folder': 'Reporting'},
    {'name': 'System Status', 'endpoint': '/api/system-status', 'method': 'GET', 'folder': 'Health'},
    {'name': 'Core Health', 'endpoint': '/health', 'method': 'GET', 'folder': 'Health'},
    {'name': 'Docker Containers', 'endpoint': '/api/reporting/docker-containers', 'method': 'GET', 'folder': 'Docker'},
    {'name': 'Docker Stats', 'endpoint': '/api/reporting/docker-stats', 'method': 'GET', 'folder': 'Docker'},
    {'name': 'Excel Export', 'endpoint': '/api/reporting/export-excel', 'method': 'GET', 'folder': 'Reporting'},
    {'name': 'PPTX Export', 'endpoint': '/api/reporting/export-pptx', 'method': 'GET', 'folder': 'Reporting'},
]

def check_endpoint(endpoint_info):
    """Kontrollon një endpoint dhe kthen statusin"""
    url = BASE_URL + endpoint_info['endpoint']
    start_time = time.time()
    
    try:
        response = requests.get(url, timeout=10)
        response_time = int((time.time() - start_time) * 1000)
        
        if response.status_code == 200:
            status = 'ONLINE'
            try:
                data = response.json()
                sample = str(data)[:100] + '...' if len(str(data)) > 100 else str(data)
            except:
                sample = f'Binary ({len(response.content)} bytes)'
        elif response.status_code >= 500:
            status = 'OFFLINE'
            sample = f'Error {response.status_code}'
        else:
            status = 'DEGRADED'
            sample = f'Status {response.status_code}'
            
    except requests.exceptions.Timeout:
        status = 'TIMEOUT'
        response_time = 10000
        sample = 'Request timeout'
    except Exception as e:
        status = 'OFFLINE'
        response_time = 0
        sample = str(e)[:50]
    
    return {
        **endpoint_info,
        'status': status,
        'response_time': response_time,
        'sample': sample,
        'last_check': datetime.now().strftime('%H:%M:%S'),
        'url': url
    }

def create_live_excel():
    """Krijon Excel LIVE me të dhëna aktuale"""
    print(f'\n🔄 Përditësoj Excel... {datetime.now().strftime("%H:%M:%S")}')
    
    # Kontrollo të gjitha endpoints
    results = []
    for ep in ENDPOINTS:
        result = check_endpoint(ep)
        results.append(result)
        status_icon = '✅' if result['status'] == 'ONLINE' else '⚠️' if result['status'] == 'DEGRADED' else '❌'
        print(f"  {status_icon} {result['name']}: {result['status']} ({result['response_time']}ms)")
    
    # Krijo workbook
    wb = Workbook()
    
    # === SHEET 1: LIVE STATUS ===
    ws = wb.active
    ws.title = 'LIVE STATUS'
    
    # LIVE indicator
    ws.merge_cells('A1:H1')
    live_cell = ws['A1']
    live_cell.value = f'🔴 LIVE - Përditësuar: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} - Auto-refresh çdo {REFRESH_INTERVAL}s'
    live_cell.fill = live_fill
    live_cell.font = live_font
    live_cell.alignment = center_align
    
    # Statistics row
    online_count = sum(1 for r in results if r['status'] == 'ONLINE')
    offline_count = sum(1 for r in results if r['status'] == 'OFFLINE')
    degraded_count = sum(1 for r in results if r['status'] in ['DEGRADED', 'TIMEOUT'])
    avg_response = sum(r['response_time'] for r in results) // len(results) if results else 0
    
    ws.merge_cells('A2:H2')
    stats_cell = ws['A2']
    stats_cell.value = f'📊 Online: {online_count} | ⚠️ Degraded: {degraded_count} | ❌ Offline: {offline_count} | ⚡ Avg Response: {avg_response}ms'
    stats_cell.font = Font(bold=True, size=12)
    stats_cell.alignment = center_align
    
    # Headers
    headers = ['#', 'API Name', 'Endpoint', 'Method', 'Status', 'Response (ms)', 'Last Check', 'Sample Response']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Data rows
    for row_idx, result in enumerate(results, 4):
        ws.cell(row=row_idx, column=1, value=row_idx - 3).border = border
        ws.cell(row=row_idx, column=2, value=result['name']).border = border
        ws.cell(row=row_idx, column=3, value=result['endpoint']).border = border
        ws.cell(row=row_idx, column=4, value=result['method']).border = border
        
        # Status me ngjyrë
        status_cell = ws.cell(row=row_idx, column=5, value=result['status'])
        status_cell.border = border
        status_cell.alignment = center_align
        if result['status'] == 'ONLINE':
            status_cell.fill = online_fill
        elif result['status'] == 'OFFLINE':
            status_cell.fill = offline_fill
        else:
            status_cell.fill = degraded_fill
        
        # Response time me ngjyrë
        rt_cell = ws.cell(row=row_idx, column=6, value=result['response_time'])
        rt_cell.border = border
        rt_cell.alignment = center_align
        if result['response_time'] < 500:
            rt_cell.fill = online_fill
        elif result['response_time'] < 2000:
            rt_cell.fill = degraded_fill
        else:
            rt_cell.fill = offline_fill
        
        ws.cell(row=row_idx, column=7, value=result['last_check']).border = border
        ws.cell(row=row_idx, column=8, value=result['sample']).border = border
        ws.cell(row=row_idx, column=8).alignment = wrap_align
    
    # Column widths
    widths = [5, 25, 35, 10, 12, 15, 12, 60]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # === SHEET 2: STATISTICS ===
    ws2 = wb.create_sheet('STATISTICS')
    
    ws2['A1'] = 'Category'
    ws2['B1'] = 'Count'
    ws2['A2'] = 'Online'
    ws2['B2'] = online_count
    ws2['A3'] = 'Degraded'
    ws2['B3'] = degraded_count
    ws2['A4'] = 'Offline'
    ws2['B4'] = offline_count
    
    # Pie chart
    pie = PieChart()
    labels = Reference(ws2, min_col=1, min_row=2, max_row=4)
    data = Reference(ws2, min_col=2, min_row=1, max_row=4)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "API Status Distribution"
    ws2.add_chart(pie, "D1")
    
    # Response times for bar chart
    ws2['A7'] = 'API'
    ws2['B7'] = 'Response (ms)'
    for idx, result in enumerate(results, 8):
        ws2.cell(row=idx, column=1, value=result['name'])
        ws2.cell(row=idx, column=2, value=result['response_time'])
    
    # Bar chart
    bar = BarChart()
    bar.title = "Response Times (ms)"
    data = Reference(ws2, min_col=2, min_row=7, max_row=7 + len(results))
    cats = Reference(ws2, min_col=1, min_row=8, max_row=7 + len(results))
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.shape = 4
    ws2.add_chart(bar, "D15")
    
    # === SHEET 3: HISTORY LOG ===
    ws3 = wb.create_sheet('HISTORY')
    ws3['A1'] = 'Timestamp'
    ws3['B1'] = 'Total Online'
    ws3['C1'] = 'Avg Response'
    ws3.cell(row=1, column=1).font = header_font
    ws3.cell(row=1, column=1).fill = header_fill
    ws3.cell(row=1, column=2).font = header_font
    ws3.cell(row=1, column=2).fill = header_fill
    ws3.cell(row=1, column=3).font = header_font
    ws3.cell(row=1, column=3).fill = header_fill
    
    # Add current data to history
    ws3['A2'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws3['B2'] = online_count
    ws3['C2'] = avg_response
    
    # Ruaj
    wb.save(EXCEL_PATH)
    print(f'💾 Ruajtur: {EXCEL_PATH}')
    return results

def run_continuous():
    """Ekzekuton përditësimin e vazhdueshëm"""
    print('=' * 60)
    print('🔴 LIVE EXCEL DASHBOARD - Clisonix API Monitor')
    print('=' * 60)
    print(f'📁 File: {EXCEL_PATH}')
    print(f'🔄 Refresh: çdo {REFRESH_INTERVAL} sekonda')
    print(f'🌐 Server: {BASE_URL}')
    print('=' * 60)
    print('Ctrl+C për të ndaluar...\n')
    
    # Hap Excel herën e parë
    create_live_excel()
    os.startfile(EXCEL_PATH)
    
    # Loop i vazhdueshëm
    while True:
        time.sleep(REFRESH_INTERVAL)
        try:
            create_live_excel()
        except PermissionError:
            print('⚠️ Excel është hapur - mbyll Excel për të përditësuar ose prit...')
        except Exception as e:
            print(f'❌ Gabim: {e}')

if __name__ == '__main__':
    try:
        run_continuous()
    except KeyboardInterrupt:
        print('\n\n👋 Dashboard u ndal. Mirupafshim!')
