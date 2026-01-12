"""
🎯 CLISONIX API REQUEST GENERATOR
Excel Template me formulas =PY() për gjenerim automatik të:
- cURL commands
- Python snippets  
- Body validation
- Headers formatting

PËRDORIM:
1. python api_request_generator.py
2. Excel 365 do të ekzekutojë Python dhe gjenerojë automatikisht

KËRKON: Microsoft 365 me Python in Excel të aktivizuar
"""

import json
import os
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.comments import Comment
except ImportError:
    print("❌ pip install openpyxl")
    exit(1)


def create_api_generator_template(output_path: str = None):
    """
    Krijon Excel template me formulas për API request generation.
    """
    base_path = Path(__file__).parent
    output_path = output_path or str(base_path / "Clisonix_API_Generator.xlsx")
    
    wb = Workbook()
    
    # ========== SHEET 1: API Requests ==========
    ws = wb.active
    ws.title = "API_Requests"
    
    # Headers
    headers = [
        "Row_ID",      # A
        "Folder",      # B
        "Method",      # C
        "Endpoint",    # D
        "Description", # E
        "cURL",        # F - Formula =PY()
        "Python",      # G - Formula =PY()
        "Headers",     # H
        "Body",        # I
        "Token",       # J
    ]
    
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
    
    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 45
    ws.column_dimensions['F'].width = 70
    ws.column_dimensions['G'].width = 70
    ws.column_dimensions['H'].width = 35
    ws.column_dimensions['I'].width = 45
    ws.column_dimensions['J'].width = 25
    
    # Token header në J2
    ws['J2'] = "YOUR_API_TOKEN"
    ws['J2'].fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    ws['J2'].font = Font(bold=True, color='000000')
    ws['J2'].alignment = Alignment(horizontal='center')
    
    # Ngarko API endpoints
    collection_path = base_path / "clisonix-ultra-mega-collection.json"
    if not collection_path.exists():
        collection_path = base_path / "clisonix-postman-collection.json"
    
    endpoints = []
    if collection_path.exists():
        with open(collection_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            
        def extract_endpoints(items, folder_name="Root"):
            for item in items:
                if 'item' in item:
                    extract_endpoints(item['item'], item.get('name', 'Unknown'))
                elif 'request' in item:
                    req = item['request']
                    method = req.get('method', 'GET')
                    
                    url = req.get('url', {})
                    if isinstance(url, str):
                        endpoint = url
                    else:
                        raw = url.get('raw', '')
                        path = url.get('path', [])
                        if path:
                            endpoint = '/' + '/'.join(path)
                        else:
                            endpoint = raw
                    
                    # Extract body
                    body = ""
                    if 'body' in req:
                        body_obj = req['body']
                        if body_obj.get('mode') == 'raw':
                            body = body_obj.get('raw', '')
                    
                    # Extract headers
                    headers_list = req.get('header', [])
                    headers_dict = {}
                    for h in headers_list:
                        if isinstance(h, dict):
                            headers_dict[h.get('key', '')] = h.get('value', '')
                    
                    endpoints.append({
                        'folder': folder_name,
                        'method': method,
                        'endpoint': endpoint,
                        'name': item.get('name', ''),
                        'body': body,
                        'headers': json.dumps(headers_dict) if headers_dict else ''
                    })
        
        if 'item' in data:
            extract_endpoints(data['item'])
    
    # Method colors
    method_colors = {
        'GET': '2ECC71',
        'POST': '3498DB',
        'PUT': 'F39C12',
        'PATCH': 'E67E22',
        'DELETE': 'E74C3C'
    }
    
    code_font = Font(name='Consolas', size=9)
    wrap_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Shto endpoints
    for idx, ep in enumerate(endpoints, 2):
        method = ep['method'].upper()
        row = idx
        
        # Row ID
        ws.cell(row=row, column=1, value=idx-1).alignment = Alignment(horizontal='center')
        
        # Folder
        ws.cell(row=row, column=2, value=ep['folder']).alignment = wrap_align
        
        # Method me ngjyrë
        method_cell = ws.cell(row=row, column=3, value=method)
        color = method_colors.get(method, 'CCCCCC')
        method_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        method_cell.font = Font(bold=True, color='FFFFFF')
        method_cell.alignment = Alignment(horizontal='center')
        
        # Endpoint
        ws.cell(row=row, column=4, value=ep['endpoint']).alignment = wrap_align
        ws.cell(row=row, column=4).font = code_font
        
        # Description
        ws.cell(row=row, column=5, value=ep['name']).alignment = wrap_align
        
        # cURL - Formulë statike (jo =PY() sepse s'funksionon pa cloud)
        curl_base = "https://api.clisonix.com"
        curl_cmd = f'curl -X {method} "{curl_base}{ep["endpoint"]}" -H "Authorization: Bearer $TOKEN"'
        if ep['body'] and method in ['POST', 'PUT', 'PATCH']:
            curl_cmd += f" -H 'Content-Type: application/json' -d '{ep['body'][:100]}...'"
        curl_cell = ws.cell(row=row, column=6, value=curl_cmd)
        curl_cell.font = code_font
        curl_cell.alignment = wrap_align
        
        # Python - Snippet
        python_snippet = f"""import requests

response = requests.{method.lower()}(
    "{curl_base}{ep['endpoint']}",
    headers={{"Authorization": f"Bearer {{token}}"}}
)
print(response.json())"""
        python_cell = ws.cell(row=row, column=7, value=python_snippet)
        python_cell.font = code_font
        python_cell.alignment = wrap_align
        
        # Headers
        if ep['headers']:
            ws.cell(row=row, column=8, value=ep['headers']).font = code_font
        
        # Body
        if ep['body']:
            body_preview = ep['body'][:500] + "..." if len(ep['body']) > 500 else ep['body']
            ws.cell(row=row, column=9, value=body_preview).font = code_font
            ws.cell(row=row, column=9).alignment = wrap_align
        
        # Apply border
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = thin_border
    
    # Freeze panes
    ws.freeze_panes = 'A2'
    
    # ========== SHEET 2: PY Formulas (Reference) ==========
    ws2 = wb.create_sheet("PY_Formulas")
    ws2['A1'] = "🐍 PYTHON IN EXCEL FORMULAS"
    ws2['A1'].font = Font(bold=True, size=16)
    
    ws2['A3'] = "Formula Name"
    ws2['B3'] = "Formula Code"
    ws2['C3'] = "Description"
    
    for col in ['A', 'B', 'C']:
        ws2[f'{col}3'].font = Font(bold=True)
        ws2[f'{col}3'].fill = header_fill
        ws2[f'{col}3'].font = header_font
    
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 80
    ws2.column_dimensions['C'].width = 40
    
    # cURL Formula Example
    curl_formula = '=PY("endpoint = xl(\'D2\'); method = xl(\'C2\'); f\'curl -X {method} https://api.clisonix.com{endpoint}\'")'
    ws2['A4'] = "cURL Generator"
    ws2['B4'] = curl_formula
    ws2['B4'].font = code_font
    ws2['B4'].alignment = wrap_align
    ws2['C4'] = "Gjeneron cURL command nga method dhe endpoint"
    
    # Python Formula Example
    python_formula = '=PY("endpoint = xl(\'D2\'); method = xl(\'C2\').lower(); f\'requests.{method}(https://api.clisonix.com{endpoint})\'")'
    ws2['A5'] = "Python Generator"
    ws2['B5'] = python_formula
    ws2['B5'].font = code_font
    ws2['B5'].alignment = wrap_align
    ws2['C5'] = "Gjeneron Python requests snippet"
    
    # Body Validator
    body_formula = '=PY("import json; body = xl(\'I2\'); \'Valid\' if not body or json.loads(str(body)) else \'Check\'")'
    ws2['A6'] = "Body Validator"
    ws2['B6'] = body_formula
    ws2['B6'].font = code_font
    ws2['B6'].alignment = wrap_align
    ws2['C6'] = "Validon JSON body"
    
    # Headers Builder
    headers_formula = '=PY("import json; token = xl(\'J2\'); json.dumps({\'Authorization\': f\'Bearer {token}\', \'Content-Type\': \'application/json\'})")'
    ws2['A7'] = "Headers Builder"
    ws2['B7'] = headers_formula
    ws2['B7'].font = code_font
    ws2['B7'].alignment = wrap_align
    ws2['C7'] = "Ndërton headers me token"
    
    # Full Request Generator
    full_formula = '=PY("method = xl(\'C2\'); ep = xl(\'D2\'); token = xl(\'J2\'); f\'curl -X {method} https://api.clisonix.com{ep} -H Authorization: Bearer {token}\'")'
    ws2['A8'] = "Full Request"
    ws2['B8'] = full_formula
    ws2['B8'].font = code_font
    ws2['B8'].alignment = wrap_align
    ws2['C8'] = "Request i plotë me token"
    
    # ========== SHEET 3: Instructions ==========
    ws3 = wb.create_sheet("Instructions")
    ws3['A1'] = "📚 SI TË PËRDORESH PYTHON IN EXCEL"
    ws3['A1'].font = Font(bold=True, size=18, color='1F4E79')
    
    instructions = [
        "",
        "1️⃣ KËRKESAT:",
        "   • Microsoft 365 (Office 365) subscription",
        "   • Python in Excel i aktivizuar (Settings → Python in Excel)",
        "",
        "2️⃣ SI FUNKSIONON:",
        "   • Formulas =PY() ekzekutojnë Python në cloud",
        "   • Rezultati shfaqet në qelizë",
        "   • xl() lexon vlera nga qelizat e tjera",
        "",
        "3️⃣ SHEMBULL I THJESHTË:",
        '   =PY("2 + 2")  → 4',
        '   =PY("xl(\'A1\') * 2")  → Dyfishon vlerën e A1',
        "",
        "4️⃣ PËR API REQUESTS:",
        "   • Vendos token-in në J2",
        "   • Kopjo formulën nga PY_Formulas sheet",
        "   • Paste në kolonën F (cURL) ose G (Python)",
        "",
        "5️⃣ SHËNIM:",
        "   • Ky file përmban snippets të gatshme",
        "   • Formulas =PY() duhet të kopjohen manualisht",
        "   • ose përdor macros VBA për automatizim",
        "",
        "6️⃣ API GENERATOR WORKFLOW:",
        "   1. Hap sheet API_Requests",
        "   2. Vendos token në qelizën J2",
        "   3. cURL commands janë në kolonën F",
        "   4. Python snippets janë në kolonën G",
        "   5. Copy-paste për ta përdorur"
    ]
    
    for i, line in enumerate(instructions, 3):
        ws3[f'A{i}'] = line
        ws3[f'A{i}'].font = Font(size=11)
    
    ws3.column_dimensions['A'].width = 80
    
    # Save
    wb.save(output_path)
    return output_path, len(endpoints)


def main():
    print("🎯 Clisonix API Request Generator")
    print("=" * 50)
    
    output_path, count = create_api_generator_template()
    
    print(f"✅ Excel Template Created: {output_path}")
    print(f"📊 {count} API endpoints loaded")
    print(f"📋 Sheets: API_Requests, PY_Formulas, Instructions")
    print(f"")
    print(f"🐍 Python In Excel Features:")
    print(f"   • cURL commands në kolonën F")
    print(f"   • Python snippets në kolonën G")
    print(f"   • Formulas =PY() në sheet 'PY_Formulas'")
    print(f"")
    print(f"📂 Opening in Excel...")
    
    # Hap Excel
    os.startfile(output_path)
    
    print(f"✅ Done!")


if __name__ == "__main__":
    main()
